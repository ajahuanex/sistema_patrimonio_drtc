import os
import tempfile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.files.storage import default_storage
from django.conf import settings
from django.db import models
from django.core.paginator import Paginator
from .models import Catalogo
from .utils import importar_catalogo_desde_excel, validar_estructura_catalogo
from .forms import CatalogoForm


@login_required
@permission_required('catalogo.view_catalogo', raise_exception=True)
def descargar_plantilla_catalogo(request):
    """Vista para descargar plantilla de ejemplo para importación de catálogo"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Catálogo"
    
    # Encabezados
    headers = ['CATÁLOGO', 'Denominación', 'Grupo', 'Clase', 'Resolución', 'Estado']
    ws.append(headers)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos de ejemplo
    ejemplos = [
        ['04220001', 'TRACTOR AGRICOLA', '04 AGRICOLA Y PESQUERO', '22 EQUIPO', 'R.D. N° 001-2020', 'ACTIVO'],
        ['05220002', 'COMPUTADORA PERSONAL', '05 EQUIPAMIENTO', '22 EQUIPO', 'R.D. N° 002-2020', 'ACTIVO'],
        ['06220003', 'ESCRITORIO DE MADERA', '06 MOBILIARIO', '22 EQUIPO', 'R.D. N° 003-2020', 'ACTIVO'],
        ['07220004', 'VEHICULO AUTOMOVIL', '07 TRANSPORTE', '22 EQUIPO', 'R.D. N° 004-2020', 'ACTIVO'],
        ['08220005', 'IMPRESORA LASER', '08 MAQUINARIA', '22 EQUIPO', 'R.D. N° 005-2020', 'EXCLUIDO'],
    ]
    
    for ejemplo in ejemplos:
        ws.append(ejemplo)
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15,  # CATÁLOGO
        'B': 40,  # Denominación
        'C': 30,  # Grupo
        'D': 20,  # Clase
        'E': 25,  # Resolución
        'F': 12,  # Estado
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Agregar instrucciones en una hoja separada
    ws_instrucciones = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR CATÁLOGO"],
        [""],
        ["1. ESTRUCTURA DEL ARCHIVO"],
        ["   - El archivo debe contener las siguientes columnas (en cualquier orden):"],
        ["     * CATÁLOGO: Código único de 8 dígitos"],
        ["     * Denominación: Nombre del bien (máximo 200 caracteres)"],
        ["     * Grupo: Grupo del catálogo (ej: 04 AGRICOLA Y PESQUERO)"],
        ["     * Clase: Clase del catálogo (ej: 22 EQUIPO)"],
        ["     * Resolución: Resolución que aprueba el bien"],
        ["     * Estado: ACTIVO o EXCLUIDO"],
        [""],
        ["2. REGLAS DE VALIDACIÓN"],
        ["   - Los códigos de catálogo deben ser únicos"],
        ["   - Las denominaciones deben ser únicas"],
        ["   - El código debe tener exactamente 8 dígitos"],
        ["   - El estado solo puede ser ACTIVO o EXCLUIDO"],
        [""],
        ["3. PROCESO DE IMPORTACIÓN"],
        ["   - Elimine estas filas de ejemplo antes de importar"],
        ["   - Complete sus datos en la hoja 'Plantilla Catálogo'"],
        ["   - Guarde el archivo en formato .xlsx o .xls"],
        ["   - Use el botón 'Validar' antes de importar"],
        ["   - Si la validación es exitosa, proceda con la importación"],
        [""],
        ["4. ACTUALIZACIÓN DE REGISTROS EXISTENTES"],
        ["   - Si marca 'Actualizar registros existentes':"],
        ["     Los registros con códigos existentes serán actualizados"],
        ["   - Si no marca la opción:"],
        ["     Los registros con códigos existentes serán omitidos"],
        [""],
        ["5. EJEMPLOS DE CÓDIGOS"],
        ["   - 04220001: Grupo 04, Clase 22, Correlativo 0001"],
        ["   - 05220002: Grupo 05, Clase 22, Correlativo 0002"],
        ["   - Los primeros 2 dígitos indican el grupo"],
        ["   - Los siguientes 2 dígitos indican la clase"],
        ["   - Los últimos 4 dígitos son el correlativo"],
    ]
    
    for row_data in instrucciones:
        ws_instrucciones.append(row_data)
    
    # Estilo para el título de instrucciones
    ws_instrucciones['A1'].font = Font(bold=True, size=14, color="366092")
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"plantilla_catalogo_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@permission_required('catalogo.add_catalogo', raise_exception=True)
def importar_catalogo_view(request):
    """Vista para importar catálogo desde Excel"""
    if request.method == 'POST':
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect('catalogo:importar')
        
        archivo = request.FILES['archivo_excel']
        actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'
        
        # Validar extensión
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
            return redirect('catalogo:importar')
        
        # Guardar archivo temporalmente
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                for chunk in archivo.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name
            
            # Procesar archivo
            resultado = importar_catalogo_desde_excel(
                temp_path, 
                actualizar_existentes,
                usuario=request.user,
                archivo_nombre=archivo.name,
                permitir_duplicados_denominacion=True  # Permitir denominaciones duplicadas
            )
            
            # Limpiar archivo temporal
            os.unlink(temp_path)
            
            # Mostrar resultados
            if resultado['exito']:
                messages.success(request, f"Importación exitosa: {resultado['resumen']}")
                
                # Mostrar advertencias si las hay
                for warning in resultado['warnings']:
                    messages.warning(request, warning)
                
                # Mostrar observaciones si las hay
                if resultado['total_observaciones'] > 0:
                    messages.info(
                        request, 
                        f"Se registraron {resultado['total_observaciones']} observaciones durante la importación. "
                        f"Puedes revisarlas en el panel de administración."
                    )
            else:
                messages.error(request, "Error en la importación:")
                for error in resultado['errores']:
                    messages.error(request, error)
            
            return redirect('catalogo:importar')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('catalogo:importar')
    
    # GET request - mostrar formulario
    context = {
        'total_catalogos': Catalogo.objects.count(),
        'catalogos_activos': Catalogo.objects.filter(estado='ACTIVO').count(),
        'catalogos_excluidos': Catalogo.objects.filter(estado='EXCLUIDO').count(),
    }
    
    return render(request, 'catalogo/importar.html', context)


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def validar_archivo_catalogo(request):
    """API endpoint para validar estructura del archivo Excel"""
    if 'archivo' not in request.FILES:
        return JsonResponse({
            'error': 'No se proporcionó archivo'
        }, status=400)
    
    archivo = request.FILES['archivo']
    
    try:
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            for chunk in archivo.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name
        
        # Validar estructura
        resultado = validar_estructura_catalogo(temp_path)
        
        # Limpiar archivo temporal
        os.unlink(temp_path)
        
        return JsonResponse(resultado)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error al validar archivo: {str(e)}'
        }, status=500)


@login_required
def lista_catalogo_view(request):
    """Vista para listar el catálogo"""
    catalogos = Catalogo.objects.all().order_by('codigo')
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    grupo_filtro = request.GET.get('grupo')
    busqueda = request.GET.get('q')
    
    if estado_filtro:
        catalogos = catalogos.filter(estado=estado_filtro)
    
    if grupo_filtro:
        catalogos = catalogos.filter(grupo__icontains=grupo_filtro)
    
    if busqueda:
        catalogos = catalogos.filter(
            models.Q(codigo__icontains=busqueda) |
            models.Q(denominacion__icontains=busqueda)
        )
    
    # Paginación
    paginator = Paginator(catalogos, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener grupos únicos para filtro (solo activos, no eliminados)
    grupos = Catalogo.objects.values_list('grupo', flat=True).distinct().order_by('grupo')
    
    context = {
        'catalogos': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'grupos': grupos,
        'estado_filtro': estado_filtro,
        'grupo_filtro': grupo_filtro,
        'busqueda': busqueda,
    }
    
    return render(request, 'catalogo/lista.html', context)


@login_required
def buscar_catalogo_api(request):
    """API para buscar en el catálogo (para autocompletado)"""
    termino = request.GET.get('q', '')
    limite = int(request.GET.get('limit', 10))
    
    if len(termino) < 2:
        return JsonResponse({'resultados': []})
    
    catalogos = Catalogo.buscar_por_denominacion(termino)[:limite]
    
    resultados = []
    for catalogo in catalogos:
        resultados.append({
            'id': catalogo.id,
            'codigo': catalogo.codigo,
            'denominacion': catalogo.denominacion,
            'grupo': catalogo.grupo,
            'clase': catalogo.clase,
            'texto': f"{catalogo.codigo} - {catalogo.denominacion}"
        })
    
    return JsonResponse({'resultados': resultados})


@login_required
def estadisticas_catalogo_view(request):
    """Vista para mostrar estadísticas del catálogo"""
    from django.db.models import Count
    
    # Estadísticas generales
    total = Catalogo.objects.count()
    activos = Catalogo.objects.filter(estado='ACTIVO').count()
    excluidos = Catalogo.objects.filter(estado='EXCLUIDO').count()
    
    # Estadísticas por grupo
    por_grupo = Catalogo.objects.values('grupo').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    # Estadísticas por clase
    por_clase = Catalogo.objects.values('clase').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    context = {
        'total': total,
        'activos': activos,
        'excluidos': excluidos,
        'por_grupo': por_grupo,
        'por_clase': por_clase,
    }
    
    return render(request, 'catalogo/estadisticas.html', context)

@login_required
@permission_required('catalogo.add_catalogo', raise_exception=True)
def agregar_catalogo_view(request):
    """Vista para agregar un nuevo catálogo"""
    if request.method == 'POST':
        form = CatalogoForm(request.POST)
        if form.is_valid():
            catalogo = form.save()
            messages.success(request, f'Catálogo "{catalogo.codigo}" creado exitosamente.')
            return redirect('catalogo:detalle', pk=catalogo.pk)
    else:
        form = CatalogoForm()
    
    context = {
        'form': form,
        'titulo': 'Agregar Catálogo'
    }
    return render(request, 'catalogo/formulario.html', context)


@login_required
def detalle_catalogo_view(request, pk):
    """Vista para ver detalles de un catálogo"""
    catalogo = get_object_or_404(Catalogo, pk=pk)
    
    # Obtener bienes relacionados si existen
    bienes_count = 0
    try:
        from apps.bienes.models import BienPatrimonial
        bienes_count = BienPatrimonial.objects.filter(catalogo=catalogo).count()
    except ImportError:
        pass
    
    context = {
        'catalogo': catalogo,
        'bienes_count': bienes_count,
    }
    return render(request, 'catalogo/detalle.html', context)


@login_required
@permission_required('catalogo.change_catalogo', raise_exception=True)
def editar_catalogo_view(request, pk):
    """Vista para editar un catálogo"""
    catalogo = get_object_or_404(Catalogo, pk=pk)
    
    if request.method == 'POST':
        form = CatalogoForm(request.POST, instance=catalogo)
        if form.is_valid():
            catalogo = form.save()
            messages.success(request, f'Catálogo "{catalogo.codigo}" actualizado exitosamente.')
            return redirect('catalogo:detalle', pk=catalogo.pk)
    else:
        form = CatalogoForm(instance=catalogo)
    
    context = {
        'form': form,
        'catalogo': catalogo,
        'titulo': f'Editar Catálogo {catalogo.codigo}'
    }
    return render(request, 'catalogo/formulario.html', context)


@login_required
@permission_required('catalogo.view_catalogo', raise_exception=True)
def exportar_catalogo_view(request):
    """Vista para exportar catálogo a Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.utils import timezone
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo SBN"
    
    # Encabezados
    headers = ['CÓDIGO', 'DENOMINACIÓN', 'GRUPO', 'CLASE', 'RESOLUCIÓN', 'ESTADO']
    ws.append(headers)
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Datos
    catalogos = Catalogo.objects.all().order_by('codigo')
    for catalogo in catalogos:
        ws.append([
            catalogo.codigo,
            catalogo.denominacion,
            catalogo.grupo or '',
            catalogo.clase or '',
            catalogo.resolucion or '',
            catalogo.estado
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"catalogo_sbn_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@permission_required('catalogo.delete_catalogo', raise_exception=True)
@require_http_methods(["POST"])
def eliminar_catalogo_view(request, pk):
    """Vista para eliminar un catálogo via AJAX (soft delete)"""
    try:
        catalogo = get_object_or_404(Catalogo, pk=pk)
        
        # Verificar si tiene bienes asociados
        bienes_count = 0
        try:
            from apps.bienes.models import BienPatrimonial
            bienes_count = BienPatrimonial.objects.filter(catalogo=catalogo).count()
        except ImportError:
            pass
        
        if bienes_count > 0:
            return JsonResponse({
                'success': False,
                'errors': {
                    'general': [f'No se puede eliminar el catálogo "{catalogo.denominacion}" porque tiene {bienes_count} bienes asociados.']
                }
            })
        
        # Guardar información para el mensaje
        codigo_catalogo = catalogo.codigo
        denominacion_catalogo = catalogo.denominacion
        
        # Obtener motivo de eliminación si se proporciona
        deletion_reason = request.POST.get('deletion_reason', 'Eliminación desde interfaz web')
        
        # Usar soft delete
        catalogo.soft_delete(user=request.user, reason=deletion_reason)
        
        # Crear entrada en RecycleBin
        from apps.core.models import RecycleBin, RecycleBinConfig
        from django.contrib.contenttypes.models import ContentType
        from django.utils import timezone
        from datetime import timedelta
        
        content_type = ContentType.objects.get_for_model(Catalogo)
        config = RecycleBinConfig.get_config_for_module('catalogo')
        auto_delete_at = timezone.now() + timedelta(days=config.retention_days)
        
        RecycleBin.objects.create(
            content_type=content_type,
            object_id=catalogo.id,
            object_repr=f"{codigo_catalogo} - {denominacion_catalogo}",
            module_name='catalogo',
            deleted_by=request.user,
            deletion_reason=deletion_reason,
            auto_delete_at=auto_delete_at,
            original_data={
                'codigo': catalogo.codigo,
                'denominacion': catalogo.denominacion,
                'grupo': catalogo.grupo,
                'clase': catalogo.clase,
                'estado': catalogo.estado
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': f'El catálogo "{denominacion_catalogo}" ({codigo_catalogo}) ha sido movido a la papelera de reciclaje.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        })