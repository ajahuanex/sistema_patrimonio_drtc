import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from .models import Catalogo


class CatalogoImporter:
    """Clase para importar catálogo desde archivos Excel"""
    
    COLUMNAS_REQUERIDAS = [
        'CATLOGO',  # Nota: puede tener variaciones en el nombre
        'Denominación',
        'Grupo',
        'Clase',
        'Resolución',
        'Estado'
    ]
    
    COLUMNAS_ALTERNATIVAS = {
        'CATLOGO': ['CATALOGO', 'CÓDIGO', 'CODIGO'],
        'Denominación': ['DENOMINACION', 'DENOMINACIÓN BIEN', 'DENOMINACION BIEN'],
        'Estado': ['ESTADO']
    }
    
    def __init__(self):
        self.errores = []
        self.warnings = []
        self.registros_procesados = 0
        self.registros_creados = 0
        self.registros_actualizados = 0
    
    def validar_archivo(self, archivo_path):
        """Valida que el archivo Excel tenga la estructura correcta"""
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener la primera fila (encabezados)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Verificar columnas requeridas
            columnas_encontradas = {}
            for col_requerida in self.COLUMNAS_REQUERIDAS:
                encontrada = False
                
                # Buscar columna exacta
                for header in headers:
                    if header.upper() == col_requerida.upper():
                        columnas_encontradas[col_requerida] = header
                        encontrada = True
                        break
                
                # Buscar alternativas
                if not encontrada and col_requerida in self.COLUMNAS_ALTERNATIVAS:
                    for alternativa in self.COLUMNAS_ALTERNATIVAS[col_requerida]:
                        for header in headers:
                            if header.upper() == alternativa.upper():
                                columnas_encontradas[col_requerida] = header
                                encontrada = True
                                break
                        if encontrada:
                            break
                
                if not encontrada:
                    self.errores.append(f"Columna requerida no encontrada: {col_requerida}")
            
            if self.errores:
                return False, columnas_encontradas
            
            return True, columnas_encontradas
            
        except Exception as e:
            self.errores.append(f"Error al leer el archivo: {str(e)}")
            return False, {}
    
    def procesar_archivo(self, archivo_path, actualizar_existentes=False):
        """Procesa el archivo Excel e importa los datos"""
        # Validar archivo
        es_valido, columnas_map = self.validar_archivo(archivo_path)
        if not es_valido:
            return self.generar_reporte()
        
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener índices de columnas
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            indices_columnas = {}
            
            for col_requerida, header_encontrado in columnas_map.items():
                try:
                    indices_columnas[col_requerida] = headers.index(header_encontrado)
                except ValueError:
                    self.errores.append(f"Error al encontrar índice de columna: {header_encontrado}")
                    return self.generar_reporte()
            
            # Procesar filas de datos
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    try:
                        self.procesar_fila(row, indices_columnas, row_num, actualizar_existentes)
                    except Exception as e:
                        self.errores.append(f"Error en fila {row_num}: {str(e)}")
                        continue
            
        except Exception as e:
            self.errores.append(f"Error general al procesar archivo: {str(e)}")
        
        return self.generar_reporte()
    
    def procesar_fila(self, row, indices_columnas, row_num, actualizar_existentes):
        """Procesa una fila individual del Excel"""
        # Extraer datos de la fila
        datos = {}
        for col_name, col_index in indices_columnas.items():
            cell_value = row[col_index].value
            datos[col_name] = str(cell_value).strip() if cell_value else ''
        
        # Validar datos requeridos
        if not datos.get('CATLOGO') or not datos.get('Denominación'):
            self.warnings.append(f"Fila {row_num}: Código o denominación vacíos, omitida")
            return
        
        # Normalizar datos
        codigo = datos['CATLOGO']
        denominacion = datos['Denominación'].upper()
        grupo = datos.get('Grupo', '')
        clase = datos.get('Clase', '')
        resolucion = datos.get('Resolución', '')
        estado = datos.get('Estado', 'ACTIVO').upper()
        
        # Validar estado
        if estado not in ['ACTIVO', 'EXCLUIDO']:
            estado = 'ACTIVO'
            self.warnings.append(f"Fila {row_num}: Estado inválido, se asignó ACTIVO")
        
        # Verificar si ya existe
        catalogo_existente = None
        try:
            catalogo_existente = Catalogo.objects.get(codigo=codigo)
        except Catalogo.DoesNotExist:
            pass
        
        if catalogo_existente:
            if actualizar_existentes:
                # Actualizar existente
                catalogo_existente.denominacion = denominacion
                catalogo_existente.grupo = grupo
                catalogo_existente.clase = clase
                catalogo_existente.resolucion = resolucion
                catalogo_existente.estado = estado
                catalogo_existente.save()
                self.registros_actualizados += 1
            else:
                self.warnings.append(f"Fila {row_num}: Código {codigo} ya existe, omitido")
        else:
            # Crear nuevo
            try:
                Catalogo.objects.create(
                    codigo=codigo,
                    denominacion=denominacion,
                    grupo=grupo,
                    clase=clase,
                    resolucion=resolucion,
                    estado=estado
                )
                self.registros_creados += 1
            except Exception as e:
                self.errores.append(f"Fila {row_num}: Error al crear registro: {str(e)}")
                return
        
        self.registros_procesados += 1
    
    def generar_reporte(self):
        """Genera un reporte del proceso de importación"""
        return {
            'exito': len(self.errores) == 0,
            'registros_procesados': self.registros_procesados,
            'registros_creados': self.registros_creados,
            'registros_actualizados': self.registros_actualizados,
            'errores': self.errores,
            'warnings': self.warnings,
            'resumen': f"Procesados: {self.registros_procesados}, "
                      f"Creados: {self.registros_creados}, "
                      f"Actualizados: {self.registros_actualizados}, "
                      f"Errores: {len(self.errores)}, "
                      f"Advertencias: {len(self.warnings)}"
        }


def importar_catalogo_desde_excel(archivo_path, actualizar_existentes=False):
    """Función helper para importar catálogo"""
    importer = CatalogoImporter()
    return importer.procesar_archivo(archivo_path, actualizar_existentes)


def validar_estructura_catalogo(archivo_path):
    """Función helper para validar estructura del archivo"""
    importer = CatalogoImporter()
    es_valido, columnas_map = importer.validar_archivo(archivo_path)
    return {
        'es_valido': es_valido,
        'columnas_encontradas': columnas_map,
        'errores': importer.errores
    }