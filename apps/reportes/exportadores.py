import os
import csv
import json
from datetime import datetime
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from django.conf import settings
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)


class ExportadorBase:
    """Clase base para todos los exportadores"""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def _preparar_datos_bienes(self, queryset):
        """Prepara los datos de bienes para exportación"""
        datos = []
        
        for bien in queryset.select_related('catalogo', 'oficina'):
            datos.append({
                'codigo_patrimonial': bien.codigo_patrimonial,
                'codigo_interno': bien.codigo_interno or '',
                'denominacion': bien.catalogo.denominacion if bien.catalogo else '',
                'grupo': bien.catalogo.grupo if bien.catalogo else '',
                'clase': bien.catalogo.clase if bien.catalogo else '',
                'estado_bien': bien.get_estado_bien_display(),
                'marca': bien.marca or '',
                'modelo': bien.modelo or '',
                'color': bien.color or '',
                'serie': bien.serie or '',
                'dimension': bien.dimension or '',
                'placa': bien.placa or '',
                'matricula': bien.matricula or '',
                'nro_motor': bien.nro_motor or '',
                'nro_chasis': bien.nro_chasis or '',
                'oficina_codigo': bien.oficina.codigo if bien.oficina else '',
                'oficina_nombre': bien.oficina.nombre if bien.oficina else '',
                'responsable': bien.oficina.responsable if bien.oficina else '',
                'fecha_adquisicion': bien.fecha_adquisicion.strftime('%d/%m/%Y') if bien.fecha_adquisicion else '',
                'valor_adquisicion': bien.valor_adquisicion or 0,
                'observaciones': bien.observaciones or '',
                'url_qr': bien.url_qr or '',
                'fecha_registro': bien.created_at.strftime('%d/%m/%Y %H:%M') if bien.created_at else '',
            })
        
        return datos


class ExportadorExcel(ExportadorBase):
    """Exportador para archivos Excel (.xlsx)"""
    
    def exportar_bienes(self, queryset, archivo_salida=None):
        """
        Exporta bienes a Excel
        
        Args:
            queryset: QuerySet de bienes
            archivo_salida: Ruta del archivo de salida (opcional)
            
        Returns:
            str: Ruta del archivo generado
        """
        if not archivo_salida:
            archivo_salida = f'inventario_{self.timestamp}.xlsx'
        
        # Preparar datos
        datos = self._preparar_datos_bienes(queryset)
        
        if not datos:
            # Crear archivo vacío
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            ws.append(["No se encontraron bienes con los filtros aplicados"])
            wb.save(archivo_salida)
            return archivo_salida
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja principal de inventario
        ws_inventario = wb.active
        ws_inventario.title = "Inventario"
        
        # Agregar encabezados institucionales
        self._agregar_encabezado_institucional(ws_inventario)
        
        # Agregar datos
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_inventario.append(r)
        
        # Aplicar formato
        self._aplicar_formato_inventario(ws_inventario, len(datos))
        
        # Crear hoja de estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        self._crear_hoja_estadisticas(ws_stats, queryset)
        
        # Crear hoja de gráficos
        ws_graficos = wb.create_sheet("Gráficos")
        self._crear_graficos_excel(ws_graficos, queryset)
        
        # Guardar archivo
        wb.save(archivo_salida)
        
        logger.info(f"Archivo Excel generado: {archivo_salida} con {len(datos)} registros")
        return archivo_salida
    
    def exportar_estadisticas(self, estadisticas, archivo_salida=None):
        """
        Exporta estadísticas a Excel
        
        Args:
            estadisticas: Dict con estadísticas
            archivo_salida: Ruta del archivo de salida
            
        Returns:
            str: Ruta del archivo generado
        """
        if not archivo_salida:
            archivo_salida = f'estadisticas_{self.timestamp}.xlsx'
        
        wb = Workbook()
        
        # Hoja de resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        self._crear_hoja_resumen_estadisticas(ws_resumen, estadisticas)
        
        # Hoja por estado
        ws_estados = wb.create_sheet("Por Estado")
        self._crear_hoja_estados(ws_estados, estadisticas.get('por_estado', []))
        
        # Hoja por oficina
        ws_oficinas = wb.create_sheet("Por Oficina")
        self._crear_hoja_oficinas(ws_oficinas, estadisticas.get('por_oficina', []))
        
        # Hoja por grupo
        ws_grupos = wb.create_sheet("Por Grupo")
        self._crear_hoja_grupos(ws_grupos, estadisticas.get('por_grupo', []))
        
        wb.save(archivo_salida)
        
        logger.info(f"Archivo de estadísticas Excel generado: {archivo_salida}")
        return archivo_salida
    
    def _agregar_encabezado_institucional(self, ws):
        """Agrega encabezado institucional al Excel"""
        # Título principal
        ws.merge_cells('A1:T1')
        ws['A1'] = "DIRECCIÓN REGIONAL DE TRANSPORTES Y COMUNICACIONES - PUNO"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Subtítulo
        ws.merge_cells('A2:T2')
        ws['A2'] = "SISTEMA DE REGISTRO DE PATRIMONIO"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A3:T3')
        ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = Alignment(horizontal='center')
        
        # Línea en blanco
        ws.append([])
    
    def _aplicar_formato_inventario(self, ws, num_filas):
        """Aplica formato a la hoja de inventario"""
        # Colores
        color_encabezado = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        color_alternado = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Fuente para encabezados
        fuente_encabezado = Font(bold=True, color="FFFFFF")
        
        # Encontrar fila de encabezados (después del encabezado institucional)
        fila_encabezado = 5  # Ajustar según el encabezado institucional
        
        # Formatear encabezados
        for col in range(1, ws.max_column + 1):
            celda = ws.cell(row=fila_encabezado, column=col)
            celda.fill = color_encabezado
            celda.font = fuente_encabezado
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formatear filas de datos (alternadas)
        for fila in range(fila_encabezado + 1, fila_encabezado + num_filas + 1):
            if (fila - fila_encabezado) % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=fila, column=col).fill = color_alternado
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _crear_hoja_estadisticas(self, ws, queryset):
        """Crea hoja de estadísticas"""
        from .utils import FiltroAvanzado
        
        # Generar estadísticas
        filtro = FiltroAvanzado()
        estadisticas = filtro.obtener_estadisticas(queryset)
        
        # Título
        ws['A1'] = "ESTADÍSTICAS DEL INVENTARIO"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        fila_actual = 3
        
        # Estadísticas por estado
        ws[f'A{fila_actual}'] = "DISTRIBUCIÓN POR ESTADO"
        ws[f'A{fila_actual}'].font = Font(bold=True, size=12)
        fila_actual += 1
        
        ws[f'A{fila_actual}'] = "Estado"
        ws[f'B{fila_actual}'] = "Cantidad"
        ws[f'C{fila_actual}'] = "Porcentaje"
        
        # Aplicar formato a encabezados
        for col in ['A', 'B', 'C']:
            celda = ws[f'{col}{fila_actual}']
            celda.font = Font(bold=True)
            celda.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        fila_actual += 1
        
        for estado in estadisticas.get('por_estado', []):
            ws[f'A{fila_actual}'] = estado.get('estado_bien', 'N/A')
            ws[f'B{fila_actual}'] = estado.get('total', 0)
            ws[f'C{fila_actual}'] = f"{estado.get('porcentaje', 0):.1f}%"
            fila_actual += 1
        
        fila_actual += 2
        
        # Estadísticas por oficina (top 10)
        ws[f'A{fila_actual}'] = "TOP 10 OFICINAS"
        ws[f'A{fila_actual}'].font = Font(bold=True, size=12)
        fila_actual += 1
        
        ws[f'A{fila_actual}'] = "Oficina"
        ws[f'B{fila_actual}'] = "Cantidad"
        
        # Aplicar formato a encabezados
        for col in ['A', 'B']:
            celda = ws[f'{col}{fila_actual}']
            celda.font = Font(bold=True)
            celda.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        fila_actual += 1
        
        for oficina in estadisticas.get('por_oficina', [])[:10]:
            ws[f'A{fila_actual}'] = oficina.get('oficina__nombre', 'Sin nombre')
            ws[f'B{fila_actual}'] = oficina.get('total', 0)
            fila_actual += 1
    
    def _crear_graficos_excel(self, ws, queryset):
        """Crea gráficos en Excel"""
        from .utils import FiltroAvanzado
        
        # Generar estadísticas
        filtro = FiltroAvanzado()
        estadisticas = filtro.obtener_estadisticas(queryset)
        
        # Preparar datos para gráfico de estados
        estados_data = estadisticas.get('por_estado', [])
        if estados_data:
            # Crear datos en la hoja
            ws['A1'] = "Estado"
            ws['B1'] = "Cantidad"
            
            fila = 2
            for estado in estados_data:
                ws[f'A{fila}'] = estado.get('estado_bien', 'N/A')
                ws[f'B{fila}'] = estado.get('total', 0)
                fila += 1
            
            # Crear gráfico de torta
            pie_chart = PieChart()
            labels = Reference(ws, min_col=1, min_row=2, max_row=fila-1)
            data = Reference(ws, min_col=2, min_row=1, max_row=fila-1)
            
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            pie_chart.title = "Distribución por Estado"
            
            ws.add_chart(pie_chart, "D2")
    
    def _crear_hoja_resumen_estadisticas(self, ws, estadisticas):
        """Crea hoja de resumen de estadísticas"""
        ws['A1'] = "RESUMEN ESTADÍSTICO"
        ws['A1'].font = Font(bold=True, size=14)
        
        fila = 3
        ws[f'A{fila}'] = "Total de Bienes:"
        ws[f'B{fila}'] = estadisticas.get('total_bienes', 0)
        fila += 1
        
        # Agregar más estadísticas según disponibilidad
        valores = estadisticas.get('valores', {})
        if valores:
            ws[f'A{fila}'] = "Valor Total:"
            ws[f'B{fila}'] = f"S/ {valores.get('total_valor', 0):,.2f}"
            fila += 1
            
            ws[f'A{fila}'] = "Valor Promedio:"
            ws[f'B{fila}'] = f"S/ {valores.get('valor_promedio', 0):,.2f}"
            fila += 1
    
    def _crear_hoja_estados(self, ws, estados_data):
        """Crea hoja de estadísticas por estado"""
        ws['A1'] = "ESTADÍSTICAS POR ESTADO"
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A3'] = "Estado"
        ws['B3'] = "Cantidad"
        ws['C3'] = "Porcentaje"
        
        fila = 4
        for estado in estados_data:
            ws[f'A{fila}'] = estado.get('estado_bien', 'N/A')
            ws[f'B{fila}'] = estado.get('total', 0)
            ws[f'C{fila}'] = f"{estado.get('porcentaje', 0):.1f}%"
            fila += 1
    
    def _crear_hoja_oficinas(self, ws, oficinas_data):
        """Crea hoja de estadísticas por oficina"""
        ws['A1'] = "ESTADÍSTICAS POR OFICINA"
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A3'] = "Oficina"
        ws['B3'] = "Cantidad"
        ws['C3'] = "Porcentaje"
        
        fila = 4
        for oficina in oficinas_data:
            ws[f'A{fila}'] = oficina.get('oficina__nombre', 'Sin nombre')
            ws[f'B{fila}'] = oficina.get('total', 0)
            ws[f'C{fila}'] = f"{oficina.get('porcentaje', 0):.1f}%"
            fila += 1
    
    def _crear_hoja_grupos(self, ws, grupos_data):
        """Crea hoja de estadísticas por grupo"""
        ws['A1'] = "ESTADÍSTICAS POR GRUPO"
        ws['A1'].font = Font(bold=True, size=12)
        
        ws['A3'] = "Grupo"
        ws['B3'] = "Cantidad"
        ws['C3'] = "Porcentaje"
        
        fila = 4
        for grupo in grupos_data:
            ws[f'A{fila}'] = grupo.get('catalogo__grupo', 'Sin grupo')
            ws[f'B{fila}'] = grupo.get('total', 0)
            ws[f'C{fila}'] = f"{grupo.get('porcentaje', 0):.1f}%"
            fila += 1


class ExportadorCSV(ExportadorBase):
    """Exportador para archivos CSV"""
    
    def exportar_bienes(self, queryset, archivo_salida=None):
        """
        Exporta bienes a CSV
        
        Args:
            queryset: QuerySet de bienes
            archivo_salida: Ruta del archivo de salida
            
        Returns:
            str: Ruta del archivo generado
        """
        if not archivo_salida:
            archivo_salida = f'inventario_{self.timestamp}.csv'
        
        # Preparar datos
        datos = self._preparar_datos_bienes(queryset)
        
        if not datos:
            # Crear archivo vacío
            with open(archivo_salida, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["No se encontraron bienes con los filtros aplicados"])
            return archivo_salida
        
        # Escribir CSV
        with open(archivo_salida, 'w', newline='', encoding='utf-8-sig') as csvfile:
            fieldnames = datos[0].keys()
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            # Escribir encabezado institucional
            writer.writerow({fieldnames[0]: "DIRECCIÓN REGIONAL DE TRANSPORTES Y COMUNICACIONES - PUNO"})
            writer.writerow({fieldnames[0]: "SISTEMA DE REGISTRO DE PATRIMONIO"})
            writer.writerow({fieldnames[0]: f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"})
            writer.writerow({})  # Línea en blanco
            
            # Escribir encabezados
            writer.writeheader()
            
            # Escribir datos
            writer.writerows(datos)
        
        logger.info(f"Archivo CSV generado: {archivo_salida} con {len(datos)} registros")
        return archivo_salida


class ExportadorZPL(ExportadorBase):
    """Exportador para plantillas ZPL (Zebra Programming Language)"""
    
    def generar_plantilla_stickers(self, queryset, archivo_salida=None):
        """
        Genera plantilla ZPL para stickers
        
        Args:
            queryset: QuerySet de bienes
            archivo_salida: Ruta del archivo de salida
            
        Returns:
            str: Ruta del archivo generado
        """
        if not archivo_salida:
            archivo_salida = f'stickers_{self.timestamp}.zpl'
        
        # Configuración del sticker (en dots, 203 DPI)
        config = {
            'ancho_sticker': 400,  # ~2 pulgadas
            'alto_sticker': 300,   # ~1.5 pulgadas
            'margen': 20,
            'tamaño_qr': 100,
            'fuente_titulo': 'A',
            'tamaño_titulo': 'N',
            'fuente_texto': 'A',
            'tamaño_texto': 'N',
        }
        
        plantilla_zpl = []
        
        # Encabezado ZPL
        plantilla_zpl.append("^XA")  # Inicio de etiqueta
        
        for bien in queryset:
            # Inicio de nueva etiqueta
            plantilla_zpl.append("^XA")
            
            # Configurar origen
            plantilla_zpl.append("^LH0,0")
            
            # Código QR (esquina superior izquierda)
            if bien.qr_code:
                plantilla_zpl.append(f"^FO{config['margen']},{config['margen']}")
                plantilla_zpl.append(f"^BQN,2,{config['tamaño_qr']//10}")
                plantilla_zpl.append(f"^FDQA,{bien.qr_code}^FS")
            
            # Código patrimonial (título principal)
            x_texto = config['margen'] + config['tamaño_qr'] + 20
            y_texto = config['margen']
            
            plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
            plantilla_zpl.append(f"^{config['fuente_titulo']}{config['tamaño_titulo']}")
            plantilla_zpl.append(f"^FD{bien.codigo_patrimonial}^FS")
            
            # Denominación (truncada)
            y_texto += 40
            denominacion = bien.catalogo.denominacion if bien.catalogo else "Sin denominación"
            if len(denominacion) > 25:
                denominacion = denominacion[:22] + "..."
            
            plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
            plantilla_zpl.append(f"^{config['fuente_texto']}{config['tamaño_texto']}")
            plantilla_zpl.append(f"^FD{denominacion}^FS")
            
            # Oficina
            y_texto += 30
            oficina = bien.oficina.codigo if bien.oficina else "Sin oficina"
            
            plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
            plantilla_zpl.append(f"^{config['fuente_texto']}{config['tamaño_texto']}")
            plantilla_zpl.append(f"^FD{oficina}^FS")
            
            # Estado
            y_texto += 30
            estado = bien.get_estado_bien_display()
            
            plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
            plantilla_zpl.append(f"^{config['fuente_texto']}{config['tamaño_texto']}")
            plantilla_zpl.append(f"^FD{estado}^FS")
            
            # Marca/Modelo (si existe)
            if bien.marca or bien.modelo:
                y_texto += 30
                marca_modelo = f"{bien.marca or ''} {bien.modelo or ''}".strip()
                if len(marca_modelo) > 20:
                    marca_modelo = marca_modelo[:17] + "..."
                
                plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
                plantilla_zpl.append(f"^{config['fuente_texto']}{config['tamaño_texto']}")
                plantilla_zpl.append(f"^FD{marca_modelo}^FS")
            
            # Serie (si existe)
            if bien.serie:
                y_texto += 25
                serie = f"S/N: {bien.serie}"
                if len(serie) > 20:
                    serie = serie[:17] + "..."
                
                plantilla_zpl.append(f"^FO{x_texto},{y_texto}")
                plantilla_zpl.append(f"^{config['fuente_texto']}{config['tamaño_texto']}")
                plantilla_zpl.append(f"^FD{serie}^FS")
            
            # Línea separadora (opcional)
            plantilla_zpl.append(f"^FO{config['margen']},{config['alto_sticker'] - 40}")
            plantilla_zpl.append(f"^GB{config['ancho_sticker'] - 2*config['margen']},2,2^FS")
            
            # Fecha de generación (pequeña)
            plantilla_zpl.append(f"^FO{config['margen']},{config['alto_sticker'] - 30}")
            plantilla_zpl.append("^A0N,15,15")
            plantilla_zpl.append(f"^FD{datetime.now().strftime('%d/%m/%Y')}^FS")
            
            # Fin de etiqueta
            plantilla_zpl.append("^XZ")
        
        # Escribir archivo ZPL
        with open(archivo_salida, 'w', encoding='utf-8') as zpl_file:
            zpl_file.write('\n'.join(plantilla_zpl))
        
        logger.info(f"Plantilla ZPL generada: {archivo_salida} con {queryset.count()} stickers")
        return archivo_salida
    
    def generar_sticker_individual(self, bien, config=None):
        """
        Genera código ZPL para un sticker individual
        
        Args:
            bien: Instancia de BienPatrimonial
            config: Configuración personalizada del sticker
            
        Returns:
            str: Código ZPL para el sticker
        """
        if not config:
            config = {
                'ancho_sticker': 400,
                'alto_sticker': 300,
                'margen': 20,
                'tamaño_qr': 100,
            }
        
        zpl_lines = []
        
        # Inicio de etiqueta
        zpl_lines.append("^XA")
        zpl_lines.append("^LH0,0")
        
        # Código QR
        if bien.qr_code:
            zpl_lines.append(f"^FO{config['margen']},{config['margen']}")
            zpl_lines.append(f"^BQN,2,{config['tamaño_qr']//10}")
            zpl_lines.append(f"^FDQA,{bien.qr_code}^FS")
        
        # Código patrimonial
        x_texto = config['margen'] + config['tamaño_qr'] + 20
        zpl_lines.append(f"^FO{x_texto},{config['margen']}")
        zpl_lines.append("^AN")
        zpl_lines.append(f"^FD{bien.codigo_patrimonial}^FS")
        
        # Denominación
        denominacion = bien.catalogo.denominacion if bien.catalogo else "Sin denominación"
        if len(denominacion) > 25:
            denominacion = denominacion[:22] + "..."
        
        zpl_lines.append(f"^FO{x_texto},{config['margen'] + 40}")
        zpl_lines.append("^AN")
        zpl_lines.append(f"^FD{denominacion}^FS")
        
        # Fin de etiqueta
        zpl_lines.append("^XZ")
        
        return '\n'.join(zpl_lines)