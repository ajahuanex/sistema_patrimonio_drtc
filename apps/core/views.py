from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
import json

from .models import UserProfile, AuditLog, RecycleBin, RecycleBinConfig
from .permissions import role_required, permission_required_custom
from .utils import (
    create_user_with_profile, update_user_role, 
    deactivate_user, activate_user, validate_user_data,
    get_user_permissions_summary, get_audit_logs_for_user,
    RecycleBinService
)
from .filters import RecycleBinFilterForm, RecycleBinQuickFilters
from .forms import RestoreForm, PermanentDeleteForm, BulkOperationForm, QuickRestoreForm
from apps.oficinas.models import Oficina


@login_required
@role_required(['administrador'])
def user_list(request):
    """Lista de usuarios del sistema"""
    search_query = request.GET.get('search', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    
    users = User.objects.select_related('profile', 'profile__oficina').all()
    
    # Aplicar filtros
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if role_filter:
        users = users.filter(profile__role=role_filter)
    
    if status_filter == 'active':
        users = users.filter(profile__is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(profile__is_active=False)
    
    # Paginación
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'roles': UserProfile.ROLES,
    }
    
    return render(request, 'core/user_list.html', context)


@login_required
@role_required(['administrador'])
def user_create(request):
    """Crear nuevo usuario"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        role = request.POST.get('role', 'consulta')
        telefono = request.POST.get('telefono', '')
        cargo = request.POST.get('cargo', '')
        oficina_id = request.POST.get('oficina')
        
        # Validar datos
        errors = validate_user_data(username, email, role)
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'core/user_form.html', {
                'oficinas': Oficina.objects.filter(estado=True),
                'roles': UserProfile.ROLES,
                'form_data': request.POST
            })
        
        # Obtener oficina si se especificó
        oficina = None
        if oficina_id:
            try:
                oficina = Oficina.objects.get(id=oficina_id)
            except Oficina.DoesNotExist:
                pass
        
        # Crear usuario
        user, error = create_user_with_profile(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            password=password,
            role=role,
            telefono=telefono,
            cargo=cargo,
            oficina=oficina
        )
        
        if user:
            messages.success(request, f'Usuario {username} creado correctamente')
            return redirect('core:user_list')
        else:
            messages.error(request, f'Error al crear usuario: {error}')
    
    context = {
        'oficinas': Oficina.objects.filter(estado=True),
        'roles': UserProfile.ROLES,
    }
    
    return render(request, 'core/user_form.html', context)


@login_required
@role_required(['administrador'])
def user_edit(request, user_id):
    """Editar usuario existente"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        # Actualizar datos básicos del usuario
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        
        # Actualizar perfil
        profile = user.profile
        old_role = profile.role
        new_role = request.POST.get('role', profile.role)
        profile.telefono = request.POST.get('telefono', '')
        profile.cargo = request.POST.get('cargo', '')
        
        # Actualizar oficina
        oficina_id = request.POST.get('oficina')
        if oficina_id:
            try:
                profile.oficina = Oficina.objects.get(id=oficina_id)
            except Oficina.DoesNotExist:
                profile.oficina = None
        else:
            profile.oficina = None
        
        try:
            user.save()
            profile.save()
            
            # Actualizar rol si cambió
            if old_role != new_role:
                success, message = update_user_role(user, new_role)
                if success:
                    messages.success(request, f'Usuario actualizado. {message}')
                else:
                    messages.error(request, f'Error al actualizar rol: {message}')
            else:
                messages.success(request, 'Usuario actualizado correctamente')
            
            return redirect('core:user_list')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar usuario: {str(e)}')
    
    context = {
        'user_obj': user,
        'oficinas': Oficina.objects.filter(estado=True),
        'roles': UserProfile.ROLES,
    }
    
    return render(request, 'core/user_form.html', context)


@login_required
@role_required(['administrador'])
def user_detail(request, user_id):
    """Detalle de usuario con historial de auditoría"""
    user = get_object_or_404(User, id=user_id)
    permissions = get_user_permissions_summary(user)
    audit_logs = get_audit_logs_for_user(user, limit=20)
    
    context = {
        'user_obj': user,
        'permissions': permissions,
        'audit_logs': audit_logs,
    }
    
    return render(request, 'core/user_detail.html', context)


@login_required
@role_required(['administrador'])
@require_http_methods(["POST"])
def user_toggle_status(request, user_id):
    """Activar/desactivar usuario"""
    user = get_object_or_404(User, id=user_id)
    
    if user.profile.is_active:
        success, message = deactivate_user(user, request.user)
        action = 'desactivado'
    else:
        success, message = activate_user(user, request.user)
        action = 'activado'
    
    if success:
        messages.success(request, f'Usuario {action} correctamente')
    else:
        messages.error(request, f'Error: {message}')
    
    return redirect('core:user_list')


@login_required
@role_required(['administrador'])
@require_http_methods(["POST"])
def user_reset_password(request, user_id):
    """Resetear contraseña de usuario"""
    user = get_object_or_404(User, id=user_id)
    new_password = request.POST.get('new_password')
    
    if not new_password:
        messages.error(request, 'Debe proporcionar una nueva contraseña')
        return redirect('core:user_detail', user_id=user_id)
    
    try:
        user.set_password(new_password)
        user.save()
        
        # Registrar en auditoría
        AuditLog.objects.create(
            user=request.user,
            action='update',
            model_name='User',
            object_id=str(user.id),
            object_repr=f"{user.username}",
            changes={'password': 'reset'}
        )
        
        messages.success(request, 'Contraseña actualizada correctamente')
        
    except Exception as e:
        messages.error(request, f'Error al actualizar contraseña: {str(e)}')
    
    return redirect('core:user_detail', user_id=user_id)


@login_required
@role_required(['administrador', 'auditor'])
def audit_logs(request):
    """Lista de logs de auditoría"""
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    model_filter = request.GET.get('model', '')
    
    logs = AuditLog.objects.select_related('user').all()
    
    # Aplicar filtros
    if user_filter:
        logs = logs.filter(user__username__icontains=user_filter)
    
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    if model_filter:
        logs = logs.filter(model_name__icontains=model_filter)
    
    # Paginación
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'user_filter': user_filter,
        'action_filter': action_filter,
        'model_filter': model_filter,
        'actions': AuditLog.ACTION_CHOICES,
    }
    
    return render(request, 'core/audit_logs.html', context)


# ============================================================================
# DELETION AUDIT REPORTS VIEWS
# ============================================================================


@login_required
@permission_required_custom('can_view_deletion_audit_logs')
def deletion_audit_reports(request):
    """
    Vista principal de reportes de auditoría de eliminaciones con filtros avanzados.
    Incluye estadísticas, gráficos de tendencias y detección de patrones sospechosos.
    """
    from apps.core.models import DeletionAuditLog
    from datetime import datetime, timedelta
    from django.db.models import Count, Q
    from django.db.models.functions import TruncDate, TruncHour
    
    # Obtener parámetros de filtro
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    module_filter = request.GET.get('module', '')
    success_filter = request.GET.get('success', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Queryset base con optimizaciones
    queryset = DeletionAuditLog.objects.select_related(
        'user', 'content_type', 'recycle_bin_entry'
    ).all()
    
    # Aplicar filtros
    if user_filter:
        queryset = queryset.filter(user__username__icontains=user_filter)
    
    if action_filter:
        queryset = queryset.filter(action=action_filter)
    
    if module_filter:
        queryset = queryset.filter(module_name=module_filter)
    
    if success_filter:
        queryset = queryset.filter(success=(success_filter == 'true'))
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            # Incluir todo el día
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            queryset = queryset.filter(timestamp__lte=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        queryset = queryset.filter(
            Q(object_repr__icontains=search_query) |
            Q(reason__icontains=search_query) |
            Q(error_message__icontains=search_query)
        )
    
    # Estadísticas generales
    total_logs = queryset.count()
    successful_operations = queryset.filter(success=True).count()
    failed_operations = queryset.filter(success=False).count()
    
    # Estadísticas por acción
    action_stats = queryset.values('action').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Estadísticas por módulo
    module_stats = queryset.values('module_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Estadísticas por usuario (top 10)
    user_stats = queryset.values('user__username').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Datos para gráfico de tendencias (últimos 30 días)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    trend_data = queryset.filter(
        timestamp__gte=thirty_days_ago
    ).annotate(
        date=TruncDate('timestamp')
    ).values('date', 'action').annotate(
        count=Count('id')
    ).order_by('date')
    
    # Organizar datos de tendencias por acción
    trend_by_action = {}
    for item in trend_data:
        action = item['action']
        if action not in trend_by_action:
            trend_by_action[action] = []
        trend_by_action[action].append({
            'date': item['date'].strftime('%Y-%m-%d'),
            'count': item['count']
        })
    
    # Detección de patrones sospechosos
    suspicious_patterns = _detect_suspicious_patterns(queryset)
    
    # Paginación
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preparar contexto
    context = {
        'page_obj': page_obj,
        'total_logs': total_logs,
        'successful_operations': successful_operations,
        'failed_operations': failed_operations,
        'action_stats': action_stats,
        'module_stats': module_stats,
        'user_stats': user_stats,
        'trend_by_action': trend_by_action,
        'suspicious_patterns': suspicious_patterns,
        'user_filter': user_filter,
        'action_filter': action_filter,
        'module_filter': module_filter,
        'success_filter': success_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'action_choices': DeletionAuditLog.ACTION_CHOICES,
        'is_admin': request.user.profile.is_administrador,
    }
    
    return render(request, 'core/deletion_audit_reports.html', context)


def _detect_suspicious_patterns(queryset):
    """
    Detecta patrones sospechosos en los logs de auditoría.
    
    Args:
        queryset: QuerySet de DeletionAuditLog
        
    Returns:
        list: Lista de patrones sospechosos detectados
    """
    from datetime import timedelta
    from django.db.models import Count
    
    patterns = []
    now = timezone.now()
    
    # Patrón 1: Múltiples eliminaciones permanentes en corto tiempo
    one_hour_ago = now - timedelta(hours=1)
    recent_permanent_deletes = queryset.filter(
        action='permanent_delete',
        timestamp__gte=one_hour_ago
    ).values('user__username').annotate(
        count=Count('id')
    ).filter(count__gte=5)
    
    for item in recent_permanent_deletes:
        patterns.append({
            'type': 'high_permanent_deletes',
            'severity': 'high',
            'message': f"Usuario {item['user__username']} realizó {item['count']} eliminaciones permanentes en la última hora",
            'user': item['user__username'],
            'count': item['count'],
            'icon': '⚠️'
        })
    
    # Patrón 2: Múltiples intentos fallidos
    recent_failures = queryset.filter(
        success=False,
        timestamp__gte=one_hour_ago
    ).values('user__username').annotate(
        count=Count('id')
    ).filter(count__gte=3)
    
    for item in recent_failures:
        patterns.append({
            'type': 'multiple_failures',
            'severity': 'medium',
            'message': f"Usuario {item['user__username']} tuvo {item['count']} operaciones fallidas en la última hora",
            'user': item['user__username'],
            'count': item['count'],
            'icon': '⚡'
        })
    
    # Patrón 3: Eliminaciones masivas en un módulo específico
    one_day_ago = now - timedelta(days=1)
    massive_deletes = queryset.filter(
        action__in=['soft_delete', 'permanent_delete'],
        timestamp__gte=one_day_ago
    ).values('module_name', 'user__username').annotate(
        count=Count('id')
    ).filter(count__gte=20)
    
    for item in massive_deletes:
        patterns.append({
            'type': 'massive_deletes',
            'severity': 'high',
            'message': f"Usuario {item['user__username']} eliminó {item['count']} elementos del módulo {item['module_name']} en las últimas 24 horas",
            'user': item['user__username'],
            'module': item['module_name'],
            'count': item['count'],
            'icon': '🔥'
        })
    
    # Patrón 4: Actividad fuera de horario laboral (10pm - 6am)
    from django.db.models.functions import ExtractHour
    
    off_hours_activity = queryset.filter(
        timestamp__gte=one_day_ago
    ).annotate(
        hour=ExtractHour('timestamp')
    ).filter(
        Q(hour__gte=22) | Q(hour__lt=6)
    ).values('user__username').annotate(
        count=Count('id')
    ).filter(count__gte=5)
    
    for item in off_hours_activity:
        patterns.append({
            'type': 'off_hours_activity',
            'severity': 'low',
            'message': f"Usuario {item['user__username']} realizó {item['count']} operaciones fuera de horario laboral",
            'user': item['user__username'],
            'count': item['count'],
            'icon': '🌙'
        })
    
    # Patrón 5: Restauraciones seguidas de eliminaciones permanentes
    suspicious_restore_delete = []
    users_with_activity = queryset.filter(
        timestamp__gte=one_day_ago
    ).values_list('user__username', flat=True).distinct()
    
    for username in users_with_activity:
        user_logs = queryset.filter(
            user__username=username,
            timestamp__gte=one_day_ago
        ).order_by('timestamp')
        
        restore_count = 0
        permanent_delete_count = 0
        
        for log in user_logs:
            if log.action == 'restore':
                restore_count += 1
            elif log.action == 'permanent_delete' and restore_count > 0:
                permanent_delete_count += 1
        
        if restore_count >= 3 and permanent_delete_count >= 3:
            patterns.append({
                'type': 'restore_then_delete',
                'severity': 'medium',
                'message': f"Usuario {username} restauró {restore_count} elementos y luego eliminó permanentemente {permanent_delete_count} en las últimas 24 horas",
                'user': username,
                'restore_count': restore_count,
                'delete_count': permanent_delete_count,
                'icon': '🔄'
            })
    
    return patterns


@login_required
@permission_required_custom('can_view_deletion_audit_logs')
def deletion_audit_export(request):
    """
    Exporta los logs de auditoría de eliminaciones a PDF o Excel según el formato solicitado.
    Aplica los mismos filtros que la vista principal.
    """
    from apps.core.models import DeletionAuditLog
    from datetime import datetime
    from django.db.models import Q
    import io
    from django.http import HttpResponse
    
    # Obtener formato de exportación
    export_format = request.GET.get('format', 'excel')
    
    # Obtener parámetros de filtro (mismos que la vista principal)
    user_filter = request.GET.get('user', '')
    action_filter = request.GET.get('action', '')
    module_filter = request.GET.get('module', '')
    success_filter = request.GET.get('success', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search_query = request.GET.get('search', '')
    
    # Queryset base
    queryset = DeletionAuditLog.objects.select_related(
        'user', 'content_type', 'recycle_bin_entry'
    ).all()
    
    # Aplicar filtros (mismo código que la vista principal)
    if user_filter:
        queryset = queryset.filter(user__username__icontains=user_filter)
    
    if action_filter:
        queryset = queryset.filter(action=action_filter)
    
    if module_filter:
        queryset = queryset.filter(module_name=module_filter)
    
    if success_filter:
        queryset = queryset.filter(success=(success_filter == 'true'))
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(timestamp__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            date_to_obj = date_to_obj.replace(hour=23, minute=59, second=59)
            queryset = queryset.filter(timestamp__lte=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        queryset = queryset.filter(
            Q(object_repr__icontains=search_query) |
            Q(reason__icontains=search_query) |
            Q(error_message__icontains=search_query)
        )
    
    # Limitar a 10000 registros para evitar problemas de memoria
    queryset = queryset[:10000]
    
    if export_format == 'excel':
        return _export_to_excel(queryset, request)
    elif export_format == 'pdf':
        return _export_to_pdf(queryset, request)
    else:
        messages.error(request, 'Formato de exportación no válido')
        return redirect('core:deletion_audit_reports')


def _export_to_excel(queryset, request):
    """
    Exporta los logs de auditoría a formato Excel.
    
    Args:
        queryset: QuerySet de DeletionAuditLog
        request: HttpRequest
        
    Returns:
        HttpResponse: Archivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import io
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría de Eliminaciones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Fecha/Hora', 'Usuario', 'Acción', 'Módulo', 'Objeto',
            'Exitoso', 'Motivo', 'IP', 'Código Seguridad'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row_num, log in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=2, value=log.user.username)
            ws.cell(row=row_num, column=3, value=log.get_action_display())
            ws.cell(row=row_num, column=4, value=log.module_name)
            ws.cell(row=row_num, column=5, value=log.object_repr)
            ws.cell(row=row_num, column=6, value='Sí' if log.success else 'No')
            ws.cell(row=row_num, column=7, value=log.reason or '')
            ws.cell(row=row_num, column=8, value=log.ip_address or '')
            ws.cell(row=row_num, column=9, value='Sí' if log.security_code_used else 'No')
        
        # Ajustar ancho de columnas
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20
        
        # Agregar hoja de estadísticas
        ws_stats = wb.create_sheet("Estadísticas")
        ws_stats.cell(row=1, column=1, value="Estadísticas de Auditoría").font = Font(bold=True, size=14)
        
        from django.db.models import Count
        
        # Estadísticas por acción
        ws_stats.cell(row=3, column=1, value="Por Acción").font = Font(bold=True)
        action_stats = queryset.values('action').annotate(count=Count('id')).order_by('-count')
        row = 4
        for stat in action_stats:
            ws_stats.cell(row=row, column=1, value=dict(queryset.model.ACTION_CHOICES).get(stat['action'], stat['action']))
            ws_stats.cell(row=row, column=2, value=stat['count'])
            row += 1
        
        # Estadísticas por módulo
        ws_stats.cell(row=row + 1, column=1, value="Por Módulo").font = Font(bold=True)
        module_stats = queryset.values('module_name').annotate(count=Count('id')).order_by('-count')
        row += 2
        for stat in module_stats:
            ws_stats.cell(row=row, column=1, value=stat['module_name'])
            ws_stats.cell(row=row, column=2, value=stat['count'])
            row += 1
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="auditoria_eliminaciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada. No se puede exportar a Excel.')
        return redirect('core:deletion_audit_reports')
    except Exception as e:
        messages.error(request, f'Error al exportar a Excel: {str(e)}')
        return redirect('core:deletion_audit_reports')


def _export_to_pdf(queryset, request):
    """
    Exporta los logs de auditoría a formato PDF.
    
    Args:
        queryset: QuerySet de DeletionAuditLog
        request: HttpRequest
        
    Returns:
        HttpResponse: Archivo PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from django.http import HttpResponse
        from django.db.models import Count
        import io
        
        # Crear buffer
        buffer = io.BytesIO()
        
        # Crear documento en orientación horizontal
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Título
        title = Paragraph("Reporte de Auditoría de Eliminaciones", title_style)
        elements.append(title)
        
        # Información del reporte
        info_style = styles['Normal']
        info_text = f"Generado por: {request.user.get_full_name() or request.user.username}<br/>"
        info_text += f"Fecha: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        info_text += f"Total de registros: {queryset.count()}"
        elements.append(Paragraph(info_text, info_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas
        elements.append(Paragraph("Estadísticas Generales", heading_style))
        
        stats_data = [
            ['Métrica', 'Valor'],
            ['Total de operaciones', str(queryset.count())],
            ['Operaciones exitosas', str(queryset.filter(success=True).count())],
            ['Operaciones fallidas', str(queryset.filter(success=False).count())],
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas por acción
        elements.append(Paragraph("Operaciones por Acción", heading_style))
        action_stats = queryset.values('action').annotate(count=Count('id')).order_by('-count')
        
        action_data = [['Acción', 'Cantidad']]
        for stat in action_stats:
            action_name = dict(queryset.model.ACTION_CHOICES).get(stat['action'], stat['action'])
            action_data.append([action_name, str(stat['count'])])
        
        action_table = Table(action_data, colWidths=[3*inch, 2*inch])
        action_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(action_table)
        elements.append(PageBreak())
        
        # Detalle de logs (limitado a primeros 100)
        elements.append(Paragraph("Detalle de Operaciones (Primeras 100)", heading_style))
        
        log_data = [['Fecha/Hora', 'Usuario', 'Acción', 'Módulo', 'Objeto', 'Exitoso']]
        
        for log in queryset[:100]:
            log_data.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M'),
                log.user.username[:15],
                log.get_action_display()[:20],
                log.module_name[:15],
                log.object_repr[:30],
                'Sí' if log.success else 'No'
            ])
        
        log_table = Table(log_data, colWidths=[1.3*inch, 1*inch, 1.5*inch, 1*inch, 2*inch, 0.7*inch])
        log_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(log_table)
        
        # Construir PDF
        doc.build(elements)
        
        # Obtener valor del buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Crear respuesta
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="auditoria_eliminaciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(pdf)
        
        return response
        
    except ImportError:
        messages.error(request, 'La librería reportlab no está instalada. No se puede exportar a PDF.')
        return redirect('core:deletion_audit_reports')
    except Exception as e:
        messages.error(request, f'Error al exportar a PDF: {str(e)}')
        return redirect('core:deletion_audit_reports')


@login_required
@permission_required_custom('can_view_deletion_audit_logs')
def deletion_audit_detail(request, log_id):
    """
    Vista de detalle de un log de auditoría específico.
    Muestra toda la información incluyendo snapshots y metadatos.
    """
    from apps.core.models import DeletionAuditLog
    
    log = get_object_or_404(DeletionAuditLog, id=log_id)
    
    # Obtener logs relacionados del mismo objeto
    related_logs = DeletionAuditLog.objects.filter(
        content_type=log.content_type,
        object_id=log.object_id
    ).exclude(id=log.id).order_by('-timestamp')[:10]
    
    context = {
        'log': log,
        'related_logs': related_logs,
        'is_admin': request.user.profile.is_administrador,
    }
    
    return render(request, 'core/deletion_audit_detail.html', context)


# API Views para el frontend React
@csrf_exempt
@login_required
@permission_required_custom('can_manage_users')
def api_users_list(request):
    """API para listar usuarios"""
    if request.method == 'GET':
        users = User.objects.select_related('profile', 'profile__oficina').all()
        
        users_data = []
        for user in users:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.profile.role if hasattr(user, 'profile') else 'consulta',
                'role_display': user.profile.get_role_display() if hasattr(user, 'profile') else 'Consulta',
                'is_active': user.profile.is_active if hasattr(user, 'profile') else True,
                'oficina': user.profile.oficina.nombre if hasattr(user, 'profile') and user.profile.oficina else None,
                'cargo': user.profile.cargo if hasattr(user, 'profile') else '',
                'created_at': user.date_joined.isoformat(),
            })
        
        return JsonResponse({'users': users_data})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
@login_required
@permission_required_custom('can_manage_users')
def api_user_create(request):
    """API para crear usuario"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar datos
            errors = validate_user_data(
                data.get('username'), 
                data.get('email'), 
                data.get('role', 'consulta')
            )
            
            if errors:
                return JsonResponse({'error': errors}, status=400)
            
            # Obtener oficina si se especificó
            oficina = None
            if data.get('oficina_id'):
                try:
                    oficina = Oficina.objects.get(id=data['oficina_id'])
                except Oficina.DoesNotExist:
                    pass
            
            # Crear usuario
            user, error = create_user_with_profile(
                username=data.get('username'),
                email=data.get('email', ''),
                first_name=data.get('first_name', ''),
                last_name=data.get('last_name', ''),
                password=data.get('password'),
                role=data.get('role', 'consulta'),
                telefono=data.get('telefono', ''),
                cargo=data.get('cargo', ''),
                oficina=oficina
            )
            
            if user:
                return JsonResponse({
                    'success': True,
                    'message': f'Usuario {user.username} creado correctamente',
                    'user_id': user.id
                })
            else:
                return JsonResponse({'error': error}, status=400)
                
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ============================================================================
# RECYCLE BIN VIEWS
# ============================================================================


@login_required
@permission_required_custom('can_view_recycle_bin')
def recycle_bin_list(request):
    """
    Vista principal de la papelera de reciclaje con paginación y filtros avanzados.
    Implementa segregación de datos por usuario según permisos.
    """
    # Verificar permisos de visualización
    can_view_all = request.user.profile.can_view_all_recycle_items()
    
    # Queryset base con optimizaciones
    queryset = RecycleBin.objects.select_related(
        'deleted_by', 'restored_by', 'content_type'
    )
    
    # Segregación de datos por usuario según permisos
    if not can_view_all:
        # Usuarios sin permiso de ver todo solo ven sus propias eliminaciones
        queryset = queryset.filter(deleted_by=request.user)
    
    # Crear formulario de filtros con datos GET
    filter_form = RecycleBinFilterForm(request.GET or None)
    
    # Aplicar filtros si el formulario es válido
    if filter_form.is_valid():
        queryset = filter_form.apply_filters(queryset, request.user)
    else:
        # Si no hay filtros, mostrar solo elementos no restaurados por defecto
        if not request.GET.get('status'):
            queryset = queryset.filter(restored_at__isnull=True)
    
    # Ordenar por fecha de eliminación (más recientes primero)
    queryset = queryset.order_by('-deleted_at')
    
    # Paginación
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Obtener estadísticas
    stats = RecycleBinService.get_recycle_bin_stats(request.user)
    
    # Obtener filtros rápidos
    quick_filters = {
        'expiring_soon': RecycleBinQuickFilters.get_expiring_soon(
            RecycleBin.objects.all()
        ).count(),
        'expired': RecycleBinQuickFilters.get_expired(
            RecycleBin.objects.all()
        ).count(),
        'my_deletions': RecycleBinQuickFilters.get_by_user(
            RecycleBin.objects.filter(restored_at__isnull=True),
            request.user
        ).count(),
    }
    
    # Resumen de filtros activos
    active_filters_summary = []
    active_filters_count = 0
    if filter_form.is_valid():
        active_filters_summary = filter_form.get_active_filters_summary()
        active_filters_count = filter_form.get_active_filters_count()
    
    # Permisos del usuario para el template
    user_permissions = {
        'can_view_all': can_view_all,
        'can_restore_items': request.user.profile.can_restore_items(),
        'can_restore_own': request.user.profile.can_restore_own_items(),
        'can_restore_others': request.user.profile.can_restore_others_items(),
        'can_permanent_delete': request.user.profile.can_permanent_delete(),
        'can_bulk_restore': request.user.profile.can_bulk_restore(),
        'can_bulk_delete': request.user.profile.can_bulk_permanent_delete(),
    }
    
    context = {
        'page_obj': page_obj,
        'stats': stats,
        'filter_form': filter_form,
        'quick_filters': quick_filters,
        'active_filters_summary': active_filters_summary,
        'active_filters_count': active_filters_count,
        'is_admin': request.user.profile.is_administrador,
        'user_permissions': user_permissions,
    }
    
    return render(request, 'core/recycle_bin_list.html', context)


@login_required
@permission_required_custom('can_view_recycle_bin')
def recycle_bin_detail(request, entry_id):
    """
    Vista de detalle de un elemento en la papelera con vista previa de datos.
    Implementa segregación de datos por usuario según permisos.
    """
    entry = get_object_or_404(RecycleBin, id=entry_id)
    
    # Verificar permisos de visualización con segregación de datos
    can_view_all = request.user.profile.can_view_all_recycle_items()
    if not can_view_all and entry.deleted_by != request.user:
        messages.error(request, 'No tiene permisos para ver este elemento')
        return redirect('core:recycle_bin_list')
    
    # Obtener el objeto original si existe
    original_object = entry.content_object
    
    # Verificar permisos de restauración
    can_restore_own = request.user.profile.can_restore_own_items()
    can_restore_others = request.user.profile.can_restore_others_items()
    
    # Determinar si puede restaurar este elemento específico
    can_restore = False
    if entry.deleted_by == request.user and can_restore_own:
        can_restore = True
    elif entry.deleted_by != request.user and can_restore_others:
        can_restore = True
    
    # Verificar conflictos potenciales
    restore_conflicts = None
    if original_object and can_restore:
        restore_conflicts = RecycleBinService._check_restore_conflicts(original_object)
    
    # Permisos del usuario para el template
    user_permissions = {
        'can_restore': can_restore,
        'can_permanent_delete': request.user.profile.can_permanent_delete(),
        'can_view_audit_logs': request.user.profile.can_view_deletion_audit_logs(),
    }
    
    context = {
        'entry': entry,
        'original_object': original_object,
        'can_restore': can_restore,
        'restore_conflicts': restore_conflicts,
        'is_admin': request.user.profile.is_administrador,
        'original_data': entry.original_data,
        'user_permissions': user_permissions,
    }
    
    return render(request, 'core/recycle_bin_detail.html', context)


@login_required
@permission_required_custom('can_restore_items')
@require_http_methods(["GET", "POST"])
def recycle_bin_restore(request, entry_id):
    """
    Vista para restaurar un elemento individual desde la papelera con formulario de validación.
    Implementa validación de permisos granular según el propietario del elemento.
    """
    entry = get_object_or_404(RecycleBin, id=entry_id)
    
    # Verificar permisos granulares de restauración
    can_restore_own = request.user.profile.can_restore_own_items()
    can_restore_others = request.user.profile.can_restore_others_items()
    
    # Validar permisos según el propietario
    if entry.deleted_by == request.user:
        if not can_restore_own:
            messages.error(request, 'No tiene permisos para restaurar sus propios elementos')
            return redirect('core:recycle_bin_list')
    else:
        if not can_restore_others:
            messages.error(request, 'No tiene permisos para restaurar elementos de otros usuarios')
            return redirect('core:recycle_bin_list')
    
    if request.method == 'POST':
        # Obtener contexto para auditoría
        ip_address = request.META.get('REMOTE_ADDR')
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        # Si viene de un POST simple (sin formulario completo), usar QuickRestoreForm
        if 'quick_restore' in request.POST:
            form = QuickRestoreForm(request.POST, user=request.user)
            if form.is_valid():
                # Intentar restaurar
                success, message, restored_object = RecycleBinService.restore_object(
                    entry, 
                    request.user,
                    ip_address=ip_address,
                    user_agent=user_agent
                )
                
                if success:
                    messages.success(request, message)
                    return _redirect_to_restored_object(entry, restored_object)
                else:
                    messages.error(request, message)
                    return redirect('core:recycle_bin_detail', entry_id=entry_id)
            else:
                for error in form.errors.values():
                    messages.error(request, error)
                return redirect('core:recycle_bin_detail', entry_id=entry_id)
        else:
            # Usar formulario completo con validaciones
            form = RestoreForm(request.POST, entry=entry, user=request.user)
            if form.is_valid():
                # Intentar restaurar
                success, message, restored_object = RecycleBinService.restore_object(
                    entry, 
                    request.user,
                    notes=form.cleaned_data.get('notes', ''),
                    ip_address=ip_address,
                    user_agent=user_agent
                )
                
                if success:
                    messages.success(request, message)
                    return _redirect_to_restored_object(entry, restored_object)
                else:
                    messages.error(request, message)
                    return redirect('core:recycle_bin_detail', entry_id=entry_id)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
                return redirect('core:recycle_bin_detail', entry_id=entry_id)
    
    # GET request - mostrar formulario
    form = RestoreForm(initial={'entry_id': entry_id}, entry=entry, user=request.user)
    
    context = {
        'entry': entry,
        'form': form,
        'original_object': entry.content_object,
    }
    
    return render(request, 'core/recycle_bin_restore_form.html', context)


def _redirect_to_restored_object(entry, restored_object):
    """Helper para redirigir al objeto restaurado"""
    if restored_object:
        try:
            module_name = entry.module_name
            if module_name == 'oficinas':
                return redirect('oficinas:detalle', pk=restored_object.pk)
            elif module_name == 'bienes':
                return redirect('bienes:detalle', pk=restored_object.pk)
            elif module_name == 'catalogo':
                return redirect('catalogo:detalle', pk=restored_object.pk)
        except:
            pass
    
    return redirect('core:recycle_bin_list')


@login_required
@permission_required_custom('can_bulk_restore')
@require_http_methods(["POST"])
def recycle_bin_bulk_restore(request):
    """
    Vista para restaurar múltiples elementos en lote desde la papelera con validación.
    Implementa validación de permisos granular para cada elemento.
    """
    # Preparar datos del formulario
    entry_ids = request.POST.getlist('entry_ids[]')
    
    if not entry_ids:
        messages.error(request, 'No se seleccionaron elementos para restaurar')
        return redirect('core:recycle_bin_list')
    
    # Crear datos del formulario
    form_data = {
        'operation': 'restore',
        'entry_ids': ','.join(entry_ids),
        'confirm': request.POST.get('confirm', 'on'),
        'notes': request.POST.get('notes', ''),
    }
    
    form = BulkOperationForm(form_data, user=request.user)
    
    if not form.is_valid():
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'{field}: {error}')
        return redirect('core:recycle_bin_list')
    
    # Obtener entradas validadas
    entries = form.get_entries()
    
    # Verificar permisos granulares
    can_restore_own = request.user.profile.can_restore_own_items()
    can_restore_others = request.user.profile.can_restore_others_items()
    
    # Contadores
    restored_count = 0
    error_count = 0
    errors = []
    restored_objects = []
    
    # Obtener contexto para auditoría
    ip_address = request.META.get('REMOTE_ADDR')
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    for entry in entries:
        try:
            # Verificar permisos granulares según el propietario
            if entry.deleted_by == request.user:
                if not can_restore_own:
                    error_count += 1
                    errors.append(f"{entry.object_repr}: Sin permisos para restaurar propios elementos")
                    continue
            else:
                if not can_restore_others:
                    error_count += 1
                    errors.append(f"{entry.object_repr}: Sin permisos para restaurar elementos de otros")
                    continue
            
            # Intentar restaurar
            success, message, restored_object = RecycleBinService.restore_object(
                entry, 
                request.user,
                notes=form.cleaned_data.get('notes', ''),
                ip_address=ip_address,
                user_agent=user_agent
            )
            
            if success:
                restored_count += 1
                if restored_object:
                    restored_objects.append(restored_object)
            else:
                error_count += 1
                errors.append(f"{entry.object_repr}: {message}")
                
        except Exception as e:
            error_count += 1
            errors.append(f"{entry.object_repr}: {str(e)}")
    
    # Registrar operación en lote en auditoría
    if restored_objects:
        from .models import DeletionAuditLog
        DeletionAuditLog.log_bulk_operation(
            action='bulk_restore',
            objects=restored_objects,
            user=request.user,
            ip_address=ip_address,
            user_agent=user_agent,
            metadata={
                'total_count': len(entries),
                'restored_count': restored_count,
                'error_count': error_count,
                'notes': form.cleaned_data.get('notes', '')
            }
        )
    
    # Mensajes de resultado
    if restored_count > 0:
        messages.success(request, f'Se restauraron {restored_count} elemento(s) correctamente')
    
    if error_count > 0:
        error_message = f'No se pudieron restaurar {error_count} elemento(s)'
        if errors:
            error_message += ': ' + '; '.join(errors[:5])  # Mostrar solo los primeros 5 errores
            if len(errors) > 5:
                error_message += f' y {len(errors) - 5} más...'
        messages.error(request, error_message)
    
    return redirect('core:recycle_bin_list')


@login_required
@permission_required_custom('can_permanent_delete')
@require_http_methods(["GET", "POST"])
def recycle_bin_permanent_delete(request, entry_id):
    """
    Vista para eliminar permanentemente un elemento con código de seguridad y validación.
    Incluye sistema de bloqueo temporal, rate limiting y CAPTCHA tras intentos fallidos.
    Requiere permisos de administrador.
    """
    entry = get_object_or_404(RecycleBin, id=entry_id)
    
    # Verificar si el usuario está bloqueado
    from .models import SecurityCodeAttempt
    from django.conf import settings
    
    # Obtener nivel de bloqueo actual
    lockout_level = SecurityCodeAttempt.get_lockout_level(request.user)
    is_locked, attempts, time_remaining = SecurityCodeAttempt.is_user_locked_out(
        request.user,
        max_attempts=lockout_level['max_attempts'],
        lockout_minutes=lockout_level['lockout_minutes']
    )
    
    if is_locked:
        if lockout_level['requires_admin_unlock']:
            messages.error(
                request,
                f'Su cuenta ha sido bloqueada por seguridad (Nivel {lockout_level["name"]}). '
                f'Contacte a un administrador para desbloquearla.'
            )
        else:
            messages.error(
                request,
                f'Su cuenta está bloqueada temporalmente por múltiples intentos fallidos (Nivel {lockout_level["name"]}). '
                f'Podrá intentar nuevamente en {time_remaining} minutos.'
            )
        return redirect('core:recycle_bin_detail', entry_id=entry_id)
    
    # Verificar rate limiting
    is_rate_limited, rate_count, time_until_reset = SecurityCodeAttempt.check_rate_limit(request.user)
    if is_rate_limited:
        messages.error(
            request,
            f'Demasiados intentos. Por favor espere {time_until_reset} minutos antes de intentar nuevamente.'
        )
        return redirect('core:recycle_bin_detail', entry_id=entry_id)
    
    # Verificar si requiere CAPTCHA
    requires_captcha = SecurityCodeAttempt.requires_captcha_validation(request.user)
    recaptcha_site_key = getattr(settings, 'RECAPTCHA_SITE_KEY', None)
    
    if request.method == 'POST':
        form = PermanentDeleteForm(request.POST, entry=entry, user=request.user)
        
        if form.is_valid():
            security_code = form.cleaned_data['security_code']
            reason = form.cleaned_data['reason']
            
            # Obtener información del request
            ip_address = request.META.get('REMOTE_ADDR')
            user_agent = request.META.get('HTTP_USER_AGENT', '')
            session_id = request.session.session_key or ''
            request_path = request.path
            referer = request.META.get('HTTP_REFERER', '')
            
            # Obtener respuesta del CAPTCHA si es requerido
            captcha_response = request.POST.get('g-recaptcha-response', None) if requires_captcha else None
            
            # Intentar eliminación permanente con todas las validaciones de seguridad
            success, message = RecycleBinService.permanent_delete(
                entry, 
                request.user, 
                security_code,
                reason=reason,
                ip_address=ip_address,
                user_agent=user_agent,
                session_id=session_id,
                request_path=request_path,
                referer=referer,
                captcha_response=captcha_response
            )
            
            if success:
                messages.success(request, message)
                return redirect('core:recycle_bin_list')
            else:
                messages.error(request, message)
                # Si el mensaje indica bloqueo o rate limit, redirigir a detalle
                if 'bloqueado' in message.lower() or 'espere' in message.lower():
                    return redirect('core:recycle_bin_detail', entry_id=entry_id)
                # Si es código incorrecto o CAPTCHA fallido, mantener en el formulario para reintentar
                return redirect('core:recycle_bin_permanent_delete', entry_id=entry_id)
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
            return redirect('core:recycle_bin_permanent_delete', entry_id=entry_id)
    
    # GET request - mostrar formulario
    form = PermanentDeleteForm(initial={'entry_id': entry_id}, entry=entry, user=request.user)
    
    # Obtener información de intentos recientes
    recent_failures = SecurityCodeAttempt.get_recent_failed_attempts(request.user)
    attempts_count = recent_failures.count()
    remaining_attempts = max(0, lockout_level['max_attempts'] - attempts_count)
    
    # Obtener resumen de seguridad
    security_summary = SecurityCodeAttempt.get_security_summary(request.user, hours=24)
    
    context = {
        'entry': entry,
        'form': form,
        'attempts_count': attempts_count,
        'remaining_attempts': remaining_attempts,
        'show_warning': attempts_count > 0,
        'requires_captcha': requires_captcha,
        'recaptcha_site_key': recaptcha_site_key,
        'lockout_level': lockout_level,
        'security_summary': security_summary,
    }
    
    return render(request, 'core/recycle_bin_permanent_delete_form.html', context)


@login_required
@permission_required_custom('can_bulk_permanent_delete')
@require_http_methods(["POST"])
def recycle_bin_bulk_permanent_delete(request):
    """
    Vista para eliminar permanentemente múltiples elementos en lote con validación.
    Incluye sistema de bloqueo temporal tras intentos fallidos.
    Requiere permisos de administrador.
    """
    
    # Verificar si el usuario está bloqueado
    from .models import SecurityCodeAttempt
    is_locked, attempts, time_remaining = SecurityCodeAttempt.is_user_locked_out(request.user)
    
    if is_locked:
        messages.error(
            request,
            f'Su cuenta está bloqueada temporalmente por múltiples intentos fallidos. '
            f'Podrá intentar nuevamente en {time_remaining} minutos.'
        )
        return redirect('core:recycle_bin_list')
    
    # Obtener IDs y código de seguridad
    entry_ids = request.POST.getlist('entry_ids[]')
    
    if not entry_ids:
        messages.error(request, 'No se seleccionaron elementos para eliminar')
        return redirect('core:recycle_bin_list')
    
    # Crear datos del formulario
    form_data = {
        'operation': 'permanent_delete',
        'entry_ids': ','.join(entry_ids),
        'security_code': request.POST.get('security_code', ''),
        'confirm': request.POST.get('confirm', 'on'),
        'notes': request.POST.get('notes', ''),
    }
    
    form = BulkOperationForm(form_data, user=request.user)
    
    if not form.is_valid():
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f'{field}: {error}')
        return redirect('core:recycle_bin_list')
    
    # Obtener entradas validadas
    entries = form.get_entries()
    security_code = form.cleaned_data['security_code']
    
    # Obtener información del request
    ip_address = request.META.get('REMOTE_ADDR')
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    # Contadores
    deleted_count = 0
    error_count = 0
    errors = []
    
    for entry in entries:
        try:
            # Intentar eliminación permanente
            success, message = RecycleBinService.permanent_delete(
                entry, 
                request.user, 
                security_code,
                reason=form.cleaned_data.get('notes', 'Eliminación en lote'),
                ip_address=ip_address,
                user_agent=user_agent
            )
            
            if success:
                deleted_count += 1
            else:
                error_count += 1
                errors.append(f"{entry.object_repr}: {message}")
                # Si el código es incorrecto o hay bloqueo, detener el proceso
                if 'incorrecto' in message.lower() or 'bloqueado' in message.lower():
                    break
                
        except Exception as e:
            error_count += 1
            errors.append(f"{entry.object_repr}: {str(e)}")
    
    # Mensajes de resultado
    if deleted_count > 0:
        messages.success(request, f'Se eliminaron permanentemente {deleted_count} elemento(s)')
    
    if error_count > 0:
        error_message = f'No se pudieron eliminar {error_count} elemento(s)'
        if errors:
            error_message += ': ' + '; '.join(errors[:3])
            if len(errors) > 3:
                error_message += f' y {len(errors) - 3} más...'
        messages.error(request, error_message)
    
    return redirect('core:recycle_bin_list')


@login_required
def recycle_bin_dashboard(request):
    """
    Dashboard de estadísticas de la papelera de reciclaje con gráficos y métricas.
    Muestra elementos por módulo, usuario, tiempo, y métricas de restauraciones vs eliminaciones.
    """
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    # Verificar permisos
    is_admin = hasattr(request.user, 'profile') and request.user.profile.is_administrador
    
    # Queryset base
    if is_admin:
        queryset = RecycleBin.objects.all()
    else:
        queryset = RecycleBin.objects.filter(deleted_by=request.user)
    
    # Filtros de fecha
    date_filter = request.GET.get('date_range', '30')  # Por defecto últimos 30 días
    try:
        days = int(date_filter)
        if days > 0:
            start_date = timezone.now() - timedelta(days=days)
            queryset = queryset.filter(deleted_at__gte=start_date)
    except (ValueError, TypeError):
        days = 30
    
    # ============================================================================
    # ESTADÍSTICAS GENERALES
    # ============================================================================
    
    total_deleted = queryset.count()
    total_restored = queryset.filter(restored_at__isnull=False).count()
    total_pending = queryset.filter(restored_at__isnull=True).count()
    
    # Elementos cerca de eliminación automática
    near_auto_delete = 0
    ready_for_auto_delete = 0
    for entry in queryset.filter(restored_at__isnull=True):
        if entry.is_near_auto_delete:
            near_auto_delete += 1
        if entry.is_ready_for_auto_delete:
            ready_for_auto_delete += 1
    
    # ============================================================================
    # ESTADÍSTICAS POR MÓDULO
    # ============================================================================
    
    stats_by_module = queryset.values('module_name').annotate(
        total=Count('id'),
        restored=Count('id', filter=Q(restored_at__isnull=False)),
        pending=Count('id', filter=Q(restored_at__isnull=True))
    ).order_by('-total')
    
    # Preparar datos para gráfico de módulos
    module_labels = []
    module_deleted = []
    module_restored = []
    module_pending = []
    
    module_names_map = {
        'oficinas': 'Oficinas',
        'bienes': 'Bienes Patrimoniales',
        'catalogo': 'Catálogo',
        'core': 'Sistema',
    }
    
    for stat in stats_by_module:
        module_labels.append(module_names_map.get(stat['module_name'], stat['module_name']))
        module_deleted.append(stat['total'])
        module_restored.append(stat['restored'])
        module_pending.append(stat['pending'])
    
    # ============================================================================
    # ESTADÍSTICAS POR USUARIO (solo para administradores)
    # ============================================================================
    
    stats_by_user = []
    user_labels = []
    user_deleted = []
    user_restored = []
    
    if is_admin:
        stats_by_user = queryset.values('deleted_by__username', 'deleted_by__first_name', 'deleted_by__last_name').annotate(
            total=Count('id'),
            restored=Count('id', filter=Q(restored_at__isnull=False)),
            pending=Count('id', filter=Q(restored_at__isnull=True))
        ).order_by('-total')[:10]  # Top 10 usuarios
        
        for stat in stats_by_user:
            username = stat['deleted_by__username']
            first_name = stat['deleted_by__first_name']
            last_name = stat['deleted_by__last_name']
            
            if first_name and last_name:
                display_name = f"{first_name} {last_name}"
            else:
                display_name = username
            
            user_labels.append(display_name)
            user_deleted.append(stat['total'])
            user_restored.append(stat['restored'])
    
    # ============================================================================
    # ESTADÍSTICAS POR TIEMPO (últimos días)
    # ============================================================================
    
    # Agrupar por día
    from django.db.models.functions import TruncDate
    
    stats_by_date = queryset.annotate(
        date=TruncDate('deleted_at')
    ).values('date').annotate(
        deleted=Count('id'),
        restored=Count('id', filter=Q(restored_at__date=TruncDate('deleted_at')))
    ).order_by('date')
    
    # Preparar datos para gráfico de tiempo
    time_labels = []
    time_deleted = []
    time_restored = []
    
    for stat in stats_by_date:
        time_labels.append(stat['date'].strftime('%d/%m/%Y'))
        time_deleted.append(stat['deleted'])
        time_restored.append(stat['restored'])
    
    # ============================================================================
    # MÉTRICAS DE RESTAURACIONES VS ELIMINACIONES PERMANENTES
    # ============================================================================
    
    # Contar eliminaciones permanentes (las que ya no están en RecycleBin)
    # Esto lo obtenemos del AuditLog
    from .models import AuditLog
    
    permanent_deletes = AuditLog.objects.filter(
        action='permanent_delete',
        model_name='RecycleBin',
        created_at__gte=timezone.now() - timedelta(days=days)
    ).count()
    
    # Tasa de restauración
    restoration_rate = 0
    if total_deleted > 0:
        restoration_rate = round((total_restored / total_deleted) * 100, 1)
    
    # Tasa de eliminación permanente
    permanent_delete_rate = 0
    if total_deleted > 0:
        permanent_delete_rate = round((permanent_deletes / total_deleted) * 100, 1)
    
    # ============================================================================
    # ELEMENTOS RECIENTES
    # ============================================================================
    
    recent_deletions = queryset.filter(restored_at__isnull=True).order_by('-deleted_at')[:5]
    recent_restorations = queryset.filter(restored_at__isnull=False).order_by('-restored_at')[:5]
    
    # ============================================================================
    # PREPARAR CONTEXTO
    # ============================================================================
    
    context = {
        'is_admin': is_admin,
        'date_range': days,
        
        # Estadísticas generales
        'total_deleted': total_deleted,
        'total_restored': total_restored,
        'total_pending': total_pending,
        'near_auto_delete': near_auto_delete,
        'ready_for_auto_delete': ready_for_auto_delete,
        'restoration_rate': restoration_rate,
        'permanent_delete_rate': permanent_delete_rate,
        'permanent_deletes': permanent_deletes,
        
        # Datos para gráficos (JSON)
        'module_chart_data': json.dumps({
            'labels': module_labels,
            'deleted': module_deleted,
            'restored': module_restored,
            'pending': module_pending,
        }),
        
        'user_chart_data': json.dumps({
            'labels': user_labels,
            'deleted': user_deleted,
            'restored': user_restored,
        }),
        
        'time_chart_data': json.dumps({
            'labels': time_labels,
            'deleted': time_deleted,
            'restored': time_restored,
        }),
        
        # Elementos recientes
        'recent_deletions': recent_deletions,
        'recent_restorations': recent_restorations,
        
        # Estadísticas detalladas
        'stats_by_module': stats_by_module,
        'stats_by_user': stats_by_user,
    }
    
    return render(request, 'core/recycle_bin_dashboard.html', context)


@login_required
@require_http_methods(["GET"])
def recycle_bin_export_report(request):
    """
    Exporta un reporte de la papelera en formato CSV o Excel.
    Incluye estadísticas y listado de elementos.
    """
    from django.http import HttpResponse
    from django.utils import timezone
    import csv
    from datetime import timedelta
    
    # Verificar permisos
    is_admin = hasattr(request.user, 'profile') and request.user.profile.is_administrador
    
    # Queryset base
    if is_admin:
        queryset = RecycleBin.objects.all()
    else:
        queryset = RecycleBin.objects.filter(deleted_by=request.user)
    
    # Filtros
    date_filter = request.GET.get('date_range', '30')
    try:
        days = int(date_filter)
        if days > 0:
            start_date = timezone.now() - timedelta(days=days)
            queryset = queryset.filter(deleted_at__gte=start_date)
    except (ValueError, TypeError):
        days = 30
    
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'pending':
        queryset = queryset.filter(restored_at__isnull=True)
    elif status_filter == 'restored':
        queryset = queryset.filter(restored_at__isnull=False)
    
    module_filter = request.GET.get('module', '')
    if module_filter:
        queryset = queryset.filter(module_name=module_filter)
    
    # Formato de exportación
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        # Crear respuesta CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="reporte_papelera_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Agregar BOM para Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'ID',
            'Módulo',
            'Tipo de Objeto',
            'Representación',
            'Eliminado Por',
            'Fecha de Eliminación',
            'Motivo',
            'Estado',
            'Restaurado Por',
            'Fecha de Restauración',
            'Eliminación Automática',
        ])
        
        # Datos
        for entry in queryset.select_related('deleted_by', 'restored_by', 'content_type'):
            status = 'Restaurado' if entry.restored_at else 'Pendiente'
            
            writer.writerow([
                entry.id,
                entry.module_name,
                entry.content_type.model if entry.content_type else '',
                entry.object_repr,
                entry.deleted_by.username if entry.deleted_by else '',
                entry.deleted_at.strftime('%Y-%m-%d %H:%M:%S'),
                entry.deletion_reason or '',
                status,
                entry.restored_by.username if entry.restored_by else '',
                entry.restored_at.strftime('%Y-%m-%d %H:%M:%S') if entry.restored_at else '',
                entry.auto_delete_at.strftime('%Y-%m-%d %H:%M:%S') if entry.auto_delete_at else '',
            ])
        
        return response
    
    elif export_format == 'json':
        # Exportar como JSON
        data = []
        for entry in queryset.select_related('deleted_by', 'restored_by', 'content_type'):
            data.append({
                'id': entry.id,
                'module_name': entry.module_name,
                'content_type': entry.content_type.model if entry.content_type else None,
                'object_repr': entry.object_repr,
                'deleted_by': entry.deleted_by.username if entry.deleted_by else None,
                'deleted_at': entry.deleted_at.isoformat(),
                'deletion_reason': entry.deletion_reason,
                'status': 'restored' if entry.restored_at else 'pending',
                'restored_by': entry.restored_by.username if entry.restored_by else None,
                'restored_at': entry.restored_at.isoformat() if entry.restored_at else None,
                'auto_delete_at': entry.auto_delete_at.isoformat() if entry.auto_delete_at else None,
            })
        
        return JsonResponse({
            'date_range_days': days,
            'total_records': len(data),
            'exported_at': timezone.now().isoformat(),
            'data': data
        })
    
    else:
        messages.error(request, 'Formato de exportación no válido')
        return redirect('core:recycle_bin_dashboard')



@login_required
@role_required(['administrador', 'auditor'])
def security_monitoring_dashboard(request):
    """
    Dashboard de monitoreo de seguridad para intentos de código de seguridad.
    Muestra estadísticas, usuarios bloqueados y actividad sospechosa.
    """
    from .models import SecurityCodeAttempt
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # Obtener período de análisis
    hours = int(request.GET.get('hours', 24))
    cutoff_time = timezone.now() - timedelta(hours=hours)
    
    # Obtener reporte de actividad sospechosa
    suspicious_report = SecurityCodeAttempt.get_suspicious_activity_report(hours=hours)
    
    # Obtener usuarios con más intentos fallidos
    users_with_failures = SecurityCodeAttempt.objects.filter(
        attempted_at__gte=cutoff_time,
        success=False
    ).values(
        'user__username', 'user__id', 'user__email'
    ).annotate(
        failure_count=Count('id')
    ).order_by('-failure_count')[:20]
    
    # Enriquecer con información de bloqueo
    users_data = []
    for user_data in users_with_failures:
        try:
            user = User.objects.get(id=user_data['user__id'])
            is_locked, attempts, time_remaining = SecurityCodeAttempt.is_user_locked_out(user)
            lockout_level = SecurityCodeAttempt.get_lockout_level(user)
            
            users_data.append({
                'username': user_data['user__username'],
                'email': user_data['user__email'],
                'failure_count': user_data['failure_count'],
                'is_locked': is_locked,
                'time_remaining': time_remaining,
                'lockout_level': lockout_level['name'],
                'requires_admin_unlock': lockout_level['requires_admin_unlock']
            })
        except User.DoesNotExist:
            pass
    
    # Obtener IPs con más intentos fallidos
    ips_with_failures = SecurityCodeAttempt.objects.filter(
        attempted_at__gte=cutoff_time,
        success=False
    ).values('ip_address').annotate(
        failure_count=Count('id')
    ).order_by('-failure_count')[:20]
    
    # Obtener intentos recientes (últimos 50)
    recent_attempts = SecurityCodeAttempt.objects.select_related('user').filter(
        attempted_at__gte=cutoff_time
    ).order_by('-attempted_at')[:50]
    
    # Estadísticas por tipo de intento
    attempt_types = SecurityCodeAttempt.objects.filter(
        attempted_at__gte=cutoff_time
    ).values('attempt_type').annotate(
        total=Count('id'),
        failed=Count('id', filter=Q(success=False)),
        successful=Count('id', filter=Q(success=True))
    )
    
    # Gráfico de intentos por hora
    attempts_by_hour = []
    for i in range(hours):
        hour_start = timezone.now() - timedelta(hours=i+1)
        hour_end = timezone.now() - timedelta(hours=i)
        
        count = SecurityCodeAttempt.objects.filter(
            attempted_at__gte=hour_start,
            attempted_at__lt=hour_end
        ).count()
        
        attempts_by_hour.append({
            'hour': hour_start.strftime('%H:00'),
            'count': count
        })
    
    attempts_by_hour.reverse()
    
    context = {
        'hours': hours,
        'suspicious_report': suspicious_report,
        'users_data': users_data,
        'ips_with_failures': ips_with_failures,
        'recent_attempts': recent_attempts,
        'attempt_types': attempt_types,
        'attempts_by_hour': attempts_by_hour,
    }
    
    return render(request, 'core/security_monitoring_dashboard.html', context)


@login_required
@role_required(['administrador'])
@require_http_methods(["POST"])
def unlock_user_security(request, user_id):
    """
    Desbloquea manualmente a un usuario que ha sido bloqueado por intentos fallidos.
    Solo administradores pueden realizar esta acción.
    """
    from .models import SecurityCodeAttempt
    
    user = get_object_or_404(User, id=user_id)
    
    # Eliminar intentos fallidos recientes para desbloquear
    recent_failures = SecurityCodeAttempt.get_recent_failed_attempts(user, minutes=120)
    count = recent_failures.count()
    recent_failures.delete()
    
    # Registrar en auditoría
    AuditLog.objects.create(
        user=request.user,
        action='update',
        model_name='User',
        object_id=str(user.id),
        object_repr=user.username,
        ip_address=request.META.get('REMOTE_ADDR'),
        changes={
            'action': 'security_unlock',
            'cleared_attempts': count,
            'unlocked_by': request.user.username
        }
    )
    
    messages.success(
        request,
        f'Usuario {user.username} desbloqueado correctamente. Se eliminaron {count} intentos fallidos.'
    )
    
    return redirect('core:security_monitoring_dashboard')


@login_required
@role_required(['administrador', 'auditor'])
def security_attempt_detail(request, attempt_id):
    """
    Vista de detalle de un intento de código de seguridad específico.
    Muestra toda la información de auditoría del intento.
    """
    from .models import SecurityCodeAttempt
    
    attempt = get_object_or_404(SecurityCodeAttempt, id=attempt_id)
    
    # Obtener otros intentos del mismo usuario en el mismo período
    related_attempts = SecurityCodeAttempt.objects.filter(
        user=attempt.user,
        attempted_at__gte=attempt.attempted_at - timedelta(hours=1),
        attempted_at__lte=attempt.attempted_at + timedelta(hours=1)
    ).exclude(id=attempt.id).order_by('-attempted_at')[:10]
    
    # Obtener otros intentos desde la misma IP
    ip_attempts = SecurityCodeAttempt.objects.filter(
        ip_address=attempt.ip_address,
        attempted_at__gte=attempt.attempted_at - timedelta(hours=24)
    ).exclude(id=attempt.id).order_by('-attempted_at')[:10]
    
    # Obtener resumen de seguridad del usuario
    security_summary = SecurityCodeAttempt.get_security_summary(attempt.user, hours=24)
    
    context = {
        'attempt': attempt,
        'related_attempts': related_attempts,
        'ip_attempts': ip_attempts,
        'security_summary': security_summary,
    }
    
    return render(request, 'core/security_attempt_detail.html', context)



# ============================================================================
# RECYCLE BIN REAL-TIME API
# ============================================================================


@login_required
@require_http_methods(["GET"])
def recycle_bin_status_api(request):
    """
    API endpoint para obtener el estado actual de la papelera de reciclaje.
    Usado para actualizaciones en tiempo real de contadores y notificaciones.
    
    Returns:
        JsonResponse: Estado actual de la papelera
    """
    # Verificar permisos
    if not hasattr(request.user, 'profile') or not request.user.profile.can_view_recycle_bin():
        return JsonResponse({
            'error': 'No tienes permisos para ver la papelera'
        }, status=403)
    
    # Obtener elementos según permisos
    if request.user.profile.can_view_all_recycle_items():
        recycle_items = RecycleBin.objects.filter(restored_at__isnull=True)
    else:
        recycle_items = RecycleBin.objects.filter(
            deleted_by=request.user,
            restored_at__isnull=True
        )
    
    # Contar elementos totales
    total_count = recycle_items.count()
    
    # Contar elementos próximos a eliminarse (7 días o menos)
    seven_days_from_now = timezone.now() + timezone.timedelta(days=7)
    near_delete_count = recycle_items.filter(
        auto_delete_at__lte=seven_days_from_now
    ).count()
    
    # Obtener elementos más urgentes (próximos 3 días)
    three_days_from_now = timezone.now() + timezone.timedelta(days=3)
    urgent_items = recycle_items.filter(
        auto_delete_at__lte=three_days_from_now
    ).select_related('deleted_by', 'content_type').order_by('auto_delete_at')[:5]
    
    # Serializar elementos urgentes
    urgent_items_data = []
    for item in urgent_items:
        urgent_items_data.append({
            'id': item.id,
            'object_repr': item.object_repr,
            'module_name': item.module_name,
            'module_display': item.get_module_display(),
            'deleted_at': item.deleted_at.isoformat(),
            'auto_delete_at': item.auto_delete_at.isoformat(),
            'days_until_delete': item.days_until_auto_delete,
            'deleted_by': item.deleted_by.username if item.deleted_by else None,
        })
    
    # Estadísticas por módulo
    module_stats = {}
    for module in ['oficinas', 'bienes', 'catalogo']:
        module_count = recycle_items.filter(module_name=module).count()
        if module_count > 0:
            module_stats[module] = module_count
    
    return JsonResponse({
        'count': total_count,
        'near_delete_count': near_delete_count,
        'urgent_items': urgent_items_data,
        'module_stats': module_stats,
        'timestamp': timezone.now().isoformat(),
    })
