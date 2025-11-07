import os
import tempfile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count
from django.core.paginator import Paginator
from .models import Oficina, HistorialOficina
from .utils import importar_oficinas_desde_excel, validar_estructura_oficinas, generar_preview_oficinas, generar_plantilla_oficinas


@login_required
def lista_oficinas_view(request):
    """Vista para listar oficinas"""
    oficinas = Oficina.objects.annotate(
        bienes_count=Count('bienpatrimonial')
    ).order_by('codigo')
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    busqueda = request.GET.get('q')
    responsable_filtro = request.GET.get('responsable')
    
    if estado_filtro == 'True':
        oficinas = oficinas.filter(estado=True)
    elif estado_filtro == 'False':
        oficinas = oficinas.filter(estado=False)
    
    if busqueda:
        oficinas = oficinas.filter(
            Q(codigo__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(responsable__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if responsable_filtro:
        oficinas = oficinas.filter(responsable__icontains=responsable_filtro)
    
    # Paginación
    paginator = Paginator(oficinas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    total_oficinas = Oficina.objects.count()
    oficinas_activas = Oficina.objects.filter(estado=True).count()
    oficinas_inactivas = Oficina.objects.filter(estado=False).count()
    
    context = {
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
        'responsable_filtro': responsable_filtro,
        'total_oficinas': total_oficinas,
        'oficinas_activas': oficinas_activas,
        'oficinas_inactivas': oficinas_inactivas,
    }
    
    return render(request, 'oficinas/lista.html', context)


@login_required
def detalle_oficina_view(request, oficina_id):
    """Vista para mostrar detalle de una oficina"""
    oficina = get_object_or_404(Oficina, id=oficina_id)
    
    # Obtener bienes de la oficina (cuando esté implementado)
    # bienes = oficina.bienpatrimonial_set.all()[:10]
    
    context = {
        'oficina': oficina,
        # 'bienes': bienes,
    }
    
    return render(request, 'oficinas/detalle.html', context)


@login_required
@permission_required('oficinas.add_oficina', raise_exception=True)
def importar_oficinas_view(request):
    """Vista para importar oficinas desde Excel"""
    if request.method == 'POST':
        if 'archivo_excel' not in request.FILES:
            messages.error(request, 'No se seleccionó ningún archivo.')
            return redirect('oficinas:importar')
        
        archivo = request.FILES['archivo_excel']
        actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'
        
        # Validar extensión
        if not archivo.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'El archivo debe ser un Excel (.xlsx o .xls).')
            return redirect('oficinas:importar')
        
        # Guardar archivo temporalmente
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                for chunk in archivo.chunks():
                    temp_file.write(chunk)
                temp_path = temp_file.name
            
            # Procesar archivo
            resultado = importar_oficinas_desde_excel(temp_path, actualizar_existentes)
            
            # Limpiar archivo temporal
            os.unlink(temp_path)
            
            # Mostrar resultados
            if resultado['exito']:
                messages.success(request, f"Importación exitosa: {resultado['resumen']}")
                
                # Mostrar advertencias si las hay
                for warning in resultado['warnings']:
                    messages.warning(request, warning)
            else:
                messages.error(request, "Error en la importación:")
                for error in resultado['errores']:
                    messages.error(request, error)
            
            return redirect('oficinas:importar')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('oficinas:importar')
    
    # GET request - mostrar formulario
    context = {
        'total_oficinas': Oficina.objects.count(),
        'oficinas_activas': Oficina.objects.filter(estado=True).count(),
        'oficinas_inactivas': Oficina.objects.filter(estado=False).count(),
    }
    
    return render(request, 'oficinas/importar.html', context)


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def validar_archivo_oficinas(request):
    """API endpoint para validar estructura del archivo Excel"""
    if 'archivo' not in request.FILES:
        return JsonResponse({
            'error': 'No se proporcionó archivo'
        }, status=400)
    
    archivo = request.FILES['archivo']
    
    try:
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            for chunk in archivo.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name
        
        # Validar estructura
        resultado = validar_estructura_oficinas(temp_path)
        
        # Limpiar archivo temporal
        os.unlink(temp_path)
        
        return JsonResponse(resultado)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error al validar archivo: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["POST"])
@login_required
def preview_oficinas(request):
    """API endpoint para generar preview de datos del archivo Excel"""
    if 'archivo' not in request.FILES:
        return JsonResponse({
            'error': 'No se proporcionó archivo'
        }, status=400)
    
    archivo = request.FILES['archivo']
    
    try:
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            for chunk in archivo.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name
        
        # Generar preview
        resultado = generar_preview_oficinas(temp_path, max_filas=15)
        
        # Limpiar archivo temporal
        os.unlink(temp_path)
        
        return JsonResponse(resultado)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Error al generar preview: {str(e)}'
        }, status=500)


@login_required
def descargar_plantilla_oficinas(request):
    """Vista para descargar plantilla Excel de oficinas"""
    try:
        from django.http import HttpResponse
        from datetime import datetime
        
        # Generar plantilla
        wb = generar_plantilla_oficinas()
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo con fecha
        filename = f"plantilla_oficinas_drtc_puno_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar plantilla: {str(e)}')
        return redirect('oficinas:importar')


@login_required
def reportes_oficinas_view(request):
    """Vista para mostrar reportes y estadísticas de oficinas"""
    # Obtener todas las oficinas con conteo de bienes
    oficinas = Oficina.objects.annotate(
        bienes_count=Count('bienpatrimonial')
    ).order_by('codigo')
    
    # Aplicar filtros
    estado_filtro = request.GET.get('estado')
    busqueda = request.GET.get('q')
    responsable_filtro = request.GET.get('responsable')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    oficinas_filtradas = oficinas
    
    if estado_filtro == 'True':
        oficinas_filtradas = oficinas_filtradas.filter(estado=True)
    elif estado_filtro == 'False':
        oficinas_filtradas = oficinas_filtradas.filter(estado=False)
    
    if busqueda:
        oficinas_filtradas = oficinas_filtradas.filter(
            Q(codigo__icontains=busqueda) |
            Q(nombre__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if responsable_filtro:
        oficinas_filtradas = oficinas_filtradas.filter(responsable__icontains=responsable_filtro)
    
    if fecha_desde:
        from datetime import datetime
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        oficinas_filtradas = oficinas_filtradas.filter(created_at__date__gte=fecha_desde_obj)
    
    if fecha_hasta:
        from datetime import datetime
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        oficinas_filtradas = oficinas_filtradas.filter(created_at__date__lte=fecha_hasta_obj)
    
    # Estadísticas generales
    estadisticas = {
        'total_oficinas': oficinas.count(),
        'oficinas_activas': oficinas.filter(estado=True).count(),
        'oficinas_inactivas': oficinas.filter(estado=False).count(),
        'total_bienes': sum(oficina.bienes_count for oficina in oficinas)
    }
    
    # Datos para gráficos
    # Top 10 oficinas con más bienes
    top_oficinas_bienes = oficinas.filter(bienes_count__gt=0).order_by('-bienes_count')[:10]
    bienes_por_oficina = [
        {
            'codigo': oficina.codigo,
            'nombre': oficina.nombre,
            'total_bienes': oficina.bienes_count
        }
        for oficina in top_oficinas_bienes
    ]
    
    graficos = {
        'bienes_por_oficina': json.dumps(bienes_por_oficina)
    }
    
    context = {
        'oficinas_filtradas': oficinas_filtradas,
        'estadisticas': estadisticas,
        'graficos': graficos,
    }
    
    return render(request, 'oficinas/reportes.html', context)


@login_required
def buscar_oficinas_api(request):
    """API para buscar oficinas (para autocompletado)"""
    termino = request.GET.get('q', '')
    limite = int(request.GET.get('limit', 10))
    solo_activas = request.GET.get('activas', 'true').lower() == 'true'
    
    if len(termino) < 2:
        return JsonResponse({'resultados': []})
    
    oficinas = Oficina.buscar_por_nombre(termino)
    
    if solo_activas:
        oficinas = oficinas.filter(estado=True)
    
    oficinas = oficinas[:limite]
    
    resultados = []
    for oficina in oficinas:
        resultados.append({
            'id': oficina.id,
            'codigo': oficina.codigo,
            'nombre': oficina.nombre,
            'responsable': oficina.responsable,
            'estado': oficina.estado,
            'texto': f"{oficina.codigo} - {oficina.nombre}"
        })
    
    return JsonResponse({'resultados': resultados})


@login_required
def estadisticas_oficinas_view(request):
    """Vista para mostrar estadísticas de oficinas"""
    # Estadísticas generales
    total = Oficina.objects.count()
    activas = Oficina.objects.filter(estado=True).count()
    inactivas = Oficina.objects.filter(estado=False).count()
    
    # Oficinas con más bienes (cuando esté implementado)
    # oficinas_con_bienes = Oficina.objects.annotate(
    #     total_bienes=Count('bienpatrimonial')
    # ).order_by('-total_bienes')[:10]
    
    # Oficinas por responsable
    responsables = Oficina.objects.values('responsable').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    context = {
        'total': total,
        'activas': activas,
        'inactivas': inactivas,
        'responsables': responsables,
        # 'oficinas_con_bienes': oficinas_con_bienes,
    }
    
    return render(request, 'oficinas/estadisticas.html', context)


@login_required
@permission_required('oficinas.change_oficina', raise_exception=True)
def activar_oficina_view(request, oficina_id):
    """Vista para activar una oficina"""
    oficina = get_object_or_404(Oficina, id=oficina_id)
    oficina.activar()
    messages.success(request, f'La oficina "{oficina.nombre}" ha sido activada.')
    return redirect('oficinas:detalle', oficina_id=oficina.id)


@login_required
@permission_required('oficinas.add_oficina', raise_exception=True)
@require_http_methods(["POST"])
def crear_oficina_view(request):
    """Vista para crear una nueva oficina via AJAX"""
    try:
        # Obtener datos del formulario
        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        responsable = request.POST.get('responsable', '').strip()
        cargo_responsable = request.POST.get('cargo_responsable', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        ubicacion = request.POST.get('ubicacion', '').strip()
        estado = request.POST.get('estado') == 'on'
        
        # Validaciones
        errors = {}
        
        if not codigo:
            errors['codigo'] = ['Este campo es requerido']
        elif Oficina.objects.filter(codigo=codigo).exists():
            errors['codigo'] = ['Ya existe una oficina con este código']
        
        if not nombre:
            errors['nombre'] = ['Este campo es requerido']
        
        if not responsable:
            errors['responsable'] = ['Este campo es requerido']
        
        if len(codigo) > 20:
            errors['codigo'] = errors.get('codigo', []) + ['El código no puede tener más de 20 caracteres']
        
        if len(nombre) > 200:
            errors['nombre'] = errors.get('nombre', []) + ['El nombre no puede tener más de 200 caracteres']
        
        if len(responsable) > 200:
            errors['responsable'] = errors.get('responsable', []) + ['El responsable no puede tener más de 200 caracteres']
        
        if email and len(email) > 254:
            errors['email'] = ['El email no puede tener más de 254 caracteres']
        
        if telefono and len(telefono) > 20:
            errors['telefono'] = ['El teléfono no puede tener más de 20 caracteres']
        
        if errors:
            return JsonResponse({
                'success': False,
                'errors': errors
            })
        
        # Crear la oficina
        oficina = Oficina.objects.create(
            codigo=codigo,
            nombre=nombre,
            descripcion=descripcion,
            responsable=responsable,
            cargo_responsable=cargo_responsable,
            telefono=telefono,
            email=email,
            ubicacion=ubicacion,
            estado=estado,
            created_by=request.user
        )
        
        # Registrar en historial
        datos_nuevos = {
            'codigo': codigo,
            'nombre': nombre,
            'descripcion': descripcion,
            'responsable': responsable,
            'cargo_responsable': cargo_responsable,
            'telefono': telefono,
            'email': email,
            'ubicacion': ubicacion,
            'estado': estado
        }
        
        HistorialOficina.registrar_cambio(
            oficina=oficina,
            accion='CREAR',
            usuario=request.user,
            datos_nuevos=datos_nuevos,
            observaciones='Oficina creada desde interfaz web',
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Oficina creada exitosamente',
            'oficina': {
                'id': oficina.id,
                'codigo': oficina.codigo,
                'nombre': oficina.nombre,
                'responsable': oficina.responsable,
                'estado': oficina.estado
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        })


@login_required
@permission_required('oficinas.change_oficina', raise_exception=True)
@require_http_methods(["POST"])
def editar_oficina_view(request, oficina_id):
    """Vista para editar una oficina via AJAX"""
    try:
        oficina = get_object_or_404(Oficina, id=oficina_id)
        
        # Obtener datos del formulario
        codigo = request.POST.get('codigo', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        responsable = request.POST.get('responsable', '').strip()
        cargo_responsable = request.POST.get('cargo_responsable', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        email = request.POST.get('email', '').strip()
        ubicacion = request.POST.get('ubicacion', '').strip()
        estado = request.POST.get('estado') == 'on'
        
        # Validaciones
        errors = {}
        
        if not codigo:
            errors['codigo'] = ['Este campo es requerido']
        elif Oficina.objects.filter(codigo=codigo).exclude(id=oficina.id).exists():
            errors['codigo'] = ['Ya existe otra oficina con este código']
        
        if not nombre:
            errors['nombre'] = ['Este campo es requerido']
        
        if not responsable:
            errors['responsable'] = ['Este campo es requerido']
        
        if len(codigo) > 20:
            errors['codigo'] = errors.get('codigo', []) + ['El código no puede tener más de 20 caracteres']
        
        if len(nombre) > 200:
            errors['nombre'] = errors.get('nombre', []) + ['El nombre no puede tener más de 200 caracteres']
        
        if len(responsable) > 200:
            errors['responsable'] = errors.get('responsable', []) + ['El responsable no puede tener más de 200 caracteres']
        
        if email and len(email) > 254:
            errors['email'] = ['El email no puede tener más de 254 caracteres']
        
        if telefono and len(telefono) > 20:
            errors['telefono'] = ['El teléfono no puede tener más de 20 caracteres']
        
        if errors:
            return JsonResponse({
                'success': False,
                'errors': errors
            })
        
        # Guardar datos anteriores para el historial
        datos_anteriores = {
            'codigo': oficina.codigo,
            'nombre': oficina.nombre,
            'descripcion': oficina.descripcion,
            'responsable': oficina.responsable,
            'cargo_responsable': oficina.cargo_responsable,
            'telefono': oficina.telefono,
            'email': oficina.email,
            'ubicacion': oficina.ubicacion,
            'estado': oficina.estado
        }
        
        # Actualizar la oficina
        oficina.codigo = codigo
        oficina.nombre = nombre
        oficina.descripcion = descripcion
        oficina.responsable = responsable
        oficina.cargo_responsable = cargo_responsable
        oficina.telefono = telefono
        oficina.email = email
        oficina.ubicacion = ubicacion
        oficina.estado = estado
        oficina.updated_by = request.user
        oficina.save()
        
        # Registrar en historial
        datos_nuevos = {
            'codigo': codigo,
            'nombre': nombre,
            'descripcion': descripcion,
            'responsable': responsable,
            'cargo_responsable': cargo_responsable,
            'telefono': telefono,
            'email': email,
            'ubicacion': ubicacion,
            'estado': estado
        }
        
        HistorialOficina.registrar_cambio(
            oficina=oficina,
            accion='EDITAR',
            usuario=request.user,
            datos_anteriores=datos_anteriores,
            datos_nuevos=datos_nuevos,
            observaciones='Oficina editada desde interfaz web',
            request=request
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Oficina actualizada exitosamente',
            'oficina': {
                'id': oficina.id,
                'codigo': oficina.codigo,
                'nombre': oficina.nombre,
                'responsable': oficina.responsable,
                'estado': oficina.estado
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        })


@login_required
@require_http_methods(["GET"])
def obtener_oficina_view(request, oficina_id):
    """Vista para obtener datos de una oficina via AJAX"""
    try:
        oficina = get_object_or_404(Oficina, id=oficina_id)
        
        return JsonResponse({
            'success': True,
            'oficina': {
                'id': oficina.id,
                'codigo': oficina.codigo,
                'nombre': oficina.nombre,
                'descripcion': oficina.descripcion,
                'responsable': oficina.responsable,
                'cargo_responsable': oficina.cargo_responsable,
                'telefono': oficina.telefono,
                'email': oficina.email,
                'ubicacion': oficina.ubicacion,
                'estado': oficina.estado
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        })


@login_required
@permission_required('oficinas.change_oficina', raise_exception=True)
def desactivar_oficina_view(request, oficina_id):
    """Vista para desactivar una oficina"""
    oficina = get_object_or_404(Oficina, id=oficina_id)
    
    # Verificar si tiene bienes asignados
    if oficina.total_bienes > 0:
        messages.error(
            request, 
            f'No se puede desactivar la oficina "{oficina.nombre}" porque tiene {oficina.total_bienes} bienes asignados.'
        )
    else:
        oficina.desactivar()
        messages.success(request, f'La oficina "{oficina.nombre}" ha sido desactivada.')
    
    return redirect('oficinas:detalle', oficina_id=oficina.id)


@login_required
@permission_required('oficinas.delete_oficina', raise_exception=True)
@require_http_methods(["POST"])
def eliminar_oficina_view(request, oficina_id):
    """Vista para eliminar una oficina via AJAX"""
    try:
        oficina = get_object_or_404(Oficina, id=oficina_id)
        
        # Verificar si puede eliminarse (no tiene bienes asignados)
        if not oficina.puede_eliminarse():
            return JsonResponse({
                'success': False,
                'errors': {
                    'general': [f'No se puede eliminar la oficina "{oficina.nombre}" porque tiene {oficina.total_bienes} bienes asignados.']
                }
            })
        
        # Guardar información para el mensaje
        nombre_oficina = oficina.nombre
        codigo_oficina = oficina.codigo
        
        # Eliminar la oficina
        oficina.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'La oficina "{nombre_oficina}" ({codigo_oficina}) ha sido eliminada exitosamente.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        })


@login_required
def exportar_oficinas_view(request):
    """Vista para exportar oficinas a Excel con filtros aplicados"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from datetime import datetime
        
        # Obtener oficinas con los mismos filtros que la lista
        oficinas = Oficina.objects.annotate(
            bienes_count=Count('bienpatrimonial')
        ).order_by('codigo')
        
        # Aplicar filtros
        estado_filtro = request.GET.get('estado')
        busqueda = request.GET.get('q')
        responsable_filtro = request.GET.get('responsable')
        
        if estado_filtro == 'True':
            oficinas = oficinas.filter(estado=True)
        elif estado_filtro == 'False':
            oficinas = oficinas.filter(estado=False)
        
        if busqueda:
            oficinas = oficinas.filter(
                Q(codigo__icontains=busqueda) |
                Q(nombre__icontains=busqueda) |
                Q(responsable__icontains=busqueda) |
                Q(descripcion__icontains=busqueda)
            )
        
        if responsable_filtro:
            oficinas = oficinas.filter(responsable__icontains=responsable_filtro)
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Oficinas DRTC Puno"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = [
            'Código', 'Nombre', 'Responsable', 'Cargo Responsable', 
            'Teléfono', 'Email', 'Ubicación', 'Descripción', 
            'Estado', 'Total Bienes', 'Fecha Registro'
        ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos
        for row, oficina in enumerate(oficinas, 2):
            ws.cell(row=row, column=1, value=oficina.codigo)
            ws.cell(row=row, column=2, value=oficina.nombre)
            ws.cell(row=row, column=3, value=oficina.responsable)
            ws.cell(row=row, column=4, value=oficina.cargo_responsable or '')
            ws.cell(row=row, column=5, value=oficina.telefono or '')
            ws.cell(row=row, column=6, value=oficina.email or '')
            ws.cell(row=row, column=7, value=oficina.ubicacion or '')
            ws.cell(row=row, column=8, value=oficina.descripcion or '')
            ws.cell(row=row, column=9, value='Activa' if oficina.estado else 'Inactiva')
            ws.cell(row=row, column=10, value=oficina.bienes_count)
            ws.cell(row=row, column=11, value=oficina.created_at.strftime('%d/%m/%Y %H:%M'))
        
        # Ajustar ancho de columnas
        column_widths = [15, 30, 25, 20, 15, 25, 20, 30, 12, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Agregar información del reporte
        ws.insert_rows(1)
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"REPORTE DE OFICINAS - DRTC PUNO - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Agregar información de filtros si los hay
        filtros_info = []
        if estado_filtro:
            filtros_info.append(f"Estado: {'Activas' if estado_filtro == 'True' else 'Inactivas'}")
        if busqueda:
            filtros_info.append(f"Búsqueda: {busqueda}")
        if responsable_filtro:
            filtros_info.append(f"Responsable: {responsable_filtro}")
        
        if filtros_info:
            ws.insert_rows(2)
            ws.merge_cells('A2:K2')
            filter_cell = ws['A2']
            filter_cell.value = f"Filtros aplicados: {' | '.join(filtros_info)}"
            filter_cell.font = Font(italic=True)
            filter_cell.alignment = Alignment(horizontal="center")
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo con fecha y hora
        filename = f"oficinas_drtc_puno_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar el archivo en la respuesta
        wb.save(response)
        
        return response
        
    except ImportError:
        messages.error(request, 'Error: La librería openpyxl no está instalada.')
        return redirect('oficinas:lista')
    except Exception as e:
        messages.error(request, f'Error al generar el archivo Excel: {str(e)}')
        return redirect('oficinas:lista')