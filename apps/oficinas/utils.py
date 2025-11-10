import openpyxl
import unicodedata
from django.core.exceptions import ValidationError
from django.db import transaction
from .models import Oficina


class OficinaImporter:
    """Clase para importar oficinas desde archivos Excel"""
    
    COLUMNAS_REQUERIDAS = [
        'CODIGO',
        'NOMBRE',
        'RESPONSABLE'
    ]
    
    COLUMNAS_OPCIONALES = [
        'DESCRIPCION',
        'CARGO_RESPONSABLE',
        'TELEFONO',
        'EMAIL',
        'UBICACION',
        'ESTADO'
    ]
    
    COLUMNAS_ALTERNATIVAS = {
        'CODIGO': ['CÓDIGO', 'COD_OFICINA', 'CODIGO_OFICINA', 'Código', 'CODIGO OFICINA', 'COD OFICINA'],
        'NOMBRE': ['NOMBRE_OFICINA', 'OFICINA', 'DENOMINACION', 'Nombre', 'NOMBRE OFICINA', 'DENOMINACIÓN', 'NOMBRE DE OFICINA', 'NOMBRE DE LA OFICINA'],
        'RESPONSABLE': ['RESPONSABLE_OFICINA', 'ENCARGADO', 'JEFE', 'Responsable', 'RESPONSABLE OFICINA', 'ENCARGADO DE OFICINA', 'JEFE DE OFICINA', 'FUNCIONARIO RESPONSABLE'],
        'DESCRIPCION': ['DESCRIPCIÓN', 'DESC', 'DETALLE', 'Descripción', 'DESCRIPCION OFICINA', 'DESCRIPCIÓN OFICINA'],
        'CARGO_RESPONSABLE': ['CARGO', 'PUESTO', 'CARGO_ENCARGADO', 'Cargo Responsable', 'Cargo del Responsable', 'CARGO DEL RESPONSABLE', 'PUESTO DEL RESPONSABLE'],
        'TELEFONO': ['TELÉFONO', 'TEL', 'CELULAR', 'Teléfono', 'TELEFONO OFICINA', 'TELÉFONO OFICINA', 'CONTACTO'],
        'EMAIL': ['CORREO', 'CORREO_ELECTRONICO', 'E-MAIL', 'Email', 'CORREO ELECTRONICO', 'CORREO ELECTRÓNICO', 'EMAIL OFICINA'],
        'UBICACION': ['UBICACIÓN', 'DIRECCION', 'DIRECCIÓN', 'Ubicación', 'UBICACION OFICINA', 'UBICACIÓN OFICINA', 'DIRECCION OFICINA'],
        'ESTADO': ['ACTIVO', 'VIGENTE', 'STATUS', 'Estado', 'ESTADO OFICINA', 'ACTIVA', 'INACTIVA']
    }
    
    @staticmethod
    def normalizar_texto(texto):
        """Normaliza texto removiendo acentos, espacios extra, asteriscos y convirtiendo a mayúsculas"""
        if not texto:
            return ''
        
        # Convertir a string y limpiar
        texto = str(texto).strip()
        
        # Remover asteriscos y otros caracteres especiales
        texto = texto.replace('*', '').replace(':', '').strip()
        
        # Remover acentos
        texto_sin_acentos = unicodedata.normalize('NFD', texto)
        texto_sin_acentos = ''.join(c for c in texto_sin_acentos if unicodedata.category(c) != 'Mn')
        
        # Convertir a mayúsculas y reemplazar espacios/guiones por underscore
        texto_normalizado = texto_sin_acentos.upper().replace(' ', '_').replace('-', '_')
        
        return texto_normalizado
    
    def __init__(self):
        self.errores = []
        self.warnings = []
        self.registros_procesados = 0
        self.registros_creados = 0
        self.registros_actualizados = 0
        self.fila_encabezados = 1  # Por defecto fila 1
    
    def detectar_fila_encabezados(self, sheet):
        """Detecta automáticamente si los encabezados están en la fila 1, 2 o 3"""
        # Función para contar coincidencias de columnas requeridas
        def contar_coincidencias(headers):
            coincidencias = 0
            for col_requerida in self.COLUMNAS_REQUERIDAS:
                col_normalizada = self.normalizar_texto(col_requerida)
                
                # Buscar columna exacta (normalizada)
                for header in headers:
                    if self.normalizar_texto(header) == col_normalizada:
                        coincidencias += 1
                        break
                else:
                    # Buscar alternativas (normalizadas)
                    if col_requerida in self.COLUMNAS_ALTERNATIVAS:
                        for alternativa in self.COLUMNAS_ALTERNATIVAS[col_requerida]:
                            alternativa_normalizada = self.normalizar_texto(alternativa)
                            for header in headers:
                                if self.normalizar_texto(header) == alternativa_normalizada:
                                    coincidencias += 1
                                    break
                            else:
                                continue
                            break
            return coincidencias
        
        # Intentar con las primeras 3 filas
        mejor_fila = 1
        mejor_headers = []
        mejor_coincidencias = 0
        
        for fila_num in range(1, min(4, sheet.max_row + 1)):  # Probar filas 1, 2 y 3
            headers = []
            for cell in sheet[fila_num]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if headers:
                coincidencias = contar_coincidencias(headers)
                
                # Si encontramos todas las columnas requeridas, usar esta fila
                if coincidencias >= len(self.COLUMNAS_REQUERIDAS):
                    return fila_num, headers
                
                # Guardar la mejor opción hasta ahora
                if coincidencias > mejor_coincidencias:
                    mejor_coincidencias = coincidencias
                    mejor_fila = fila_num
                    mejor_headers = headers
        
        # Retornar la mejor opción encontrada
        return mejor_fila, mejor_headers

    def validar_archivo(self, archivo_path):
        """Valida que el archivo Excel tenga la estructura correcta"""
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Detectar automáticamente la fila de encabezados
            self.fila_encabezados, headers = self.detectar_fila_encabezados(sheet)
            
            # Verificar columnas requeridas
            columnas_encontradas = {}
            
            # Buscar columnas requeridas
            for col_requerida in self.COLUMNAS_REQUERIDAS:
                encontrada = False
                col_normalizada = self.normalizar_texto(col_requerida)
                
                # Buscar columna exacta (normalizada)
                for header in headers:
                    if self.normalizar_texto(header) == col_normalizada:
                        columnas_encontradas[col_requerida] = header
                        encontrada = True
                        break
                
                # Buscar alternativas (normalizadas)
                if not encontrada and col_requerida in self.COLUMNAS_ALTERNATIVAS:
                    for alternativa in self.COLUMNAS_ALTERNATIVAS[col_requerida]:
                        alternativa_normalizada = self.normalizar_texto(alternativa)
                        for header in headers:
                            if self.normalizar_texto(header) == alternativa_normalizada:
                                columnas_encontradas[col_requerida] = header
                                encontrada = True
                                break
                        if encontrada:
                            break
                
                if not encontrada:
                    self.errores.append(f"Columna requerida no encontrada: {col_requerida}")
            
            # Buscar columnas opcionales
            for col_opcional in self.COLUMNAS_OPCIONALES:
                encontrada = False
                col_normalizada = self.normalizar_texto(col_opcional)
                
                # Buscar columna exacta (normalizada)
                for header in headers:
                    if self.normalizar_texto(header) == col_normalizada:
                        columnas_encontradas[col_opcional] = header
                        encontrada = True
                        break
                
                # Buscar alternativas (normalizadas)
                if not encontrada and col_opcional in self.COLUMNAS_ALTERNATIVAS:
                    for alternativa in self.COLUMNAS_ALTERNATIVAS[col_opcional]:
                        alternativa_normalizada = self.normalizar_texto(alternativa)
                        for header in headers:
                            if self.normalizar_texto(header) == alternativa_normalizada:
                                columnas_encontradas[col_opcional] = header
                                encontrada = True
                                break
                        if encontrada:
                            break
            
            if self.errores:
                return False, columnas_encontradas
            
            # Agregar información sobre la fila de encabezados detectada
            self.warnings.append(f"Encabezados detectados en la fila {self.fila_encabezados}")
            
            return True, columnas_encontradas
            
        except Exception as e:
            self.errores.append(f"Error al leer el archivo: {str(e)}")
            return False, {}
    
    def procesar_archivo(self, archivo_path, actualizar_existentes=False):
        """Procesa el archivo Excel e importa los datos"""
        # Validar archivo
        es_valido, columnas_map = self.validar_archivo(archivo_path)
        if not es_valido:
            return self.generar_reporte()
        
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener índices de columnas usando la fila de encabezados detectada
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[self.fila_encabezados]]
            indices_columnas = {}
            
            for col_name, header_encontrado in columnas_map.items():
                try:
                    indices_columnas[col_name] = headers.index(header_encontrado)
                except ValueError:
                    # Columna opcional no encontrada
                    if col_name not in self.COLUMNAS_REQUERIDAS:
                        continue
                    self.errores.append(f"Error al encontrar índice de columna: {header_encontrado}")
                    return self.generar_reporte()
            
            # Procesar filas de datos (empezar después de la fila de encabezados)
            fila_inicio_datos = self.fila_encabezados + 1
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=fila_inicio_datos), start=fila_inicio_datos):
                    try:
                        self.procesar_fila(row, indices_columnas, row_num, actualizar_existentes)
                    except Exception as e:
                        self.errores.append(f"Error en fila {row_num}: {str(e)}")
                        continue
            
        except Exception as e:
            self.errores.append(f"Error general al procesar archivo: {str(e)}")
        
        return self.generar_reporte()
    
    def generar_preview(self, archivo_path, max_filas=10):
        """Genera un preview de los datos que se van a importar"""
        try:
            # Validar archivo primero
            es_valido, columnas_map = self.validar_archivo(archivo_path)
            if not es_valido:
                return {
                    'exito': False,
                    'errores': self.errores,
                    'warnings': self.warnings
                }
            
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener índices de columnas usando la fila de encabezados detectada
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[self.fila_encabezados]]
            indices_columnas = {}
            
            for col_name, header_encontrado in columnas_map.items():
                try:
                    indices_columnas[col_name] = headers.index(header_encontrado)
                except ValueError:
                    continue
            
            # Generar preview de datos
            preview_data = []
            fila_inicio_datos = self.fila_encabezados + 1
            filas_procesadas = 0
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=fila_inicio_datos), start=fila_inicio_datos):
                if filas_procesadas >= max_filas:
                    break
                
                # Extraer datos de la fila
                datos_fila = {}
                tiene_datos = False
                
                for col_name, col_index in indices_columnas.items():
                    if col_index < len(row):
                        cell_value = row[col_index].value
                        valor = str(cell_value).strip() if cell_value else ''
                        datos_fila[col_name] = valor
                        if valor:
                            tiene_datos = True
                    else:
                        datos_fila[col_name] = ''
                
                # Solo agregar filas que tengan al menos un dato
                if tiene_datos:
                    datos_fila['_fila_numero'] = row_num
                    preview_data.append(datos_fila)
                    filas_procesadas += 1
            
            return {
                'exito': True,
                'fila_encabezados': self.fila_encabezados,
                'columnas_detectadas': columnas_map,
                'total_filas_datos': sheet.max_row - self.fila_encabezados,
                'preview_data': preview_data,
                'warnings': self.warnings
            }
            
        except Exception as e:
            return {
                'exito': False,
                'errores': [f'Error al generar preview: {str(e)}']
            }
    
    def procesar_fila(self, row, indices_columnas, row_num, actualizar_existentes):
        """Procesa una fila individual del Excel"""
        # Extraer datos de la fila
        datos = {}
        tiene_algun_dato = False
        
        for col_name, col_index in indices_columnas.items():
            if col_index < len(row):
                cell_value = row[col_index].value
                valor = str(cell_value).strip() if cell_value else ''
                datos[col_name] = valor
                if valor:  # Si hay algún valor, la fila tiene datos
                    tiene_algun_dato = True
            else:
                datos[col_name] = ''
        
        # Si la fila está completamente vacía, omitirla sin generar warning
        if not tiene_algun_dato:
            return
        
        # Validar datos requeridos (solo si la fila tiene algún dato)
        if not datos.get('CODIGO') or not datos.get('NOMBRE') or not datos.get('RESPONSABLE'):
            self.warnings.append(f"Fila {row_num}: Código, nombre o responsable vacíos, omitida")
            return
        
        # Normalizar datos
        codigo = datos['CODIGO'].upper()
        nombre = datos['NOMBRE'].strip()
        responsable = datos['RESPONSABLE'].strip()
        descripcion = datos.get('DESCRIPCION', '')
        cargo_responsable = datos.get('CARGO_RESPONSABLE', '')
        telefono = datos.get('TELEFONO', '')
        email = datos.get('EMAIL', '')
        ubicacion = datos.get('UBICACION', '')
        
        # Procesar estado
        estado_texto = datos.get('ESTADO', 'ACTIVO').upper()
        estado = True  # Default activo
        if estado_texto in ['INACTIVO', 'INACTIVA', 'FALSE', '0', 'NO']:
            estado = False
        
        # Verificar si ya existe
        oficina_existente = None
        try:
            oficina_existente = Oficina.objects.get(codigo=codigo)
        except Oficina.DoesNotExist:
            pass
        
        if oficina_existente:
            if actualizar_existentes:
                # Actualizar existente
                oficina_existente.nombre = nombre
                oficina_existente.descripcion = descripcion
                oficina_existente.responsable = responsable
                oficina_existente.cargo_responsable = cargo_responsable
                oficina_existente.telefono = telefono
                oficina_existente.email = email
                oficina_existente.ubicacion = ubicacion
                oficina_existente.estado = estado
                oficina_existente.save()
                self.registros_actualizados += 1
            else:
                self.warnings.append(f"Fila {row_num}: Código {codigo} ya existe, omitido")
        else:
            # Crear nuevo
            try:
                Oficina.objects.create(
                    codigo=codigo,
                    nombre=nombre,
                    descripcion=descripcion,
                    responsable=responsable,
                    cargo_responsable=cargo_responsable,
                    telefono=telefono,
                    email=email,
                    ubicacion=ubicacion,
                    estado=estado
                )
                self.registros_creados += 1
            except Exception as e:
                self.errores.append(f"Fila {row_num}: Error al crear registro: {str(e)}")
                return
        
        self.registros_procesados += 1
    
    def generar_reporte(self):
        """Genera un reporte del proceso de importación"""
        return {
            'exito': len(self.errores) == 0,
            'registros_procesados': self.registros_procesados,
            'registros_creados': self.registros_creados,
            'registros_actualizados': self.registros_actualizados,
            'errores': self.errores,
            'warnings': self.warnings,
            'resumen': f"Procesados: {self.registros_procesados}, "
                      f"Creados: {self.registros_creados}, "
                      f"Actualizados: {self.registros_actualizados}, "
                      f"Errores: {len(self.errores)}, "
                      f"Advertencias: {len(self.warnings)}"
        }


def importar_oficinas_desde_excel(archivo_path, actualizar_existentes=False):
    """Función helper para importar oficinas"""
    importer = OficinaImporter()
    return importer.procesar_archivo(archivo_path, actualizar_existentes)


def validar_estructura_oficinas(archivo_path):
    """Función helper para validar estructura del archivo"""
    importer = OficinaImporter()
    es_valido, columnas_map = importer.validar_archivo(archivo_path)
    return {
        'es_valido': es_valido,
        'columnas_encontradas': columnas_map,
        'fila_encabezados': getattr(importer, 'fila_encabezados', 1),
        'errores': importer.errores,
        'warnings': importer.warnings
    }


def generar_preview_oficinas(archivo_path, max_filas=10):
    """Función helper para generar preview de datos"""
    importer = OficinaImporter()
    return importer.generar_preview(archivo_path, max_filas)


def generar_plantilla_oficinas():
    """Genera una plantilla Excel vacía para importar oficinas"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from datetime import datetime
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Oficinas"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        instruction_font = Font(bold=True, color="D35400")
        instruction_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título principal
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = f"PLANTILLA PARA IMPORTAR OFICINAS - DRTC PUNO"
        title_cell.font = Font(bold=True, size=14, color="2C3E50")
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
        
        # Instrucciones
        ws.merge_cells('A2:K2')
        instruction_cell = ws['A2']
        instruction_cell.value = "INSTRUCCIONES: Complete los datos en las filas siguientes. Los campos marcados con (*) son obligatorios."
        instruction_cell.font = instruction_font
        instruction_cell.alignment = Alignment(horizontal="center")
        instruction_cell.fill = instruction_fill
        
        # Encabezados (fila 3)
        headers = [
            'Código*', 'Nombre*', 'Responsable*', 'Cargo Responsable', 
            'Teléfono', 'Email', 'Ubicación', 'Descripción', 
            'Estado', 'Fecha Registro', 'Total Bienes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ejemplos de datos (fila 4)
        ejemplos = [
            'ADM-001', 'Administración General', 'Juan Pérez García', 'Jefe de Administración',
            '051-123456', 'admin@drtcpuno.gob.pe', 'Piso 2, Oficina 201', 'Oficina encargada de la administración general',
            'ACTIVO', datetime.now().strftime('%d/%m/%Y'), '0'
        ]
        
        for col, ejemplo in enumerate(ejemplos, 1):
            cell = ws.cell(row=4, column=col, value=ejemplo)
            cell.font = Font(italic=True, color="7F8C8D")
            cell.border = border
        
        # Más ejemplos
        ejemplos2 = [
            'FIN-001', 'Finanzas y Contabilidad', 'María García López', 'Contadora',
            '051-654321', 'finanzas@drtcpuno.gob.pe', 'Piso 1, Oficina 105', 'Oficina de finanzas y contabilidad',
            'ACTIVO', datetime.now().strftime('%d/%m/%Y'), '0'
        ]
        
        for col, ejemplo in enumerate(ejemplos2, 1):
            cell = ws.cell(row=5, column=col, value=ejemplo)
            cell.font = Font(italic=True, color="7F8C8D")
            cell.border = border
        
        # Agregar filas vacías para que el usuario complete
        for row in range(6, 16):  # 10 filas vacías
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col, value="")
                cell.border = border
        
        # Ajustar ancho de columnas
        column_widths = [15, 30, 25, 20, 15, 25, 20, 35, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Agregar hoja de instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        
        instrucciones_texto = [
            "INSTRUCCIONES PARA IMPORTAR OFICINAS",
            "",
            "CAMPOS OBLIGATORIOS (*):",
            "• Código: Código único de la oficina (ej: ADM-001, FIN-002)",
            "• Nombre: Nombre completo de la oficina",
            "• Responsable: Nombre completo del responsable de la oficina",
            "",
            "CAMPOS OPCIONALES:",
            "• Cargo Responsable: Puesto o cargo del responsable",
            "• Teléfono: Número de contacto de la oficina",
            "• Email: Correo electrónico de la oficina",
            "• Ubicación: Ubicación física (piso, oficina, etc.)",
            "• Descripción: Descripción de las funciones de la oficina",
            "• Estado: ACTIVO o INACTIVO (por defecto ACTIVO)",
            "",
            "NOTAS IMPORTANTES:",
            "• Los códigos de oficina deben ser únicos",
            "• No elimine ni modifique los encabezados de la fila 3",
            "• Puede eliminar las filas de ejemplo (4 y 5) antes de importar",
            "• El sistema detecta automáticamente variaciones en los nombres de columnas",
            "• Use el botón 'Validar' antes de importar para verificar la estructura",
            "",
            "VARIANTES ACEPTADAS DE NOMBRES DE COLUMNAS:",
            "• Código: CÓDIGO, COD_OFICINA, CODIGO_OFICINA, etc.",
            "• Nombre: NOMBRE_OFICINA, DENOMINACION, OFICINA, etc.",
            "• Responsable: ENCARGADO, JEFE, RESPONSABLE_OFICINA, etc.",
            "",
            f"Plantilla generada el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ]
        
        for row, texto in enumerate(instrucciones_texto, 1):
            cell = ws_instrucciones.cell(row=row, column=1, value=texto)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif texto.startswith("•"):
                cell.font = Font(color="2C3E50")
            elif texto.endswith(":"):
                cell.font = Font(bold=True, color="D35400")
        
        ws_instrucciones.column_dimensions['A'].width = 80
        
        return wb
        
    except Exception as e:
        raise Exception(f'Error al generar plantilla: {str(e)}')