import json
from io import BytesIO
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, Http404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.conf import settings
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import models, transaction
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from .models import BienPatrimonial, HistorialEstado
from .utils import QRCodeGenerator, QRCodeValidator, importar_bienes_desde_excel, exportar_bienes_a_excel
from .forms import BienPatrimonialForm, MovimientoBienForm, BuscarBienForm, ImportarBienesForm
from apps.catalogo.models import Catalogo
from apps.oficinas.models import Oficina


class QRCodeDetailView(DetailView):
    """Vista pública para mostrar información del bien mediante QR"""
    model = BienPatrimonial
    template_name = 'bienes/qr_detail.html'
    context_object_name = 'bien'
    
    def get_object(self):
        qr_code = self.kwargs.get('qr_code')
        try:
            return BienPatrimonial.objects.select_related('catalogo', 'oficina').get(qr_code=qr_code)
        except BienPatrimonial.DoesNotExist:
            raise Http404("Bien patrimonial no encontrado")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_mobile'] = self.request.user_agent.is_mobile if hasattr(self.request, 'user_agent') else False
        context['user_is_admin'] = self.request.user.is_authenticated and (
            self.request.user.is_superuser or 
            self.request.user.has_perm('bienes.change_bienpatrimonial')
        )
        return context


class QRCodeMobileView(DetailView):
    """Vista optimizada para móviles al escanear QR"""
    model = BienPatrimonial
    template_name = 'bienes/qr_mobile.html'
    context_object_name = 'bien'
    
    def get_object(self):
        qr_code = self.kwargs.get('qr_code')
        try:
            return BienPatrimonial.objects.select_related('catalogo', 'oficina').get(qr_code=qr_code)
        except BienPatrimonial.DoesNotExist:
            raise Http404("Bien patrimonial no encontrado")
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_is_admin'] = self.request.user.is_authenticated and (
            self.request.user.is_superuser or 
            self.request.user.has_perm('bienes.change_bienpatrimonial')
        )
        context['estados_bien'] = BienPatrimonial.ESTADOS_BIEN
        return context


class MobileScannerView(LoginRequiredMixin, View):
    """Vista para el escáner QR móvil"""
    template_name = 'bienes/mobile_scanner.html'
    
    def get(self, request):
        return render(request, self.template_name)


class QRScanAPIView(APIView):
    """API endpoint para procesamiento de escaneo QR desde móvil"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        qr_code = request.data.get('qr_code')
        device_info = request.data.get('device_info', {})
        
        if not qr_code:
            return Response({
                'error': 'Código QR requerido',
                'error_code': 'QR_REQUIRED'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar formato del QR
        is_valid, message = QRCodeValidator.validar_formato_qr(qr_code)
        if not is_valid:
            return Response({
                'error': f'Código QR inválido: {message}',
                'error_code': 'QR_INVALID_FORMAT'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            bien = BienPatrimonial.objects.select_related(
                'catalogo', 'oficina', 'created_by'
            ).get(qr_code=qr_code)
            
            # Determinar permisos del usuario
            user_permissions = {
                'can_edit': request.user.has_perm('bienes.change_bienpatrimonial'),
                'can_update_estado': request.user.has_perm('bienes.change_bienpatrimonial'),
                'can_view_history': request.user.has_perm('bienes.view_historialestado'),
                'is_admin': request.user.is_superuser,
                'is_staff': request.user.is_staff
            }
            
            # Obtener historial reciente si tiene permisos
            historial_reciente = []
            if user_permissions['can_view_history']:
                historial_reciente = list(
                    HistorialEstado.objects.filter(bien=bien)
                    .select_related('created_by')
                    .order_by('-created_at')[:5]
                    .values(
                        'estado_anterior', 'estado_nuevo', 'observaciones',
                        'created_at', 'created_by__username', 'ubicacion_gps'
                    )
                )
            
            # Información completa del bien
            bien_data = {
                'id': bien.id,
                'codigo_patrimonial': bien.codigo_patrimonial,
                'codigo_interno': bien.codigo_interno,
                'denominacion': bien.denominacion,
                'estado_bien': bien.estado_bien,
                'estado_bien_texto': bien.estado_bien_texto,
                'marca': bien.marca or '',
                'modelo': bien.modelo or '',
                'color': bien.color or '',
                'serie': bien.serie or '',
                'dimension': bien.dimension or '',
                'placa': bien.placa or '',
                'matricula': bien.matricula or '',
                'nro_motor': bien.nro_motor or '',
                'nro_chasis': bien.nro_chasis or '',
                'observaciones': bien.observaciones or '',
                'oficina': {
                    'id': bien.oficina.id if bien.oficina else None,
                    'nombre': bien.oficina.nombre if bien.oficina else '',
                    'codigo': bien.oficina.codigo if bien.oficina else '',
                    'responsable': bien.oficina.responsable if bien.oficina else ''
                },
                'catalogo': {
                    'id': bien.catalogo.id if bien.catalogo else None,
                    'codigo': bien.catalogo.codigo if bien.catalogo else '',
                    'denominacion': bien.catalogo.denominacion if bien.catalogo else '',
                    'grupo': bien.catalogo.grupo if bien.catalogo else '',
                    'clase': bien.catalogo.clase if bien.catalogo else ''
                },
                'qr_code': bien.qr_code,
                'url_qr': bien.url_qr,
                'created_at': bien.created_at.isoformat() if bien.created_at else None,
                'updated_at': bien.updated_at.isoformat() if bien.updated_at else None,
                'es_vehiculo': bien.es_vehiculo
            }
            
            return Response({
                'success': True,
                'bien': bien_data,
                'permissions': user_permissions,
                'estados_disponibles': [
                    {'codigo': codigo, 'nombre': nombre} 
                    for codigo, nombre in BienPatrimonial.ESTADOS_BIEN
                ],
                'historial_reciente': historial_reciente,
                'scan_info': {
                    'timestamp': timezone.now().isoformat(),
                    'user': request.user.username,
                    'device_info': device_info
                }
            })
            
        except BienPatrimonial.DoesNotExist:
            return Response({
                'error': 'Bien patrimonial no encontrado',
                'error_code': 'BIEN_NOT_FOUND',
                'qr_code': qr_code
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'error': f'Error interno del servidor: {str(e)}',
                'error_code': 'INTERNAL_ERROR'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UpdateEstadoAPIView(APIView):
    """API endpoint para actualizar estado desde móvil"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        # Verificar permisos
        if not request.user.has_perm('bienes.change_bienpatrimonial'):
            return Response({
                'error': 'No tiene permisos para actualizar bienes',
                'error_code': 'PERMISSION_DENIED'
            }, status=status.HTTP_403_FORBIDDEN)
        
        bien_id = request.data.get('bien_id')
        nuevo_estado = request.data.get('estado')
        observaciones = request.data.get('observaciones', '')
        ubicacion_gps = request.data.get('ubicacion_gps', '')
        device_info = request.data.get('device_info', {})
        
        if not bien_id or not nuevo_estado:
            return Response({
                'error': 'ID del bien y nuevo estado son requeridos',
                'error_code': 'MISSING_REQUIRED_FIELDS'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar estado
        estados_validos = [codigo for codigo, _ in BienPatrimonial.ESTADOS_BIEN]
        if nuevo_estado not in estados_validos:
            return Response({
                'error': f'Estado inválido. Estados válidos: {", ".join(estados_validos)}',
                'error_code': 'INVALID_STATE',
                'estados_validos': estados_validos
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar observaciones para estados críticos
        estados_criticos = ['M', 'E', 'C']  # Malo, RAEE, Chatarra
        if nuevo_estado in estados_criticos and not observaciones.strip():
            return Response({
                'error': f'Las observaciones son obligatorias para el estado {dict(BienPatrimonial.ESTADOS_BIEN)[nuevo_estado]}',
                'error_code': 'OBSERVATIONS_REQUIRED'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                bien = BienPatrimonial.objects.select_for_update().get(id=bien_id)
                estado_anterior = bien.estado_bien
                
                # Verificar si realmente hay cambio
                if estado_anterior == nuevo_estado:
                    return Response({
                        'success': True,
                        'message': 'El bien ya tiene el estado seleccionado',
                        'warning': 'NO_CHANGE_NEEDED',
                        'bien': {
                            'id': bien.id,
                            'codigo_patrimonial': bien.codigo_patrimonial,
                            'estado_actual': dict(BienPatrimonial.ESTADOS_BIEN)[nuevo_estado]
                        }
                    })
                
                # Actualizar estado
                bien.estado_bien = nuevo_estado
                bien.save(update_fields=['estado_bien', 'updated_at'])
                
                # Registrar en historial
                historial = HistorialEstado.objects.create(
                    bien=bien,
                    estado_anterior=estado_anterior,
                    estado_nuevo=nuevo_estado,
                    observaciones=observaciones,
                    ubicacion_gps=ubicacion_gps,
                    created_by=request.user
                )
                
                # Procesar foto si se envió
                foto_url = None
                if 'foto' in request.FILES:
                    historial.foto = request.FILES['foto']
                    historial.save()
                    foto_url = historial.foto.url if historial.foto else None
                
                # Log de la acción
                import logging
                logger = logging.getLogger('patrimonio')
                logger.info(
                    f'Estado actualizado desde móvil - Usuario: {request.user.username}, '
                    f'Bien: {bien.codigo_patrimonial}, '
                    f'Estado: {estado_anterior} -> {nuevo_estado}, '
                    f'GPS: {ubicacion_gps}, '
                    f'Device: {device_info.get("user_agent", "Unknown")}'
                )
                
                return Response({
                    'success': True,
                    'message': 'Estado actualizado correctamente',
                    'bien': {
                        'id': bien.id,
                        'codigo_patrimonial': bien.codigo_patrimonial,
                        'denominacion': bien.denominacion,
                        'estado_anterior': {
                            'codigo': estado_anterior,
                            'nombre': dict(BienPatrimonial.ESTADOS_BIEN)[estado_anterior]
                        },
                        'estado_nuevo': {
                            'codigo': nuevo_estado,
                            'nombre': dict(BienPatrimonial.ESTADOS_BIEN)[nuevo_estado]
                        },
                        'fecha_actualizacion': historial.created_at.isoformat(),
                        'usuario': request.user.username,
                        'ubicacion_gps': ubicacion_gps,
                        'foto_url': foto_url
                    },
                    'historial_id': historial.id
                })
                
        except BienPatrimonial.DoesNotExist:
            return Response({
                'error': 'Bien patrimonial no encontrado',
                'error_code': 'BIEN_NOT_FOUND'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import logging
            logger = logging.getLogger('patrimonio')
            logger.error(f'Error actualizando estado desde móvil: {str(e)}', exc_info=True)
            
            return Response({
                'error': f'Error interno del servidor: {str(e)}',
                'error_code': 'INTERNAL_ERROR'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class BatchQRScanAPIView(APIView):
    """API endpoint para escaneo masivo de códigos QR"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        qr_codes = request.data.get('qr_codes', [])
        device_info = request.data.get('device_info', {})
        
        if not qr_codes or not isinstance(qr_codes, list):
            return Response({
                'error': 'Lista de códigos QR requerida',
                'error_code': 'QR_LIST_REQUIRED'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if len(qr_codes) > 100:  # Límite de seguridad
            return Response({
                'error': 'Máximo 100 códigos QR por solicitud',
                'error_code': 'TOO_MANY_QR_CODES'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        results = []
        found_count = 0
        not_found_count = 0
        
        for qr_code in qr_codes:
            try:
                # Validar formato
                is_valid, message = QRCodeValidator.validar_formato_qr(qr_code)
                if not is_valid:
                    results.append({
                        'qr_code': qr_code,
                        'status': 'invalid',
                        'error': f'Formato inválido: {message}'
                    })
                    continue
                
                # Buscar bien
                bien = BienPatrimonial.objects.select_related(
                    'catalogo', 'oficina'
                ).get(qr_code=qr_code)
                
                results.append({
                    'qr_code': qr_code,
                    'status': 'found',
                    'bien': {
                        'id': bien.id,
                        'codigo_patrimonial': bien.codigo_patrimonial,
                        'denominacion': bien.denominacion,
                        'estado_bien': bien.estado_bien,
                        'estado_bien_texto': bien.estado_bien_texto,
                        'oficina': bien.oficina.nombre if bien.oficina else '',
                        'marca': bien.marca or '',
                        'modelo': bien.modelo or '',
                        'serie': bien.serie or '',
                        'placa': bien.placa or ''
                    }
                })
                found_count += 1
                
            except BienPatrimonial.DoesNotExist:
                results.append({
                    'qr_code': qr_code,
                    'status': 'not_found',
                    'error': 'Bien no encontrado'
                })
                not_found_count += 1
                
            except Exception as e:
                results.append({
                    'qr_code': qr_code,
                    'status': 'error',
                    'error': str(e)
                })
        
        return Response({
            'success': True,
            'summary': {
                'total_scanned': len(qr_codes),
                'found': found_count,
                'not_found': not_found_count,
                'errors': len(qr_codes) - found_count - not_found_count
            },
            'results': results,
            'scan_info': {
                'timestamp': timezone.now().isoformat(),
                'user': request.user.username,
                'device_info': device_info
            }
        })


class MobileInventoryAPIView(APIView):
    """API endpoint para inventario móvil con filtros"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Parámetros de filtro
        oficina_id = request.GET.get('oficina_id')
        estado = request.GET.get('estado')
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        page_size = min(int(request.GET.get('page_size', 20)), 100)  # Máximo 100
        
        # Construir queryset
        queryset = BienPatrimonial.objects.select_related('catalogo', 'oficina')
        
        if oficina_id:
            queryset = queryset.filter(oficina_id=oficina_id)
        
        if estado:
            queryset = queryset.filter(estado_bien=estado)
        
        if search:
            queryset = queryset.filter(
                models.Q(codigo_patrimonial__icontains=search) |
                models.Q(catalogo__denominacion__icontains=search) |
                models.Q(marca__icontains=search) |
                models.Q(modelo__icontains=search) |
                models.Q(serie__icontains=search) |
                models.Q(placa__icontains=search)
            )
        
        # Paginación
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        bienes = queryset.order_by('codigo_patrimonial')[start:end]
        
        # Serializar datos
        bienes_data = []
        for bien in bienes:
            bienes_data.append({
                'id': bien.id,
                'codigo_patrimonial': bien.codigo_patrimonial,
                'denominacion': bien.denominacion,
                'estado_bien': bien.estado_bien,
                'estado_bien_texto': bien.estado_bien_texto,
                'marca': bien.marca or '',
                'modelo': bien.modelo or '',
                'serie': bien.serie or '',
                'placa': bien.placa or '',
                'oficina': bien.oficina.nombre if bien.oficina else '',
                'qr_code': bien.qr_code,
                'url_qr': bien.url_qr
            })
        
        return Response({
            'success': True,
            'pagination': {
                'page': page,
                'page_size': page_size,
                'total': total,
                'total_pages': (total + page_size - 1) // page_size,
                'has_next': end < total,
                'has_previous': page > 1
            },
            'bienes': bienes_data,
            'filters_applied': {
                'oficina_id': oficina_id,
                'estado': estado,
                'search': search
            }
        })


class QRImageView(LoginRequiredMixin, View):
    """Vista para generar imagen del código QR"""
    
    def get(self, request, pk):
        bien = get_object_or_404(BienPatrimonial, pk=pk)
        
        # Generar imagen QR
        generator = QRCodeGenerator()
        img = generator.generar_imagen_qr(bien.url_qr)
        
        # Convertir a bytes
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='image/png')
        response['Content-Disposition'] = f'inline; filename="qr_{bien.codigo_patrimonial}.png"'
        return response


class RegenerarQRView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para regenerar códigos QR masivamente"""
    permission_required = 'bienes.change_bienpatrimonial'
    
    def post(self, request):
        generator = QRCodeGenerator()
        
        # Obtener bienes sin QR o con QR inválido
        bienes_sin_qr = BienPatrimonial.objects.filter(
            models.Q(qr_code='') | models.Q(qr_code__isnull=True) |
            models.Q(url_qr='') | models.Q(url_qr__isnull=True)
        )
        
        contador = generator.regenerar_qr_masivo(bienes_sin_qr)
        
        messages.success(request, f'Se regeneraron {contador} códigos QR correctamente.')
        return redirect('bienes:list')


# Vistas básicas CRUD
class BienListView(LoginRequiredMixin, ListView):
    """Vista para listar bienes patrimoniales"""
    model = BienPatrimonial
    template_name = 'bienes/list.html'
    context_object_name = 'bienes'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = BienPatrimonial.objects.select_related('catalogo', 'oficina')
        
        # Filtros
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(codigo_patrimonial__icontains=search) |
                models.Q(catalogo__denominacion__icontains=search) |
                models.Q(marca__icontains=search) |
                models.Q(modelo__icontains=search) |
                models.Q(serie__icontains=search) |
                models.Q(placa__icontains=search)
            )
        
        estado = self.request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado_bien=estado)
        
        oficina = self.request.GET.get('oficina')
        if oficina:
            queryset = queryset.filter(oficina_id=oficina)
        
        return queryset.order_by('codigo_patrimonial')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estados_bien'] = BienPatrimonial.ESTADOS_BIEN
        context['oficinas'] = Oficina.objects.filter(estado=True).order_by('nombre')
        context['search'] = self.request.GET.get('search', '')
        context['estado_selected'] = self.request.GET.get('estado', '')
        context['oficina_selected'] = self.request.GET.get('oficina', '')
        return context


class BienDetailView(LoginRequiredMixin, DetailView):
    """Vista para mostrar detalle de un bien patrimonial"""
    model = BienPatrimonial
    template_name = 'bienes/detail.html'
    context_object_name = 'bien'
    
    def get_queryset(self):
        return BienPatrimonial.objects.select_related('catalogo', 'oficina')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['historial_estados'] = HistorialEstado.objects.filter(
            bien=self.object
        ).order_by('-created_at')[:10]
        return context


class BienCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Vista para crear un nuevo bien patrimonial"""
    model = BienPatrimonial
    form_class = BienPatrimonialForm
    template_name = 'bienes/form.html'
    permission_required = 'bienes.add_bienpatrimonial'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Bien patrimonial creado correctamente.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('bienes:detail', kwargs={'pk': self.object.pk})


class BienUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Vista para actualizar un bien patrimonial"""
    model = BienPatrimonial
    form_class = BienPatrimonialForm
    template_name = 'bienes/form.html'
    permission_required = 'bienes.change_bienpatrimonial'
    
    def form_valid(self, form):
        messages.success(self.request, 'Bien patrimonial actualizado correctamente.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('bienes:detail', kwargs={'pk': self.object.pk})


class BienDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Vista para eliminar un bien patrimonial (soft delete)"""
    model = BienPatrimonial
    template_name = 'bienes/confirm_delete.html'
    permission_required = 'bienes.delete_bienpatrimonial'
    success_url = reverse_lazy('bienes:list')
    
    def delete(self, request, *args, **kwargs):
        """Sobrescribe delete para usar soft delete"""
        self.object = self.get_object()
        success_url = self.get_success_url()
        
        # Obtener motivo de eliminación si se proporciona
        deletion_reason = request.POST.get('deletion_reason', 'Eliminación desde interfaz web')
        
        # Usar soft delete
        self.object.soft_delete(user=request.user, reason=deletion_reason)
        
        # Crear entrada en RecycleBin
        from apps.core.models import RecycleBin, RecycleBinConfig
        from django.contrib.contenttypes.models import ContentType
        from django.utils import timezone
        from datetime import timedelta
        
        content_type = ContentType.objects.get_for_model(BienPatrimonial)
        config = RecycleBinConfig.get_config_for_module('bienes')
        auto_delete_at = timezone.now() + timedelta(days=config.retention_days)
        
        RecycleBin.objects.create(
            content_type=content_type,
            object_id=self.object.id,
            object_repr=f"{self.object.codigo_patrimonial} - {self.object.denominacion}",
            module_name='bienes',
            deleted_by=request.user,
            deletion_reason=deletion_reason,
            auto_delete_at=auto_delete_at,
            original_data={
                'codigo_patrimonial': self.object.codigo_patrimonial,
                'denominacion': self.object.denominacion,
                'estado_bien': self.object.estado_bien,
                'marca': self.object.marca,
                'modelo': self.object.modelo,
                'serie': self.object.serie
            }
        )
        
        messages.success(request, 'Bien patrimonial movido a la papelera de reciclaje.')
        return redirect(success_url)


class DescargarPlantillaBienesView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para descargar plantilla de ejemplo para importación de bienes"""
    permission_required = 'bienes.view_bienpatrimonial'
    
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from django.utils import timezone
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Bienes"
        
        # Encabezados
        headers = [
            'CODIGO_PATRIMONIAL', 'CODIGO_CATALOGO', 'DENOMINACION', 'MARCA', 'MODELO',
            'SERIE', 'COLOR', 'DIMENSIONES', 'VALOR_ADQUISICION', 'FECHA_ADQUISICION',
            'ESTADO', 'CODIGO_OFICINA', 'OBSERVACIONES'
        ]
        ws.append(headers)
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos de ejemplo
        ejemplos = [
            ['BP-2024-001', '04220001', 'TRACTOR AGRICOLA JOHN DEERE', 'JOHN DEERE', '5075E',
             'SN123456', 'VERDE', '4.5m x 2.2m x 2.8m', '85000.00', '2024-01-15',
             'bueno', 'OF-001', 'Tractor nuevo para área agrícola'],
            ['BP-2024-002', '05220002', 'COMPUTADORA PERSONAL HP', 'HP', 'ELITEDESK 800 G6',
             'SN789012', 'NEGRO', '35cm x 17cm x 34cm', '3500.00', '2024-02-20',
             'bueno', 'OF-002', 'Computadora para oficina administrativa'],
            ['BP-2024-003', '06220003', 'ESCRITORIO DE MADERA', 'MUEBLES PERU', 'EJECUTIVO-150',
             'N/A', 'CAOBA', '150cm x 75cm x 75cm', '850.00', '2024-03-10',
             'bueno', 'OF-002', 'Escritorio para gerencia'],
            ['BP-2024-004', '07220004', 'VEHICULO AUTOMOVIL TOYOTA', 'TOYOTA', 'HILUX 4X4',
             'VIN-ABC123XYZ', 'BLANCO', '5.3m x 1.8m x 1.8m', '125000.00', '2024-01-05',
             'bueno', 'OF-001', 'Vehículo para transporte de personal'],
            ['BP-2024-005', '08220005', 'IMPRESORA LASER HP', 'HP', 'LASERJET PRO M404DN',
             'SN345678', 'GRIS', '36cm x 36cm x 25cm', '1200.00', '2023-12-15',
             'regular', 'OF-003', 'Impresora para área de documentos'],
        ]
        
        for ejemplo in ejemplos:
            ws.append(ejemplo)
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 18,  # CODIGO_PATRIMONIAL
            'B': 16,  # CODIGO_CATALOGO
            'C': 35,  # DENOMINACION
            'D': 18,  # MARCA
            'E': 20,  # MODELO
            'F': 15,  # SERIE
            'G': 12,  # COLOR
            'H': 20,  # DIMENSIONES
            'I': 18,  # VALOR_ADQUISICION
            'J': 18,  # FECHA_ADQUISICION
            'K': 12,  # ESTADO
            'L': 16,  # CODIGO_OFICINA
            'M': 30,  # OBSERVACIONES
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Agregar instrucciones en una hoja separada
        ws_instrucciones = wb.create_sheet("Instrucciones")
        
        instrucciones = [
            ["INSTRUCCIONES PARA IMPORTAR BIENES PATRIMONIALES"],
            [""],
            ["1. ESTRUCTURA DEL ARCHIVO"],
            ["   El archivo debe contener las siguientes columnas (en cualquier orden):"],
            [""],
            ["   COLUMNAS REQUERIDAS:"],
            ["   - CODIGO_PATRIMONIAL: Código único del bien (ej: BP-2024-001)"],
            ["   - CODIGO_CATALOGO: Código del catálogo SBN (8 dígitos, ej: 04220001)"],
            ["   - DENOMINACION: Nombre descriptivo del bien"],
            ["   - VALOR_ADQUISICION: Valor en soles (formato: 1000.00)"],
            ["   - FECHA_ADQUISICION: Fecha en formato YYYY-MM-DD (ej: 2024-01-15)"],
            ["   - ESTADO: bueno, regular, malo, muy_malo, chatarra, RAEE"],
            ["   - CODIGO_OFICINA: Código de la oficina asignada"],
            [""],
            ["   COLUMNAS OPCIONALES:"],
            ["   - MARCA: Marca del bien"],
            ["   - MODELO: Modelo del bien"],
            ["   - SERIE: Número de serie"],
            ["   - COLOR: Color del bien"],
            ["   - DIMENSIONES: Dimensiones físicas"],
            ["   - OBSERVACIONES: Notas adicionales"],
            [""],
            ["2. REGLAS DE VALIDACIÓN"],
            ["   - Los códigos patrimoniales deben ser únicos"],
            ["   - El código de catálogo debe existir en el sistema"],
            ["   - El código de oficina debe existir en el sistema"],
            ["   - El valor de adquisición debe ser un número positivo"],
            ["   - La fecha debe estar en formato correcto (YYYY-MM-DD)"],
            ["   - El estado debe ser uno de los valores permitidos"],
            [""],
            ["3. ESTADOS PERMITIDOS"],
            ["   - bueno: Bien en buen estado"],
            ["   - regular: Bien en estado regular"],
            ["   - malo: Bien en mal estado"],
            ["   - muy_malo: Bien en muy mal estado"],
            ["   - chatarra: Bien dado de baja como chatarra"],
            ["   - RAEE: Residuo de Aparatos Eléctricos y Electrónicos"],
            [""],
            ["4. FORMATO DE FECHAS"],
            ["   - Use el formato: YYYY-MM-DD"],
            ["   - Ejemplos válidos: 2024-01-15, 2023-12-31, 2024-06-01"],
            ["   - NO use: 15/01/2024, 01-15-2024, 15-01-24"],
            [""],
            ["5. FORMATO DE VALORES"],
            ["   - Use punto (.) como separador decimal"],
            ["   - Ejemplos: 1000.00, 3500.50, 125000.00"],
            ["   - NO use comas ni símbolos de moneda"],
            [""],
            ["6. PROCESO DE IMPORTACIÓN"],
            ["   - Elimine estas filas de ejemplo antes de importar"],
            ["   - Complete sus datos en la hoja 'Plantilla Bienes'"],
            ["   - Verifique que los códigos de catálogo y oficina existan"],
            ["   - Guarde el archivo en formato .xlsx o .xls"],
            ["   - Use el botón 'Validar' antes de importar"],
            ["   - Si la validación es exitosa, proceda con la importación"],
            [""],
            ["7. ACTUALIZACIÓN DE REGISTROS EXISTENTES"],
            ["   - Si marca 'Actualizar registros existentes':"],
            ["     Los bienes con códigos existentes serán actualizados"],
            ["   - Si no marca la opción:"],
            ["     Los bienes con códigos existentes serán omitidos"],
            [""],
            ["8. CÓDIGOS DE EJEMPLO"],
            ["   Código Patrimonial: BP-2024-001 (BP = Bien Patrimonial, 2024 = Año, 001 = Correlativo)"],
            ["   Código Catálogo: 04220001 (debe existir en el catálogo SBN)"],
            ["   Código Oficina: OF-001 (debe existir en el sistema de oficinas)"],
            [""],
            ["9. NOTAS IMPORTANTES"],
            ["   - Asegúrese de que los catálogos y oficinas existan antes de importar"],
            ["   - Los códigos patrimoniales deben seguir el formato de su institución"],
            ["   - Revise los datos de ejemplo para entender el formato correcto"],
            ["   - Use la función de validación para detectar errores antes de importar"],
        ]
        
        for row_data in instrucciones:
            ws_instrucciones.append(row_data)
        
        # Estilo para el título de instrucciones
        ws_instrucciones['A1'].font = Font(bold=True, size=14, color="366092")
        ws_instrucciones.column_dimensions['A'].width = 90
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"plantilla_bienes_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class ImportarBienesView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Vista para importar bienes desde Excel"""
    permission_required = 'bienes.add_bienpatrimonial'
    template_name = 'bienes/importar.html'
    
    def get(self, request):
        context = {
            'total_bienes': BienPatrimonial.objects.count(),
            'bienes_activos': BienPatrimonial.objects.filter(estado_bien='bueno').count(),
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        if 'archivo' not in request.FILES:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return render(request, self.template_name)
        
        archivo = request.FILES['archivo']
        actualizar_existentes = request.POST.get('actualizar_existentes') == 'on'
        
        try:
            # Guardar archivo temporalmente
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                for chunk in archivo.chunks():
                    tmp_file.write(chunk)
                tmp_path = tmp_file.name
            
            # Procesar importación
            resultado = importar_bienes_desde_excel(tmp_path, actualizar_existentes)
            
            # Limpiar archivo temporal
            os.unlink(tmp_path)
            
            if resultado['exito']:
                messages.success(request, f"Importación exitosa: {resultado['resumen']}")
            else:
                messages.error(request, f"Errores en importación: {'; '.join(resultado['errores'])}")
                if resultado['warnings']:
                    messages.warning(request, f"Advertencias: {'; '.join(resultado['warnings'])}")
            
        except Exception as e:
            messages.error(request, f'Error al procesar archivo: {str(e)}')
        
        return render(request, self.template_name, {'resultado': resultado})


class ExportarBienesView(LoginRequiredMixin, View):
    """Vista para exportar bienes a Excel"""
    
    def get(self, request):
        # Aplicar filtros si existen
        queryset = BienPatrimonial.objects.select_related('catalogo', 'oficina')
        
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(codigo_patrimonial__icontains=search) |
                models.Q(catalogo__denominacion__icontains=search) |
                models.Q(marca__icontains=search) |
                models.Q(modelo__icontains=search)
            )
        
        estado = request.GET.get('estado')
        if estado:
            queryset = queryset.filter(estado_bien=estado)
        
        oficina = request.GET.get('oficina')
        if oficina:
            queryset = queryset.filter(oficina_id=oficina)
        
        # Generar archivo Excel
        return exportar_bienes_a_excel(queryset, incluir_qr_url=True)