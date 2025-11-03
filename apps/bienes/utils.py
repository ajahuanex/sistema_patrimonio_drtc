import openpyxl
import qrcode
import uuid
from io import BytesIO
from django.core.exceptions import ValidationError
from django.db import transaction
from django.conf import settings
from django.urls import reverse
from django.db.models import Q
from apps.catalogo.models import Catalogo
from apps.oficinas.models import Oficina
from .models import BienPatrimonial


class QRCodeGenerator:
    """Clase para generar códigos QR únicos para bienes patrimoniales"""
    
    def __init__(self):
        self.base_url = getattr(settings, 'BASE_URL', 'http://localhost:8000')
    
    def generar_codigo_unico(self):
        """Genera un código QR único que no exista en la base de datos"""
        while True:
            codigo = str(uuid.uuid4())
            if not BienPatrimonial.objects.filter(qr_code=codigo).exists():
                return codigo
    
    def generar_url_qr(self, qr_code):
        """Genera la URL específica para un código QR"""
        return f"{self.base_url}/qr/{qr_code}/"
    
    def generar_imagen_qr(self, url_qr, size=10, border=4):
        """Genera la imagen del código QR"""
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=size,
            border=border,
        )
        qr.add_data(url_qr)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        return img
    
    def generar_qr_para_bien(self, bien_patrimonial):
        """Genera código QR y URL para un bien patrimonial específico"""
        if not bien_patrimonial.qr_code:
            bien_patrimonial.qr_code = self.generar_codigo_unico()
        
        if not bien_patrimonial.url_qr:
            bien_patrimonial.url_qr = self.generar_url_qr(bien_patrimonial.qr_code)
        
        return bien_patrimonial.qr_code, bien_patrimonial.url_qr
    
    def regenerar_qr_masivo(self, queryset=None):
        """Regenera códigos QR para múltiples bienes"""
        if queryset is None:
            queryset = BienPatrimonial.objects.filter(
                Q(qr_code='') | Q(qr_code__isnull=True) |
                Q(url_qr='') | Q(url_qr__isnull=True)
            )
        
        contador = 0
        for bien in queryset:
            qr_code, url_qr = self.generar_qr_para_bien(bien)
            bien.save(update_fields=['qr_code', 'url_qr'])
            contador += 1
        
        return contador
    
    def validar_qr_unico(self, qr_code, excluir_id=None):
        """Valida que un código QR sea único"""
        queryset = BienPatrimonial.objects.filter(qr_code=qr_code)
        if excluir_id:
            queryset = queryset.exclude(id=excluir_id)
        return not queryset.exists()


class QRCodeValidator:
    """Clase para validar códigos QR y URLs"""
    
    @staticmethod
    def validar_formato_qr(qr_code):
        """Valida que el código QR tenga el formato correcto"""
        if not qr_code:
            return False, "Código QR vacío"
        
        try:
            # Validar que sea un UUID válido
            uuid.UUID(qr_code)
            return True, "Código QR válido"
        except ValueError:
            return False, "Código QR no tiene formato UUID válido"
    
    @staticmethod
    def validar_url_qr(url_qr):
        """Valida que la URL del QR sea correcta"""
        if not url_qr:
            return False, "URL QR vacía"
        
        base_url = getattr(settings, 'BASE_URL', 'http://localhost:8000')
        if not url_qr.startswith(f"{base_url}/qr/"):
            return False, "URL QR no tiene el formato correcto"
        
        return True, "URL QR válida"
    
    @staticmethod
    def extraer_qr_de_url(url_qr):
        """Extrae el código QR de una URL"""
        try:
            # Formato esperado: http://domain/qr/uuid/
            parts = url_qr.rstrip('/').split('/')
            if len(parts) >= 2 and parts[-2] == 'qr':
                return parts[-1]
        except:
            pass
        return None


class BienPatrimonialImporter:
    """Clase para importar bienes patrimoniales desde archivos Excel"""
    
    COLUMNAS_REQUERIDAS = [
        'CODIGO_PATRIMONIAL',
        'DENOMINACION_BIEN',
        'OFICINA'
    ]
    
    COLUMNAS_OPCIONALES = [
        'CODIGO_INTERNO',
        'ESTADO_BIEN',
        'MARCA',
        'MODELO',
        'COLOR',
        'SERIE',
        'DIMENSION',
        'PLACA',
        'MATRICULAS',
        'NRO_MOTOR',
        'NRO_CHASIS',
        'OBSERVACIONES'
    ]
    
    COLUMNAS_ALTERNATIVAS = {
        'CODIGO_PATRIMONIAL': ['CODIGO PATRIMONIAL', 'COD_PATRIMONIAL', 'PATRIMONIO'],
        'CODIGO_INTERNO': ['CODIGO INTERNO', 'COD_INTERNO', 'INTERNO'],
        'DENOMINACION_BIEN': ['DENOMINACION BIEN', 'DENOMINACION', 'BIEN', 'DESCRIPCION'],
        'OFICINA': ['OFICINA_UBICACION', 'UBICACION', 'DEPENDENCIA'],
        'ESTADO_BIEN': ['ESTADO BIEN', 'ESTADO', 'CONDICION'],
        'MATRICULAS': ['MATRICULA', 'MATRÍCULA'],
        'NRO_MOTOR': ['NRO MOTOR', 'NUMERO_MOTOR', 'MOTOR'],
        'NRO_CHASIS': ['NRO CHASIS', 'NUMERO_CHASIS', 'CHASIS']
    }
    
    ESTADOS_VALIDOS = {
        'N': ['NUEVO', 'N'],
        'B': ['BUENO', 'B', 'BIEN'],
        'R': ['REGULAR', 'R'],
        'M': ['MALO', 'M', 'MAL'],
        'E': ['RAEE', 'E'],
        'C': ['CHATARRA', 'C']
    }
    
    def __init__(self):
        self.errores = []
        self.warnings = []
        self.registros_procesados = 0
        self.registros_creados = 0
        self.registros_actualizados = 0
        self.qr_generados = 0
    
    def validar_archivo(self, archivo_path):
        """Valida que el archivo Excel tenga la estructura correcta"""
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener la primera fila (encabezados)
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Verificar columnas requeridas
            columnas_encontradas = {}
            
            # Buscar columnas requeridas
            for col_requerida in self.COLUMNAS_REQUERIDAS:
                encontrada = False
                
                # Buscar columna exacta
                for header in headers:
                    if header.upper().replace(' ', '_') == col_requerida.upper():
                        columnas_encontradas[col_requerida] = header
                        encontrada = True
                        break
                
                # Buscar alternativas
                if not encontrada and col_requerida in self.COLUMNAS_ALTERNATIVAS:
                    for alternativa in self.COLUMNAS_ALTERNATIVAS[col_requerida]:
                        for header in headers:
                            if header.upper().replace(' ', '_') == alternativa.upper().replace(' ', '_'):
                                columnas_encontradas[col_requerida] = header
                                encontrada = True
                                break
                        if encontrada:
                            break
                
                if not encontrada:
                    self.errores.append(f"Columna requerida no encontrada: {col_requerida}")
            
            # Buscar columnas opcionales
            for col_opcional in self.COLUMNAS_OPCIONALES:
                encontrada = False
                
                for header in headers:
                    if header.upper().replace(' ', '_') == col_opcional.upper():
                        columnas_encontradas[col_opcional] = header
                        encontrada = True
                        break
                
                if not encontrada and col_opcional in self.COLUMNAS_ALTERNATIVAS:
                    for alternativa in self.COLUMNAS_ALTERNATIVAS[col_opcional]:
                        for header in headers:
                            if header.upper().replace(' ', '_') == alternativa.upper().replace(' ', '_'):
                                columnas_encontradas[col_opcional] = header
                                encontrada = True
                                break
                        if encontrada:
                            break
            
            if self.errores:
                return False, columnas_encontradas
            
            return True, columnas_encontradas
            
        except Exception as e:
            self.errores.append(f"Error al leer el archivo: {str(e)}")
            return False, {}
    
    def procesar_archivo(self, archivo_path, actualizar_existentes=False):
        """Procesa el archivo Excel e importa los datos"""
        # Validar archivo
        es_valido, columnas_map = self.validar_archivo(archivo_path)
        if not es_valido:
            return self.generar_reporte()
        
        try:
            workbook = openpyxl.load_workbook(archivo_path)
            sheet = workbook.active
            
            # Obtener índices de columnas
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            indices_columnas = {}
            
            for col_name, header_encontrado in columnas_map.items():
                try:
                    indices_columnas[col_name] = headers.index(header_encontrado)
                except ValueError:
                    if col_name not in self.COLUMNAS_REQUERIDAS:
                        continue
                    self.errores.append(f"Error al encontrar índice de columna: {header_encontrado}")
                    return self.generar_reporte()
            
            # Procesar filas de datos
            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    try:
                        self.procesar_fila(row, indices_columnas, row_num, actualizar_existentes)
                    except Exception as e:
                        self.errores.append(f"Error en fila {row_num}: {str(e)}")
                        continue
            
        except Exception as e:
            self.errores.append(f"Error general al procesar archivo: {str(e)}")
        
        return self.generar_reporte()
    
    def procesar_fila(self, row, indices_columnas, row_num, actualizar_existentes):
        """Procesa una fila individual del Excel"""
        # Extraer datos de la fila
        datos = {}
        for col_name, col_index in indices_columnas.items():
            if col_index < len(row):
                cell_value = row[col_index].value
                datos[col_name] = str(cell_value).strip() if cell_value else ''
            else:
                datos[col_name] = ''
        
        # Validar datos requeridos
        if not datos.get('CODIGO_PATRIMONIAL'):
            self.warnings.append(f"Fila {row_num}: Código patrimonial vacío, omitida")
            return
        
        # Normalizar datos
        codigo_patrimonial = datos['CODIGO_PATRIMONIAL'].strip()
        codigo_interno = datos.get('CODIGO_INTERNO', '')
        denominacion_bien = datos.get('DENOMINACION_BIEN', '').upper()
        oficina_nombre = datos.get('OFICINA', '').strip()
        
        # Buscar catálogo por denominación
        catalogo = None
        if denominacion_bien:
            try:
                catalogo = Catalogo.objects.filter(
                    denominacion__icontains=denominacion_bien,
                    estado='ACTIVO'
                ).first()
                
                if not catalogo:
                    # Buscar por coincidencia parcial
                    palabras = denominacion_bien.split()
                    if palabras:
                        catalogo = Catalogo.objects.filter(
                            denominacion__icontains=palabras[0],
                            estado='ACTIVO'
                        ).first()
            except Exception:
                pass
        
        if not catalogo:
            self.warnings.append(f"Fila {row_num}: No se encontró catálogo para '{denominacion_bien}', omitida")
            return
        
        # Buscar oficina
        oficina = None
        if oficina_nombre:
            try:
                oficina = Oficina.objects.filter(
                    Q(nombre__icontains=oficina_nombre) |
                    Q(codigo__icontains=oficina_nombre),
                    estado=True
                ).first()
            except Exception:
                pass
        
        if not oficina:
            self.warnings.append(f"Fila {row_num}: No se encontró oficina para '{oficina_nombre}', omitida")
            return
        
        # Procesar estado
        estado_texto = datos.get('ESTADO_BIEN', 'B').upper()
        estado_bien = 'B'  # Default Bueno
        
        for codigo_estado, variantes in self.ESTADOS_VALIDOS.items():
            if estado_texto in [v.upper() for v in variantes]:
                estado_bien = codigo_estado
                break
        
        # Otros campos
        marca = datos.get('MARCA', '')
        modelo = datos.get('MODELO', '')
        color = datos.get('COLOR', '')
        serie = datos.get('SERIE', '')
        dimension = datos.get('DIMENSION', '')
        placa = datos.get('PLACA', '')
        matricula = datos.get('MATRICULAS', '')
        nro_motor = datos.get('NRO_MOTOR', '')
        nro_chasis = datos.get('NRO_CHASIS', '')
        observaciones = datos.get('OBSERVACIONES', '')
        
        # Verificar si ya existe
        bien_existente = None
        try:
            bien_existente = BienPatrimonial.objects.get(codigo_patrimonial=codigo_patrimonial)
        except BienPatrimonial.DoesNotExist:
            pass
        
        if bien_existente:
            if actualizar_existentes:
                # Actualizar existente
                bien_existente.codigo_interno = codigo_interno
                bien_existente.catalogo = catalogo
                bien_existente.oficina = oficina
                bien_existente.estado_bien = estado_bien
                bien_existente.marca = marca
                bien_existente.modelo = modelo
                bien_existente.color = color
                bien_existente.serie = serie
                bien_existente.dimension = dimension
                bien_existente.placa = placa
                bien_existente.matricula = matricula
                bien_existente.nro_motor = nro_motor
                bien_existente.nro_chasis = nro_chasis
                bien_existente.observaciones = observaciones
                bien_existente.save()
                self.registros_actualizados += 1
            else:
                self.warnings.append(f"Fila {row_num}: Código {codigo_patrimonial} ya existe, omitido")
        else:
            # Crear nuevo
            try:
                bien = BienPatrimonial.objects.create(
                    codigo_patrimonial=codigo_patrimonial,
                    codigo_interno=codigo_interno,
                    catalogo=catalogo,
                    oficina=oficina,
                    estado_bien=estado_bien,
                    marca=marca,
                    modelo=modelo,
                    color=color,
                    serie=serie,
                    dimension=dimension,
                    placa=placa,
                    matricula=matricula,
                    nro_motor=nro_motor,
                    nro_chasis=nro_chasis,
                    observaciones=observaciones
                )
                self.registros_creados += 1
                if bien.qr_code:
                    self.qr_generados += 1
            except Exception as e:
                self.errores.append(f"Fila {row_num}: Error al crear registro: {str(e)}")
                return
        
        self.registros_procesados += 1
    
    def generar_reporte(self):
        """Genera un reporte del proceso de importación"""
        return {
            'exito': len(self.errores) == 0,
            'registros_procesados': self.registros_procesados,
            'registros_creados': self.registros_creados,
            'registros_actualizados': self.registros_actualizados,
            'qr_generados': self.qr_generados,
            'errores': self.errores,
            'warnings': self.warnings,
            'resumen': f"Procesados: {self.registros_procesados}, "
                      f"Creados: {self.registros_creados}, "
                      f"Actualizados: {self.registros_actualizados}, "
                      f"QR generados: {self.qr_generados}, "
                      f"Errores: {len(self.errores)}, "
                      f"Advertencias: {len(self.warnings)}"
        }


def importar_bienes_desde_excel(archivo_path, actualizar_existentes=False):
    """Función helper para importar bienes patrimoniales"""
    importer = BienPatrimonialImporter()
    return importer.procesar_archivo(archivo_path, actualizar_existentes)


def validar_estructura_bienes(archivo_path):
    """Función helper para validar estructura del archivo"""
    importer = BienPatrimonialImporter()
    es_valido, columnas_map = importer.validar_archivo(archivo_path)
    return {
        'es_valido': es_valido,
        'columnas_encontradas': columnas_map,
        'errores': importer.errores
    }


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime


class BienPatrimonialExporter:
    """Clase para exportar bienes patrimoniales a Excel"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Inventario Patrimonial"
        
        # Estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
    
    def exportar_bienes(self, queryset, incluir_qr_url=True):
        """Exporta un queryset de bienes a Excel"""
        # Definir columnas
        columnas = [
            ('A', 'CODIGO PATRIMONIAL', 'codigo_patrimonial'),
            ('B', 'CODIGO INTERNO', 'codigo_interno'),
            ('C', 'DENOMINACION BIEN', 'catalogo__denominacion'),
            ('D', 'ESTADO BIEN', 'get_estado_bien_display'),
            ('E', 'MARCA', 'marca'),
            ('F', 'MODELO', 'modelo'),
            ('G', 'COLOR', 'color'),
            ('H', 'SERIE', 'serie'),
            ('I', 'DIMENSION', 'dimension'),
            ('J', 'PLACA', 'placa'),
            ('K', 'MATRICULAS', 'matricula'),
            ('L', 'NRO MOTOR', 'nro_motor'),
            ('M', 'NRO CHASIS', 'nro_chasis'),
            ('N', 'OFICINA', 'oficina__nombre'),
            ('O', 'OBSERVACIONES', 'observaciones'),
        ]
        
        if incluir_qr_url:
            columnas.append(('P', 'URL QR', 'url_qr'))
        
        # Escribir encabezados
        for col_letter, header, _ in columnas:
            cell = self.worksheet[f'{col_letter}1']
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Escribir datos
        for row_num, bien in enumerate(queryset, start=2):
            for col_letter, _, field_name in columnas:
                cell = self.worksheet[f'{col_letter}{row_num}']
                
                # Obtener valor del campo
                if '__' in field_name:
                    # Campo relacionado
                    parts = field_name.split('__')
                    value = bien
                    for part in parts:
                        if hasattr(value, part):
                            value = getattr(value, part)
                        else:
                            value = ''
                            break
                elif hasattr(bien, field_name):
                    value = getattr(bien, field_name)
                    # Si es un método, llamarlo
                    if callable(value):
                        value = value()
                else:
                    value = ''
                
                cell.value = str(value) if value else ''
        
        # Ajustar ancho de columnas
        for col_letter, _, _ in columnas:
            self.worksheet.column_dimensions[col_letter].width = 15
        
        # Columnas más anchas para ciertos campos
        self.worksheet.column_dimensions['C'].width = 40  # DENOMINACION
        self.worksheet.column_dimensions['N'].width = 25  # OFICINA
        self.worksheet.column_dimensions['O'].width = 30  # OBSERVACIONES
        if incluir_qr_url:
            self.worksheet.column_dimensions['P'].width = 50  # URL QR
        
        return self.workbook
    
    def generar_respuesta_http(self, filename=None):
        """Genera una respuesta HTTP con el archivo Excel"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"inventario_patrimonial_{timestamp}.xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.workbook.save(response)
        return response


def exportar_bienes_a_excel(queryset, incluir_qr_url=True, filename=None):
    """Función helper para exportar bienes a Excel"""
    exporter = BienPatrimonialExporter()
    exporter.exportar_bienes(queryset, incluir_qr_url)
    return exporter.generar_respuesta_http(filename)


def generar_plantilla_importacion():
    """Genera una plantilla Excel para importación de bienes"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Plantilla Importación"
    
    # Encabezados de ejemplo
    headers = [
        'CODIGO PATRIMONIAL',
        'CODIGO INTERNO',
        'DENOMINACION BIEN',
        'ESTADO BIEN',
        'MARCA',
        'MODELO',
        'COLOR',
        'SERIE',
        'DIMENSION',
        'PLACA',
        'MATRICULAS',
        'NRO MOTOR',
        'NRO CHASIS',
        'OFICINA',
        'OBSERVACIONES'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Fila de ejemplo
    ejemplo = [
        'PAT-001-2024',
        'INT-001',
        'ELECTROEYACULADOR PARA BOVINOS',
        'B',
        'MARCA EJEMPLO',
        'MODELO EJEMPLO',
        'AZUL',
        'SER123456',
        '50x30x20 cm',
        'ABC-123',
        'MAT-456',
        'MOT789',
        'CHA012',
        'Dirección Regional',
        'Observaciones del bien'
    ]
    
    for col_num, valor in enumerate(ejemplo, 1):
        worksheet.cell(row=2, column=col_num, value=valor)
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(headers) + 1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_num).column_letter].width = 20
    
    return workbook