"""
Pruebas de flujos completos de usuario
"""
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
import time
import os
from io import BytesIO
from openpyxl import Workbook
from apps.catalogo.models import Catalogo
from apps.oficinas.models import Oficina
from apps.bienes.models import BienPatrimonial
from apps.core.models import UserProfile


class UserRegistrationFlowTest(TestCase):
    """Pruebas de flujo de registro de usuario"""
    
    def setUp(self):
        self.client = Client()
        
        # Crear usuario administrador
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123',
            email='admin@test.com',
            is_staff=True,
            is_superuser=True
        )
        
        # Crear oficina
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
    
    def test_complete_user_registration_flow(self):
        """Prueba flujo completo de registro de usuario"""
        # 1. Login como administrador
        login_success = self.client.login(username='admin', password='admin123')
        self.assertTrue(login_success)
        
        # 2. Acceder a la página de creación de usuario
        response = self.client.get(reverse('core:user_create'))
        self.assertEqual(response.status_code, 200)
        
        # 3. Crear nuevo usuario
        user_data = {
            'username': 'nuevo_funcionario',
            'email': 'funcionario@test.com',
            'first_name': 'Juan',
            'last_name': 'Pérez',
            'password1': 'password123',
            'password2': 'password123',
            'role': 'funcionario',
            'oficina': self.oficina.id,
            'is_active': True
        }
        
        response = self.client.post(reverse('core:user_create'), user_data)
        self.assertEqual(response.status_code, 302)  # Redirect después de crear
        
        # 4. Verificar que el usuario se creó
        new_user = User.objects.get(username='nuevo_funcionario')
        self.assertEqual(new_user.email, 'funcionario@test.com')
        self.assertTrue(hasattr(new_user, 'profile'))
        self.assertEqual(new_user.profile.role, 'funcionario')
        
        # 5. Verificar que el nuevo usuario puede hacer login
        self.client.logout()
        login_success = self.client.login(username='nuevo_funcionario', password='password123')
        self.assertTrue(login_success)
        
        # 6. Verificar acceso a dashboard
        response = self.client.get(reverse('home'))
        self.assertEqual(response.status_code, 200)


class InventoryManagementFlowTest(TestCase):
    """Pruebas de flujo de gestión de inventario"""
    
    def setUp(self):
        self.client = Client()
        
        # Crear usuario funcionario
        self.user = User.objects.create_user(
            username='funcionario',
            password='func123',
            email='funcionario@test.com'
        )
        
        # Crear perfil de usuario
        UserProfile.objects.create(
            user=self.user,
            role='funcionario'
        )
        
        # Crear datos base
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='ELECTROEYACULADOR PARA BOVINOS',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.client.login(username='funcionario', password='func123')
    
    def test_complete_asset_registration_flow(self):
        """Prueba flujo completo de registro de activo"""
        # 1. Acceder a la página de registro
        response = self.client.get(reverse('bienes:create'))
        self.assertEqual(response.status_code, 200)
        
        # 2. Registrar nuevo bien
        bien_data = {
            'codigo_patrimonial': 'PAT-FLOW-001',
            'codigo_interno': 'INT-001',
            'catalogo': self.catalogo.id,
            'oficina': self.oficina.id,
            'estado_bien': 'N',
            'marca': 'MARCA FLOW',
            'modelo': 'MODELO FLOW',
            'color': 'AZUL',
            'serie': 'SER001',
            'dimension': '10x20x30',
            'observaciones': 'Bien registrado en flujo de prueba'
        }
        
        response = self.client.post(reverse('bienes:create'), bien_data)
        self.assertEqual(response.status_code, 302)  # Redirect después de crear
        
        # 3. Verificar que el bien se creó
        bien = BienPatrimonial.objects.get(codigo_patrimonial='PAT-FLOW-001')
        self.assertEqual(bien.marca, 'MARCA FLOW')
        self.assertIsNotNone(bien.qr_code)
        self.assertIsNotNone(bien.url_qr)
        
        # 4. Acceder al detalle del bien
        response = self.client.get(reverse('bienes:detail', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PAT-FLOW-001')
        
        # 5. Actualizar el bien
        update_data = bien_data.copy()
        update_data['estado_bien'] = 'B'
        update_data['observaciones'] = 'Estado actualizado en flujo de prueba'
        
        response = self.client.post(reverse('bienes:update', kwargs={'pk': bien.id}), update_data)
        self.assertEqual(response.status_code, 302)
        
        # 6. Verificar actualización
        bien.refresh_from_db()
        self.assertEqual(bien.estado_bien, 'B')
        
        # 7. Buscar el bien en la lista
        response = self.client.get(reverse('bienes:list'), {'search': 'PAT-FLOW-001'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PAT-FLOW-001')
    
    def test_asset_movement_flow(self):
        """Prueba flujo de movimiento de activo"""
        # Crear bien inicial
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-MOVE-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.user
        )
        
        # Crear oficina destino
        oficina_destino = Oficina.objects.create(
            codigo='ADM-001',
            nombre='Administración',
            responsable='Administrador'
        )
        
        # 1. Acceder a la página de movimiento
        response = self.client.get(reverse('bienes:move', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 200)
        
        # 2. Registrar movimiento
        movement_data = {
            'oficina_destino': oficina_destino.id,
            'motivo': 'Reasignación administrativa',
            'observaciones': 'Movimiento de prueba'
        }
        
        response = self.client.post(reverse('bienes:move', kwargs={'pk': bien.id}), movement_data)
        self.assertEqual(response.status_code, 302)
        
        # 3. Verificar que se creó el movimiento
        from apps.bienes.models import MovimientoBien
        movimiento = MovimientoBien.objects.filter(bien=bien).first()
        self.assertIsNotNone(movimiento)
        self.assertEqual(movimiento.oficina_destino, oficina_destino)
        
        # 4. Confirmar movimiento
        response = self.client.post(reverse('bienes:confirm_movement', kwargs={'pk': movimiento.id}))
        self.assertEqual(response.status_code, 302)
        
        # 5. Verificar que el bien cambió de oficina
        bien.refresh_from_db()
        self.assertEqual(bien.oficina, oficina_destino)


class ImportExportFlowTest(TestCase):
    """Pruebas de flujo de importación y exportación"""
    
    def setUp(self):
        self.client = Client()
        
        # Crear usuario administrador
        self.user = User.objects.create_user(
            username='admin',
            password='admin123',
            email='admin@test.com'
        )
        
        UserProfile.objects.create(
            user=self.user,
            role='administrador'
        )
        
        self.client.login(username='admin', password='admin123')
    
    def create_test_excel_file(self):
        """Crear archivo Excel de prueba"""
        wb = Workbook()
        ws = wb.active
        
        # Encabezados para catálogo
        headers = ['CATALOGO', 'Denominación', 'Grupo', 'Clase', 'Resolución', 'Estado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de prueba
        test_data = [
            ['04220001', 'ELECTROEYACULADOR PARA BOVINOS', '04-AGRÍCOLA Y PESQUERO', '22-EQUIPO', 'R.D. 001-2024', 'ACTIVO'],
            ['04220002', 'BOMBA DE AGUA CENTRIFUGA', '04-AGRÍCOLA Y PESQUERO', '22-EQUIPO', 'R.D. 001-2024', 'ACTIVO']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_catalog_import_export_flow(self):
        """Prueba flujo completo de importación y exportación de catálogo"""
        # 1. Acceder a página de importación
        response = self.client.get(reverse('catalogo:importar'))
        self.assertEqual(response.status_code, 200)
        
        # 2. Importar catálogo
        excel_file = self.create_test_excel_file()
        uploaded_file = SimpleUploadedFile(
            "catalogo_test.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('catalogo:importar'),
            {'archivo_excel': uploaded_file},
            follow=True
        )
        
        self.assertEqual(response.status_code, 200)
        
        # 3. Verificar que se importaron los datos
        self.assertEqual(Catalogo.objects.count(), 2)
        catalogo = Catalogo.objects.get(codigo='04220001')
        self.assertEqual(catalogo.denominacion, 'ELECTROEYACULADOR PARA BOVINOS')
        
        # 4. Acceder a lista de catálogo
        response = self.client.get(reverse('catalogo:list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ELECTROEYACULADOR PARA BOVINOS')
        
        # 5. Exportar catálogo
        response = self.client.get(reverse('catalogo:exportar'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_complete_inventory_import_flow(self):
        """Prueba flujo completo de importación de inventario"""
        # Primero crear catálogo y oficina necesarios
        catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='ELECTROEYACULADOR PARA BOVINOS',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        # Crear archivo Excel para bienes
        wb = Workbook()
        ws = wb.active
        
        headers = [
            'CODIGO PATRIMONIAL', 'CODIGO INTERNO', 'DENOMINACION BIEN', 'ESTADO BIEN',
            'MARCA', 'MODELO', 'COLOR', 'SERIE', 'DIMENSION', 'PLACA', 'MATRICULAS',
            'NRO MOTOR', 'NRO CHASIS', 'OFICINA', 'OBSERVACIONES'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        test_data = [
            ['PAT-IMPORT-001', 'INT-001', 'ELECTROEYACULADOR PARA BOVINOS', 'B', 'MARCA A', 'MODELO A', 'AZUL', 'SER001', '10x20x30', '', '', '', '', 'DIR-001', 'Bien importado']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Importar bienes
        uploaded_file = SimpleUploadedFile(
            "bienes_test.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('bienes:importar'),
            {'archivo_excel': uploaded_file},
            follow=True
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se importó el bien
        bien = BienPatrimonial.objects.get(codigo_patrimonial='PAT-IMPORT-001')
        self.assertEqual(bien.marca, 'MARCA A')
        self.assertIsNotNone(bien.qr_code)


class ReportGenerationFlowTest(TestCase):
    """Pruebas de flujo de generación de reportes"""
    
    def setUp(self):
        self.client = Client()
        
        # Crear usuario
        self.user = User.objects.create_user(
            username='reportuser',
            password='report123',
            email='report@test.com'
        )
        
        UserProfile.objects.create(
            user=self.user,
            role='funcionario'
        )
        
        # Crear datos de prueba
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        # Crear bienes de prueba
        for i in range(10):
            BienPatrimonial.objects.create(
                codigo_patrimonial=f'PAT-RPT-{i:03d}',
                catalogo=self.catalogo,
                oficina=self.oficina,
                estado_bien='B' if i % 2 == 0 else 'R',
                marca=f'MARCA {i}',
                created_by=self.user
            )
        
        self.client.login(username='reportuser', password='report123')
    
    def test_complete_report_generation_flow(self):
        """Prueba flujo completo de generación de reportes"""
        # 1. Acceder al dashboard de reportes
        response = self.client.get(reverse('reportes:dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Dashboard de Reportes')
        
        # 2. Acceder a filtros avanzados
        response = self.client.get(reverse('reportes:filtros_avanzados'))
        self.assertEqual(response.status_code, 200)
        
        # 3. Generar reporte con filtros
        filter_data = {
            'estado_bien': 'B',
            'oficina_id': self.oficina.id,
            'formato': 'excel'
        }
        
        response = self.client.post(reverse('reportes:generar'), filter_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # 4. Generar reporte estadístico
        response = self.client.get(reverse('reportes:estadisticas'))
        self.assertEqual(response.status_code, 200)
        
        # 5. Configurar y generar stickers
        response = self.client.get(reverse('reportes:configurar_stickers'))
        self.assertEqual(response.status_code, 200)
        
        # Generar stickers para algunos bienes
        bienes_ids = list(BienPatrimonial.objects.values_list('id', flat=True)[:3])
        sticker_data = {
            'bienes_ids': bienes_ids,
            'tamano_sticker': 'pequeno',
            'incluir_qr': True,
            'incluir_codigo': True
        }
        
        response = self.client.post(reverse('reportes:generar_stickers'), sticker_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/octet-stream')


class MobileWorkflowTest(TestCase):
    """Pruebas de flujo de trabajo móvil (simulado)"""
    
    def setUp(self):
        self.client = Client()
        
        # Crear usuario móvil
        self.user = User.objects.create_user(
            username='mobileuser',
            password='mobile123',
            email='mobile@test.com'
        )
        
        # Crear datos de prueba
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN MOVIL',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-MOBILE-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.user
        )
    
    def test_mobile_qr_scan_workflow(self):
        """Prueba flujo de escaneo QR móvil"""
        # 1. Acceder a la vista QR pública (sin login)
        response = self.client.get(reverse('bienes:qr_detail', kwargs={'qr_code': self.bien.qr_code}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PAT-MOBILE-001')
        
        # 2. Login de usuario móvil
        self.client.login(username='mobileuser', password='mobile123')
        
        # 3. Acceder a vista móvil con permisos
        response = self.client.get(reverse('bienes:qr_mobile', kwargs={'qr_code': self.bien.qr_code}))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PAT-MOBILE-001')
        
        # 4. Simular actualización desde móvil
        update_data = {
            'estado_bien': 'R',
            'observaciones': 'Actualizado desde móvil',
            'ubicacion_gps': '-15.8422,-70.0199'
        }
        
        response = self.client.post(
            reverse('bienes:mobile_update', kwargs={'pk': self.bien.id}),
            update_data
        )
        self.assertEqual(response.status_code, 302)
        
        # 5. Verificar actualización
        self.bien.refresh_from_db()
        self.assertEqual(self.bien.estado_bien, 'R')
        
        # 6. Verificar que se creó historial
        from apps.bienes.models import HistorialEstado
        historial = HistorialEstado.objects.filter(bien=self.bien).first()
        self.assertIsNotNone(historial)
        self.assertEqual(historial.estado_nuevo, 'R')


class ErrorHandlingFlowTest(TestCase):
    """Pruebas de manejo de errores en flujos"""
    
    def setUp(self):
        self.client = Client()
        
        self.user = User.objects.create_user(
            username='erroruser',
            password='error123',
            email='error@test.com'
        )
        
        self.client.login(username='erroruser', password='error123')
    
    def test_invalid_data_handling(self):
        """Prueba manejo de datos inválidos"""
        # Intentar crear bien sin datos requeridos
        response = self.client.post(reverse('bienes:create'), {})
        self.assertEqual(response.status_code, 200)  # Vuelve al formulario con errores
        
        # Intentar acceder a bien inexistente
        response = self.client.get(reverse('bienes:detail', kwargs={'pk': 99999}))
        self.assertEqual(response.status_code, 404)
    
    def test_permission_denied_handling(self):
        """Prueba manejo de permisos denegados"""
        # Crear usuario sin permisos especiales
        limited_user = User.objects.create_user(
            username='limited',
            password='limited123'
        )
        
        self.client.logout()
        self.client.login(username='limited', password='limited123')
        
        # Intentar acceder a funciones administrativas
        response = self.client.get(reverse('core:user_list'))
        # Debería redirigir o mostrar error de permisos
        self.assertIn(response.status_code, [302, 403])
    
    def test_file_upload_error_handling(self):
        """Prueba manejo de errores en carga de archivos"""
        # Intentar subir archivo inválido
        invalid_file = SimpleUploadedFile(
            "invalid.txt",
            b"contenido invalido",
            content_type="text/plain"
        )
        
        response = self.client.post(
            reverse('catalogo:importar'),
            {'archivo_excel': invalid_file}
        )
        
        # Debería manejar el error graciosamente
        self.assertEqual(response.status_code, 200)
        # Y mostrar mensaje de error
        self.assertContains(response, 'error', status_code=200)