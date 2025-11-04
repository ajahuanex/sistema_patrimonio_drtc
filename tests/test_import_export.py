"""
Pruebas para funcionalidades de importación y exportación
"""
import os
import tempfile
from io import BytesIO
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from openpyxl import Workbook
from apps.catalogo.models import Catalogo
from apps.catalogo.utils import CatalogoImporter
from apps.oficinas.models import Oficina
from apps.oficinas.utils import OficinaImporter
from apps.bienes.models import BienPatrimonial
from apps.bienes.utils import BienPatrimonialImporter, BienPatrimonialExporter
from apps.reportes.exportadores import ExcelExporter, PDFExporter, CSVExporter


class CatalogoImportTest(TestCase):
    """Pruebas para importación de catálogo"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        self.importer = CatalogoImporter()
    
    def create_test_excel_catalogo(self):
        """Crear archivo Excel de prueba para catálogo"""
        wb = Workbook()
        ws = wb.active
        
        # Encabezados
        headers = ['CATALOGO', 'Denominación', 'Grupo', 'Clase', 'Resolución', 'Estado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de prueba
        test_data = [
            ['04220001', 'ELECTROEYACULADOR PARA BOVINOS', '04-AGRÍCOLA Y PESQUERO', '22-EQUIPO', 'R.D. 001-2024', 'ACTIVO'],
            ['04220002', 'BOMBA DE AGUA CENTRIFUGA', '04-AGRÍCOLA Y PESQUERO', '22-EQUIPO', 'R.D. 001-2024', 'ACTIVO'],
            ['05110001', 'COMPUTADORA PERSONAL', '05-INFORMÁTICA Y COMUNICACIONES', '11-EQUIPO', 'R.D. 002-2024', 'ACTIVO']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Guardar en BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_importar_catalogo_exitoso(self):
        """Prueba importación exitosa de catálogo"""
        excel_file = self.create_test_excel_catalogo()
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])
        self.assertEqual(resultado['total_procesados'], 3)
        self.assertEqual(resultado['total_creados'], 3)
        self.assertEqual(resultado['total_errores'], 0)
        
        # Verificar que se crearon los registros
        self.assertEqual(Catalogo.objects.count(), 3)
        
        # Verificar datos específicos
        catalogo = Catalogo.objects.get(codigo='04220001')
        self.assertEqual(catalogo.denominacion, 'ELECTROEYACULADOR PARA BOVINOS')
        self.assertEqual(catalogo.grupo, '04-AGRÍCOLA Y PESQUERO')
    
    def test_importar_catalogo_con_duplicados(self):
        """Prueba importación con códigos duplicados"""
        # Crear catálogo existente
        Catalogo.objects.create(
            codigo='04220001',
            denominacion='BIEN EXISTENTE',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        excel_file = self.create_test_excel_catalogo()
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])
        self.assertEqual(resultado['total_procesados'], 3)
        self.assertEqual(resultado['total_actualizados'], 1)  # El duplicado se actualiza
        self.assertEqual(resultado['total_creados'], 2)
    
    def test_importar_catalogo_archivo_invalido(self):
        """Prueba importación con archivo inválido"""
        # Crear archivo con estructura incorrecta
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='COLUMNA_INCORRECTA')
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertFalse(resultado['success'])
        self.assertIn('error', resultado)
    
    def test_validar_estructura_archivo(self):
        """Prueba validación de estructura de archivo"""
        excel_file = self.create_test_excel_catalogo()
        
        es_valido, mensaje = self.importer.validar_estructura_archivo(excel_file)
        
        self.assertTrue(es_valido)
        self.assertEqual(mensaje, "Estructura válida")


class OficinaImportTest(TestCase):
    """Pruebas para importación de oficinas"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        self.importer = OficinaImporter()
    
    def create_test_excel_oficinas(self):
        """Crear archivo Excel de prueba para oficinas"""
        wb = Workbook()
        ws = wb.active
        
        # Encabezados
        headers = ['Código', 'Nombre', 'Descripción', 'Responsable', 'Cargo Responsable', 'Teléfono', 'Email', 'Ubicación', 'Estado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de prueba
        test_data = [
            ['DIR-001', 'Dirección Regional', 'Oficina principal', 'Juan Pérez', 'Director', '051-123456', 'director@test.com', 'Primer piso', 'ACTIVO'],
            ['ADM-001', 'Administración', 'Área administrativa', 'María García', 'Administradora', '051-123457', 'admin@test.com', 'Segundo piso', 'ACTIVO'],
            ['LOG-001', 'Logística', 'Área de logística', 'Carlos López', 'Jefe de Logística', '051-123458', 'logistica@test.com', 'Primer piso', 'ACTIVO']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_importar_oficinas_exitoso(self):
        """Prueba importación exitosa de oficinas"""
        excel_file = self.create_test_excel_oficinas()
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])
        self.assertEqual(resultado['total_procesados'], 3)
        self.assertEqual(resultado['total_creados'], 3)
        self.assertEqual(resultado['total_errores'], 0)
        
        # Verificar que se crearon los registros
        self.assertEqual(Oficina.objects.count(), 3)
        
        # Verificar datos específicos
        oficina = Oficina.objects.get(codigo='DIR-001')
        self.assertEqual(oficina.nombre, 'Dirección Regional')
        self.assertEqual(oficina.responsable, 'Juan Pérez')
    
    def test_importar_oficinas_con_duplicados(self):
        """Prueba importación con códigos duplicados"""
        # Crear oficina existente
        Oficina.objects.create(
            codigo='DIR-001',
            nombre='Oficina Existente',
            responsable='Responsable Existente'
        )
        
        excel_file = self.create_test_excel_oficinas()
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])
        self.assertEqual(resultado['total_procesados'], 3)
        self.assertEqual(resultado['total_actualizados'], 1)
        self.assertEqual(resultado['total_creados'], 2)


class BienPatrimonialImportTest(TestCase):
    """Pruebas para importación de bienes patrimoniales"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Crear catálogo y oficina necesarios
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='ELECTROEYACULADOR PARA BOVINOS',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.importer = BienPatrimonialImporter()
    
    def create_test_excel_bienes(self):
        """Crear archivo Excel de prueba para bienes"""
        wb = Workbook()
        ws = wb.active
        
        # Encabezados
        headers = [
            'CODIGO PATRIMONIAL', 'CODIGO INTERNO', 'DENOMINACION BIEN', 'ESTADO BIEN',
            'MARCA', 'MODELO', 'COLOR', 'SERIE', 'DIMENSION', 'PLACA', 'MATRICULAS',
            'NRO MOTOR', 'NRO CHASIS', 'OFICINA', 'OBSERVACIONES'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos de prueba
        test_data = [
            ['PAT-001-2024', 'INT-001', 'ELECTROEYACULADOR PARA BOVINOS', 'B', 'MARCA A', 'MODELO A', 'AZUL', 'SER001', '10x20x30', '', '', '', '', 'DIR-001', 'Bien en buen estado'],
            ['PAT-002-2024', 'INT-002', 'ELECTROEYACULADOR PARA BOVINOS', 'N', 'MARCA B', 'MODELO B', 'ROJO', 'SER002', '15x25x35', '', '', '', '', 'DIR-001', 'Bien nuevo'],
            ['PAT-003-2024', 'INT-003', 'ELECTROEYACULADOR PARA BOVINOS', 'R', 'MARCA C', 'MODELO C', 'VERDE', 'SER003', '12x22x32', '', '', '', '', 'DIR-001', 'Requiere mantenimiento']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def test_importar_bienes_exitoso(self):
        """Prueba importación exitosa de bienes"""
        excel_file = self.create_test_excel_bienes()
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])
        self.assertEqual(resultado['total_procesados'], 3)
        self.assertEqual(resultado['total_creados'], 3)
        self.assertEqual(resultado['total_errores'], 0)
        
        # Verificar que se crearon los registros
        self.assertEqual(BienPatrimonial.objects.count(), 3)
        
        # Verificar que se generaron QR codes
        for bien in BienPatrimonial.objects.all():
            self.assertIsNotNone(bien.qr_code)
            self.assertIsNotNone(bien.url_qr)
    
    def test_importar_bienes_con_errores(self):
        """Prueba importación con errores de validación"""
        wb = Workbook()
        ws = wb.active
        
        # Encabezados
        headers = [
            'CODIGO PATRIMONIAL', 'CODIGO INTERNO', 'DENOMINACION BIEN', 'ESTADO BIEN',
            'MARCA', 'MODELO', 'COLOR', 'SERIE', 'DIMENSION', 'PLACA', 'MATRICULAS',
            'NRO MOTOR', 'NRO CHASIS', 'OFICINA', 'OBSERVACIONES'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos con errores
        test_data = [
            ['', 'INT-001', 'BIEN SIN CODIGO', 'B', 'MARCA A', 'MODELO A', 'AZUL', 'SER001', '10x20x30', '', '', '', '', 'DIR-001', ''],  # Sin código patrimonial
            ['PAT-002-2024', 'INT-002', 'BIEN INEXISTENTE', 'B', 'MARCA B', 'MODELO B', 'ROJO', 'SER002', '15x25x35', '', '', '', '', 'OFICINA-INEXISTENTE', ''],  # Oficina inexistente
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        resultado = self.importer.importar_desde_excel(excel_file, self.user)
        
        self.assertTrue(resultado['success'])  # El proceso completa pero con errores
        self.assertEqual(resultado['total_procesados'], 2)
        self.assertEqual(resultado['total_errores'], 2)
        self.assertEqual(resultado['total_creados'], 0)
        
        # Verificar que se registraron los errores
        self.assertGreater(len(resultado['errores']), 0)


class BienPatrimonialExportTest(TestCase):
    """Pruebas para exportación de bienes patrimoniales"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Crear datos de prueba
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='ELECTROEYACULADOR PARA BOVINOS',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        # Crear bienes de prueba
        for i in range(5):
            BienPatrimonial.objects.create(
                codigo_patrimonial=f'PAT-{i:03d}-2024',
                catalogo=self.catalogo,
                oficina=self.oficina,
                estado_bien='B',
                marca=f'MARCA {i}',
                modelo=f'MODELO {i}',
                serie=f'SER{i:03d}',
                created_by=self.user
            )
        
        self.exporter = BienPatrimonialExporter()
    
    def test_exportar_excel_completo(self):
        """Prueba exportación completa a Excel"""
        bienes = BienPatrimonial.objects.all()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            resultado = self.exporter.exportar_excel(bienes, tmp_file.name)
            
            self.assertTrue(resultado['success'])
            self.assertEqual(resultado['total_exportados'], 5)
            self.assertTrue(os.path.exists(tmp_file.name))
            
            # Limpiar archivo temporal
            os.unlink(tmp_file.name)
    
    def test_exportar_excel_con_filtros(self):
        """Prueba exportación con filtros aplicados"""
        # Cambiar estado de algunos bienes
        bienes = BienPatrimonial.objects.all()[:2]
        for bien in bienes:
            bien.estado_bien = 'R'
            bien.save()
        
        # Exportar solo bienes en estado 'B'
        bienes_filtrados = BienPatrimonial.objects.filter(estado_bien='B')
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            resultado = self.exporter.exportar_excel(bienes_filtrados, tmp_file.name)
            
            self.assertTrue(resultado['success'])
            self.assertEqual(resultado['total_exportados'], 3)  # Solo 3 en estado 'B'
            
            os.unlink(tmp_file.name)
    
    def test_exportar_con_urls_qr(self):
        """Prueba que la exportación incluya URLs de QR"""
        bienes = BienPatrimonial.objects.all()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            resultado = self.exporter.exportar_excel(bienes, tmp_file.name, incluir_qr_urls=True)
            
            self.assertTrue(resultado['success'])
            self.assertIn('qr_urls_incluidas', resultado)
            self.assertTrue(resultado['qr_urls_incluidas'])
            
            os.unlink(tmp_file.name)


class ReportesExportTest(TestCase):
    """Pruebas para exportadores de reportes"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        
        # Crear datos de prueba
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-001-2024',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.user
        )
    
    def test_excel_exporter(self):
        """Prueba exportador Excel de reportes"""
        exporter = ExcelExporter()
        bienes = BienPatrimonial.objects.all()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            resultado = exporter.exportar(bienes, tmp_file.name)
            
            self.assertTrue(resultado['success'])
            self.assertTrue(os.path.exists(tmp_file.name))
            self.assertIn('archivo_generado', resultado)
            
            os.unlink(tmp_file.name)
    
    def test_pdf_exporter(self):
        """Prueba exportador PDF de reportes"""
        exporter = PDFExporter()
        bienes = BienPatrimonial.objects.all()
        
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp_file:
            resultado = exporter.exportar(bienes, tmp_file.name)
            
            self.assertTrue(resultado['success'])
            self.assertTrue(os.path.exists(tmp_file.name))
            
            os.unlink(tmp_file.name)
    
    def test_csv_exporter(self):
        """Prueba exportador CSV de reportes"""
        exporter = CSVExporter()
        bienes = BienPatrimonial.objects.all()
        
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as tmp_file:
            resultado = exporter.exportar(bienes, tmp_file.name)
            
            self.assertTrue(resultado['success'])
            self.assertTrue(os.path.exists(tmp_file.name))
            
            os.unlink(tmp_file.name)


class ImportExportIntegrationTest(TestCase):
    """Pruebas de integración para importación y exportación"""
    
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )
        self.client.login(username='testuser', password='testpass123')
        
        # Crear catálogo y oficina
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='ELECTROEYACULADOR PARA BOVINOS',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
    
    def test_flujo_completo_import_export_catalogo(self):
        """Prueba flujo completo de importación y exportación de catálogo"""
        # Crear archivo Excel para importar
        wb = Workbook()
        ws = wb.active
        
        headers = ['CATALOGO', 'Denominación', 'Grupo', 'Clase', 'Resolución', 'Estado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        test_data = [
            ['04220002', 'BOMBA DE AGUA', '04-AGRÍCOLA Y PESQUERO', '22-EQUIPO', 'R.D. 001-2024', 'ACTIVO']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Importar via vista web
        uploaded_file = SimpleUploadedFile(
            "catalogo_test.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('catalogo:importar'),
            {'archivo_excel': uploaded_file},
            follow=True
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se importó
        self.assertTrue(Catalogo.objects.filter(codigo='04220002').exists())
        
        # Exportar catálogo
        response = self.client.get(reverse('catalogo:exportar'))
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_flujo_completo_import_export_bienes(self):
        """Prueba flujo completo de importación y exportación de bienes"""
        # Crear archivo Excel para importar bienes
        wb = Workbook()
        ws = wb.active
        
        headers = [
            'CODIGO PATRIMONIAL', 'CODIGO INTERNO', 'DENOMINACION BIEN', 'ESTADO BIEN',
            'MARCA', 'MODELO', 'COLOR', 'SERIE', 'DIMENSION', 'PLACA', 'MATRICULAS',
            'NRO MOTOR', 'NRO CHASIS', 'OFICINA', 'OBSERVACIONES'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        test_data = [
            ['PAT-TEST-001', 'INT-001', 'ELECTROEYACULADOR PARA BOVINOS', 'B', 'MARCA TEST', 'MODELO TEST', 'AZUL', 'SER001', '10x20x30', '', '', '', '', 'DIR-001', 'Bien de prueba']
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Importar via vista web
        uploaded_file = SimpleUploadedFile(
            "bienes_test.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('bienes:importar'),
            {'archivo_excel': uploaded_file},
            follow=True
        )
        
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se importó
        self.assertTrue(BienPatrimonial.objects.filter(codigo_patrimonial='PAT-TEST-001').exists())
        
        # Verificar que se generó QR
        bien = BienPatrimonial.objects.get(codigo_patrimonial='PAT-TEST-001')
        self.assertIsNotNone(bien.qr_code)
        self.assertIsNotNone(bien.url_qr)
        
        # Exportar bienes
        response = self.client.get(reverse('bienes:exportar'))
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )