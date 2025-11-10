"""
Pruebas de regresión para mantener compatibilidad del sistema de papelera
"""
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.db.models import Q
from apps.core.models import RecycleBin, RecycleBinConfig
from apps.core.services import RecycleBinService
from apps.oficinas.models import Oficina
from apps.bienes.models import BienPatrimonial
from apps.catalogo.models import Catalogo


class BackwardCompatibilityTest(TestCase):
    """Pruebas de compatibilidad hacia atrás"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123',
            is_staff=True
        )
        
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.service = RecycleBinService()
    
    def test_existing_queries_still_work(self):
        """Prueba que consultas existentes siguen funcionando"""
        # Crear bienes activos y eliminados
        bien_activo = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-REG-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        bien_eliminado = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-REG-002',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        self.service.soft_delete_object(bien_eliminado, self.admin_user, 'Test')
        
        # Consultas existentes deberían seguir funcionando
        # 1. Consulta simple
        bienes = BienPatrimonial.objects.all()
        self.assertEqual(bienes.count(), 1)
        self.assertIn(bien_activo, bienes)
        self.assertNotIn(bien_eliminado, bienes)
        
        # 2. Consulta con filtros
        bienes_buenos = BienPatrimonial.objects.filter(estado_bien='B')
        self.assertEqual(bienes_buenos.count(), 1)
        
        # 3. Consulta con select_related
        bienes_con_catalogo = BienPatrimonial.objects.select_related('catalogo').all()
        self.assertEqual(bienes_con_catalogo.count(), 1)
        
        # 4. Consulta con prefetch_related
        bienes_con_oficina = BienPatrimonial.objects.prefetch_related('oficina').all()
        self.assertEqual(bienes_con_oficina.count(), 1)
    
    def test_existing_views_still_work(self):
        """Prueba que vistas existentes siguen funcionando"""
        client = Client()
        client.login(username='admin', password='admin123')
        
        # Crear bien
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-VIEW-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # 1. Vista de lista
        response = client.get(reverse('bienes:list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'PAT-VIEW-001')
        
        # 2. Vista de detalle
        response = client.get(reverse('bienes:detail', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 200)
        
        # 3. Vista de edición
        response = client.get(reverse('bienes:update', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 200)
        
        # 4. Eliminar (ahora soft delete)
        response = client.post(reverse('bienes:delete', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 302)
        
        # Verificar que se eliminó lógicamente
        bien.refresh_from_db()
        self.assertTrue(bien.is_deleted)
        
        # 5. Ya no debería aparecer en lista
        response = client.get(reverse('bienes:list'))
        self.assertNotContains(response, 'PAT-VIEW-001')

    def test_existing_api_endpoints_still_work(self):
        """Prueba que endpoints de API existentes siguen funcionando"""
        from rest_framework.test import APIClient
        from rest_framework_simplejwt.tokens import RefreshToken
        
        client = APIClient()
        
        # Autenticar
        refresh = RefreshToken.for_user(self.admin_user)
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {refresh.access_token}')
        
        # Crear bien
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-API-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # 1. GET list
        response = client.get(reverse('bienes:api_list'))
        self.assertEqual(response.status_code, 200)
        
        # 2. GET detail
        response = client.get(reverse('bienes:api_detail', kwargs={'pk': bien.id}))
        self.assertEqual(response.status_code, 200)
        
        # 3. PUT update
        response = client.put(
            reverse('bienes:api_detail', kwargs={'pk': bien.id}),
            {
                'codigo_patrimonial': 'PAT-API-001',
                'catalogo': self.catalogo.id,
                'oficina': self.oficina.id,
                'estado_bien': 'R',
                'marca': 'UPDATED'
            },
            format='json'
        )
        self.assertEqual(response.status_code, 200)
        
        # 4. DELETE (ahora soft delete)
        response = client.delete(reverse('bienes:api_detail', kwargs={'pk': bien.id}))
        # Debería aceptar la eliminación
        self.assertIn(response.status_code, [200, 204])
        
        # Verificar que se eliminó lógicamente
        bien.refresh_from_db()
        self.assertTrue(bien.is_deleted)
    
    def test_existing_reports_still_work(self):
        """Prueba que reportes existentes siguen funcionando"""
        client = Client()
        client.login(username='admin', password='admin123')
        
        # Crear bienes activos y eliminados
        bien_activo = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-REPORT-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        bien_eliminado = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-REPORT-002',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        self.service.soft_delete_object(bien_eliminado, self.admin_user, 'Test')
        
        # Generar reporte (debería incluir solo activos por defecto)
        response = client.post(
            reverse('reportes:generar'),
            {'formato': 'excel', 'estado_bien': 'B'}
        )
        
        self.assertEqual(response.status_code, 200)
        # El reporte debería generarse correctamente
    
    def test_existing_imports_still_work(self):
        """Prueba que importaciones existentes siguen funcionando"""
        from io import BytesIO
        from openpyxl import Workbook
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        client = Client()
        client.login(username='admin', password='admin123')
        
        # Crear archivo Excel de prueba
        wb = Workbook()
        ws = wb.active
        
        headers = ['CATALOGO', 'Denominación', 'Grupo', 'Clase', 'Resolución', 'Estado']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        data = ['99999999', 'BIEN IMPORTADO', '99-TEST', '99-TEST', 'R.D. 999-2024', 'ACTIVO']
        for col, value in enumerate(data, 1):
            ws.cell(row=2, column=col, value=value)
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        uploaded_file = SimpleUploadedFile(
            "test_import.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Importar
        response = client.post(
            reverse('catalogo:importar'),
            {'archivo_excel': uploaded_file}
        )
        
        # Debería funcionar correctamente
        self.assertEqual(response.status_code, 200)
        
        # Verificar que se importó
        catalogo_importado = Catalogo.objects.filter(codigo='99999999').first()
        self.assertIsNotNone(catalogo_importado)


class ModelMethodsCompatibilityTest(TestCase):
    """Pruebas de compatibilidad de métodos de modelo"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123'
        )
        
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.service = RecycleBinService()
    
    def test_model_save_method_compatibility(self):
        """Prueba compatibilidad del método save()"""
        # Crear bien normalmente
        bien = BienPatrimonial(
            codigo_patrimonial='PAT-SAVE-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # save() debería funcionar normalmente
        bien.save()
        
        self.assertIsNotNone(bien.id)
        self.assertIsNone(bien.deleted_at)
        
        # Actualizar
        bien.marca = 'MARCA ACTUALIZADA'
        bien.save()
        
        bien.refresh_from_db()
        self.assertEqual(bien.marca, 'MARCA ACTUALIZADA')
    
    def test_model_delete_method_compatibility(self):
        """Prueba compatibilidad del método delete()"""
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-DELETE-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # delete() ahora hace soft delete
        bien.delete()
        
        # Debería estar marcado como eliminado
        bien.refresh_from_db()
        self.assertTrue(bien.is_deleted)
        self.assertIsNotNone(bien.deleted_at)
        
        # No debería aparecer en consultas normales
        bienes = BienPatrimonial.objects.all()
        self.assertNotIn(bien, bienes)
    
    def test_model_str_method_compatibility(self):
        """Prueba compatibilidad del método __str__()"""
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-STR-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # __str__() debería funcionar normalmente
        str_representation = str(bien)
        self.assertIsNotNone(str_representation)
        self.assertIn('PAT-STR-001', str_representation)
        
        # Incluso después de eliminar
        self.service.soft_delete_object(bien, self.admin_user, 'Test')
        str_representation_deleted = str(bien)
        self.assertIsNotNone(str_representation_deleted)
    
    def test_model_custom_methods_compatibility(self):
        """Prueba compatibilidad de métodos personalizados"""
        oficina = Oficina.objects.create(
            codigo='CUSTOM-001',
            nombre='Oficina Custom',
            responsable='Responsable'
        )
        
        # Método personalizado puede_eliminarse()
        puede_eliminar = oficina.puede_eliminarse()
        self.assertIsInstance(puede_eliminar, bool)
        
        # Debería funcionar incluso después de soft delete
        self.service.soft_delete_object(oficina, self.admin_user, 'Test')
        
        # Los métodos personalizados deberían seguir funcionando
        # (aunque el objeto esté eliminado lógicamente)
        oficina.refresh_from_db()
        self.assertTrue(oficina.is_deleted)


class QuerySetCompatibilityTest(TestCase):
    """Pruebas de compatibilidad de QuerySets"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123'
        )
        
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.service = RecycleBinService()
        
        # Crear bienes activos y eliminados
        self.bien_activo = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-QS-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        self.bien_eliminado = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-QS-002',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        self.service.soft_delete_object(self.bien_eliminado, self.admin_user, 'Test')

    def test_filter_compatibility(self):
        """Prueba compatibilidad de filter()"""
        # Filtros simples
        bienes = BienPatrimonial.objects.filter(estado_bien='B')
        self.assertEqual(bienes.count(), 1)
        self.assertIn(self.bien_activo, bienes)
        
        # Filtros con Q objects
        bienes = BienPatrimonial.objects.filter(
            Q(estado_bien='B') | Q(estado_bien='R')
        )
        self.assertEqual(bienes.count(), 1)
        
        # Filtros con relaciones
        bienes = BienPatrimonial.objects.filter(oficina=self.oficina)
        self.assertEqual(bienes.count(), 1)
    
    def test_exclude_compatibility(self):
        """Prueba compatibilidad de exclude()"""
        # Excluir por estado
        bienes = BienPatrimonial.objects.exclude(estado_bien='R')
        self.assertIn(self.bien_activo, bienes)
        
        # Excluir por oficina
        bienes = BienPatrimonial.objects.exclude(oficina__codigo='INEXISTENTE')
        self.assertEqual(bienes.count(), 1)
    
    def test_get_compatibility(self):
        """Prueba compatibilidad de get()"""
        # Get por código
        bien = BienPatrimonial.objects.get(codigo_patrimonial='PAT-QS-001')
        self.assertEqual(bien, self.bien_activo)
        
        # Get de bien eliminado debería fallar
        with self.assertRaises(BienPatrimonial.DoesNotExist):
            BienPatrimonial.objects.get(codigo_patrimonial='PAT-QS-002')
        
        # Pero debería funcionar con all_objects
        bien_eliminado = BienPatrimonial.all_objects.get(codigo_patrimonial='PAT-QS-002')
        self.assertEqual(bien_eliminado, self.bien_eliminado)
    
    def test_count_compatibility(self):
        """Prueba compatibilidad de count()"""
        # Count normal
        count = BienPatrimonial.objects.count()
        self.assertEqual(count, 1)
        
        # Count con filtros
        count = BienPatrimonial.objects.filter(estado_bien='B').count()
        self.assertEqual(count, 1)
        
        # Count de todos (incluyendo eliminados)
        count = BienPatrimonial.all_objects.count()
        self.assertEqual(count, 2)
    
    def test_exists_compatibility(self):
        """Prueba compatibilidad de exists()"""
        # Exists de activo
        exists = BienPatrimonial.objects.filter(
            codigo_patrimonial='PAT-QS-001'
        ).exists()
        self.assertTrue(exists)
        
        # Exists de eliminado
        exists = BienPatrimonial.objects.filter(
            codigo_patrimonial='PAT-QS-002'
        ).exists()
        self.assertFalse(exists)
        
        # Pero debería existir en all_objects
        exists = BienPatrimonial.all_objects.filter(
            codigo_patrimonial='PAT-QS-002'
        ).exists()
        self.assertTrue(exists)
    
    def test_order_by_compatibility(self):
        """Prueba compatibilidad de order_by()"""
        # Crear más bienes
        for i in range(5):
            BienPatrimonial.objects.create(
                codigo_patrimonial=f'PAT-ORDER-{i:03d}',
                catalogo=self.catalogo,
                oficina=self.oficina,
                estado_bien='B',
                created_by=self.admin_user
            )
        
        # Order by ascendente
        bienes = BienPatrimonial.objects.order_by('codigo_patrimonial')
        self.assertEqual(bienes.count(), 6)
        
        # Order by descendente
        bienes = BienPatrimonial.objects.order_by('-codigo_patrimonial')
        self.assertEqual(bienes.count(), 6)
    
    def test_values_compatibility(self):
        """Prueba compatibilidad de values()"""
        # Values simple
        values = BienPatrimonial.objects.values('codigo_patrimonial')
        self.assertEqual(len(values), 1)
        
        # Values con múltiples campos
        values = BienPatrimonial.objects.values('codigo_patrimonial', 'estado_bien')
        self.assertEqual(len(values), 1)
    
    def test_values_list_compatibility(self):
        """Prueba compatibilidad de values_list()"""
        # Values list simple
        codes = BienPatrimonial.objects.values_list('codigo_patrimonial', flat=True)
        self.assertEqual(len(codes), 1)
        self.assertIn('PAT-QS-001', codes)
        
        # Values list con múltiples campos
        data = BienPatrimonial.objects.values_list('codigo_patrimonial', 'estado_bien')
        self.assertEqual(len(data), 1)
    
    def test_select_related_compatibility(self):
        """Prueba compatibilidad de select_related()"""
        # Select related con catalogo
        bienes = BienPatrimonial.objects.select_related('catalogo').all()
        self.assertEqual(bienes.count(), 1)
        
        # Acceder a relación no debería generar consultas adicionales
        for bien in bienes:
            _ = bien.catalogo.denominacion
    
    def test_prefetch_related_compatibility(self):
        """Prueba compatibilidad de prefetch_related()"""
        # Prefetch related
        bienes = BienPatrimonial.objects.prefetch_related('oficina').all()
        self.assertEqual(bienes.count(), 1)
        
        # Acceder a relación no debería generar consultas adicionales
        for bien in bienes:
            _ = bien.oficina.nombre
    
    def test_aggregate_compatibility(self):
        """Prueba compatibilidad de aggregate()"""
        from django.db.models import Count, Q
        
        # Aggregate simple
        result = BienPatrimonial.objects.aggregate(total=Count('id'))
        self.assertEqual(result['total'], 1)
        
        # Aggregate con filtros condicionales
        result = BienPatrimonial.objects.aggregate(
            total=Count('id'),
            buenos=Count('id', filter=Q(estado_bien='B'))
        )
        self.assertEqual(result['total'], 1)
        self.assertEqual(result['buenos'], 1)
    
    def test_annotate_compatibility(self):
        """Prueba compatibilidad de annotate()"""
        from django.db.models import Count
        
        # Annotate
        oficinas = Oficina.objects.annotate(
            num_bienes=Count('bienpatrimonial')
        )
        
        # Debería contar solo bienes activos
        for oficina in oficinas:
            if oficina.codigo == 'DIR-001':
                self.assertEqual(oficina.num_bienes, 1)


class AdminInterfaceCompatibilityTest(TestCase):
    """Pruebas de compatibilidad de interfaz de administración"""
    
    def setUp(self):
        """Configurar datos de prueba"""
        self.admin_user = User.objects.create_user(
            username='admin',
            password='admin123',
            is_staff=True,
            is_superuser=True
        )
        
        self.catalogo = Catalogo.objects.create(
            codigo='04220001',
            denominacion='TEST BIEN',
            grupo='04-AGRÍCOLA Y PESQUERO',
            clase='22-EQUIPO',
            resolucion='R.D. 001-2024',
            estado='ACTIVO'
        )
        
        self.oficina = Oficina.objects.create(
            codigo='DIR-001',
            nombre='Dirección Regional',
            responsable='Director Regional'
        )
        
        self.client = Client()
        self.client.login(username='admin', password='admin123')
    
    def test_admin_list_view_compatibility(self):
        """Prueba compatibilidad de vista de lista en admin"""
        # Crear bien
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-ADMIN-001',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # Acceder a lista en admin
        response = self.client.get('/admin/bienes/bienpatrimonial/')
        self.assertEqual(response.status_code, 200)
        
        # Debería mostrar el bien
        self.assertContains(response, 'PAT-ADMIN-001')
    
    def test_admin_change_view_compatibility(self):
        """Prueba compatibilidad de vista de cambio en admin"""
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-ADMIN-002',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # Acceder a vista de cambio
        response = self.client.get(f'/admin/bienes/bienpatrimonial/{bien.id}/change/')
        self.assertEqual(response.status_code, 200)
    
    def test_admin_delete_view_compatibility(self):
        """Prueba compatibilidad de vista de eliminación en admin"""
        bien = BienPatrimonial.objects.create(
            codigo_patrimonial='PAT-ADMIN-003',
            catalogo=self.catalogo,
            oficina=self.oficina,
            estado_bien='B',
            created_by=self.admin_user
        )
        
        # Acceder a vista de eliminación
        response = self.client.get(f'/admin/bienes/bienpatrimonial/{bien.id}/delete/')
        self.assertEqual(response.status_code, 200)
        
        # Eliminar desde admin (ahora soft delete)
        response = self.client.post(
            f'/admin/bienes/bienpatrimonial/{bien.id}/delete/',
            {'post': 'yes'}
        )
        
        # Verificar que se eliminó lógicamente
        bien.refresh_from_db()
        self.assertTrue(bien.is_deleted)
