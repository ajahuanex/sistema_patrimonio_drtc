#!/usr/bin/env python
"""
Script para probar la funcionalidad de preview
"""
import os
import sys
import django
import openpyxl

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'patrimonio.settings')
django.setup()

from apps.oficinas.utils import generar_preview_oficinas

def crear_archivo_prueba():
    """Crea un archivo Excel de prueba"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Encabezados con acentos (como los exportamos)
    headers = ['Código', 'Nombre', 'Responsable', 'Teléfono', 'Email']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de prueba
    datos = [
        ['ADM-001', 'Administración', 'Juan Pérez', '051-123456', 'admin@test.com'],
        ['FIN-001', 'Finanzas', 'María García', '051-654321', 'finanzas@test.com']
    ]
    
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    filename = 'test_preview.xlsx'
    wb.save(filename)
    return filename

def probar_preview():
    """Prueba la funcionalidad de preview"""
    print("🧪 Probando funcionalidad de preview...")
    
    # Crear archivo de prueba
    archivo = crear_archivo_prueba()
    
    try:
        # Probar preview
        resultado = generar_preview_oficinas(archivo, max_filas=5)
        
        print(f"✅ Éxito: {resultado.get('exito', False)}")
        
        if resultado.get('exito'):
            print(f"📊 Fila encabezados: {resultado.get('fila_encabezados')}")
            print(f"📊 Total filas: {resultado.get('total_filas_datos')}")
            print(f"📊 Columnas detectadas: {list(resultado.get('columnas_detectadas', {}).keys())}")
            print(f"📊 Preview data: {len(resultado.get('preview_data', []))} filas")
            
            # Mostrar mapeo de columnas
            print("\n🔍 Mapeo de columnas:")
            for col, header in resultado.get('columnas_detectadas', {}).items():
                print(f"   '{header}' → {col}")
            
            # Mostrar datos de preview
            print("\n📋 Datos de preview:")
            for i, fila in enumerate(resultado.get('preview_data', [])[:3]):
                print(f"   Fila {fila.get('_fila_numero', i+1)}: {fila}")
        else:
            print(f"❌ Errores: {resultado.get('errores', [])}")
        
        return resultado.get('exito', False)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return False
    
    finally:
        # Limpiar archivo
        try:
            os.remove(archivo)
        except:
            pass

if __name__ == '__main__':
    success = probar_preview()
    print(f"\n{'🎉 Preview funcionando!' if success else '❌ Preview con problemas'}")